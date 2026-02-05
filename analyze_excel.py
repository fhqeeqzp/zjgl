#!/usr/bin/env python3
"""分析Excel文件结构"""
import openpyxl
from openpyxl import load_workbook

# 读取原始文件
wb1 = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
print('=== 原始文件工作簿列表 ===')
for name in wb1.sheetnames:
    print(f'  - {name}')

# 读取正确识别示例
wb2 = load_workbook('WD/陈台沟填充站项目/投标附件/土建正确识别.xlsx', data_only=True)
print('\n=== 正确识别文件工作簿列表 ===')
for name in wb2.sheetnames:
    print(f'  - {name}')

# 分析原始文件中的辅助斜坡道空气预热间土建工作簿
sheet = wb1['1.辅助斜坡道空气预热间土建']
print('\n=== 原始数据前50行 ===')
for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True), 1):
    row_str = ' | '.join([str(cell)[:25] if cell else '' for cell in row[:6]])
    print(f'{i:3d}: {row_str}')

# 分析关键行
print('\n=== 关键行分析 ===')
print('第19行（被截断的项目名称）:')
row19 = list(sheet.iter_rows(min_row=19, max_row=19, values_only=True))[0]
print(f'  项目名称: {row19[3]}')
print(f'  项目特征: {row19[5]}')

print('\n第28行（续行）:')
row28 = list(sheet.iter_rows(min_row=28, max_row=28, values_only=True))[0]
print(f'  项目名称: {row28[3]}')
print(f'  项目特征: {row28[5]}')

print('\n第7行（分部名称）:')
row7 = list(sheet.iter_rows(min_row=7, max_row=7, values_only=True))[0]
print(f'  内容: {row7}')

print('\n第16行（分部名称）:')
row16 = list(sheet.iter_rows(min_row=16, max_row=16, values_only=True))[0]
print(f'  内容: {row16}')

# 查看正确识别文件
sheet2 = wb2['1土建']
print('\n=== 正确识别文件数据 ===')
for i, row in enumerate(sheet2.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    row_str = ' | '.join([str(cell)[:25] if cell else '' for cell in row[:6]])
    print(f'{i:3d}: {row_str}')

wb1.close()
wb2.close()
