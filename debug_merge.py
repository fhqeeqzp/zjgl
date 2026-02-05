#!/usr/bin/env python3
"""调试合并逻辑"""
from openpyxl import load_workbook
import re
from typing import Dict

def _should_merge_with_previous(current_row: tuple, previous_item: Dict, mapping: Dict[str, int]) -> bool:
    if not previous_item:
        return False
    
    name_col = mapping.get('name')
    desc_col = mapping.get('description')
    sequence_col = mapping.get('sequence')
    code_col = mapping.get('code')
    
    current_name = str(current_row[name_col]).strip() if name_col is not None and name_col < len(current_row) and current_row[name_col] else ""
    current_desc = str(current_row[desc_col]).strip() if desc_col is not None and desc_col < len(current_row) and current_row[desc_col] else ""
    current_seq = str(current_row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(current_row) and current_row[sequence_col] else ""
    current_code = str(current_row[code_col]).strip() if code_col is not None and code_col < len(current_row) and current_row[code_col] else ""
    
    prev_name = str(previous_item.get('name', '')).strip()
    prev_desc = str(previous_item.get('description', '')).strip()
    
    print(f"  当前行: seq='{current_seq}', code='{current_code}', name='{current_name[:20]}...'")
    print(f"  前一行: seq='{prev_name[:20]}...', desc='{prev_desc[:30]}...'")
    
    # 检查是否有项目编码
    if current_code and re.match(r'^\d{9,}$', current_code.replace('-', '')):
        print(f"  → 有项目编码，不合并")
        return False
    
    # 检查是否有序号
    if current_seq and re.match(r'^\d+$', current_seq):
        print(f"  → 有序号，不合并")
        return False
    
    # 如果没有序号，但有项目名称
    if not current_seq and current_name and not current_code:
        print(f"  → 无序号有名称，合并")
        return True
    
    # 如果当前行项目名称非常短
    if len(current_name) < 3 and len(prev_name) > 0:
        print(f"  → 名称太短，合并")
        return True
    
    # 如果前一行项目名称以不完整的方式结束
    incomplete_endings_name = ['混凝', '预拌', '泵送', '强度', '等级', '基础', '独立', '类型', '种类', '形式']
    for ending in incomplete_endings_name:
        if prev_name.endswith(ending):
            print(f"  → 前一行名称以'{ending}'结尾，合并")
            return True
    
    # 如果前一行项目特征描述以不完整的方式结束
    incomplete_endings_desc = ['混凝', '预拌', '泵送', '强度等级:', '基础类型:', '混凝土种类:', '种类:', '等级:', '类型:']
    for ending in incomplete_endings_desc:
        if prev_desc.endswith(ending):
            print(f"  → 前一行描述以'{ending}'结尾，合并")
            return True
    
    print(f"  → 不合并")
    return False

# 加载Excel文件
wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

mapping = {
    'sequence': 0,
    'code': 1,
    'name': 3,
    'description': 5,
}

# 读取第19-20行（Excel中的行号）
print("=== 调试第19-20行 ===")
row19 = list(sheet.iter_rows(min_row=19, max_row=19, values_only=True))[0]
row20 = list(sheet.iter_rows(min_row=20, max_row=20, values_only=True))[0]

print(f"第19行: {row19}")
print(f"第20行: {row20}")

previous_item = {
    'sequence': str(row19[0]) if row19[0] else '',
    'code': str(row19[1]) if row19[1] else '',
    'name': str(row19[3]) if row19[3] else '',
    'description': str(row19[5]) if row19[5] else '',
}

print(f"\n前一行数据: {previous_item}")
print(f"\n检查第20行是否应该合并:")
should_merge = _should_merge_with_previous(row20, previous_item, mapping)
print(f"\n结果: {'合并' if should_merge else '不合并'}")

wb.close()
