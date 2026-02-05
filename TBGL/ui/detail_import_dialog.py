"""
投标明细Excel导入对话框
用于从Excel文件导入明细数据，支持工作簿选择、列映射、数据预览
字段与汇总表类似，但针对明细数据优化
"""
import os
import re
from pathlib import Path
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PySide6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QComboBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QPushButton,
    QFrame,
    QSplitter,
    QWidget,
    QProgressBar,
    QTextEdit,
    QApplication,
    QScrollArea,
    QGridLayout,
    QFormLayout
)

from ui.message_dialog import MessageDialog
from PySide6.QtCore import Qt, Signal, QThread
from PySide6.QtGui import QFont

from ui.fluent_widgets import PushButton, PrimaryPushButton
from .excel_import_dialog import ColumnMappingWidget, ExcelLoadWorker


class ImportProgressDialog(QDialog):
    """导入进度对话框 - 显示导入进度和明细"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("正在导入数据...")
        self.setMinimumSize(800, 600)
        self.setMaximumSize(800, 600)
        self.setModal(True)
        
        # 设置对象名以便QSS样式应用
        self.setObjectName("importProgressDialog")
        
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 状态标签
        self.status_label = QLabel("正在准备导入...")
        self.status_label.setObjectName("statusLabel")
        layout.addWidget(self.status_label)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("%p% (%v/%m)")
        layout.addWidget(self.progress_bar)
        
        # 当前行信息
        self.current_row_label = QLabel("")
        self.current_row_label.setObjectName("currentRowLabel")
        layout.addWidget(self.current_row_label)
        
        # 明细显示区域
        detail_label = QLabel("📋 导入明细（最近20行）：")
        detail_label.setObjectName("detailLabel")
        layout.addWidget(detail_label)
        
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        layout.addWidget(self.detail_text)
        
        # 统计信息
        self.stats_label = QLabel("已导入: 0 行")
        self.stats_label.setObjectName("statsLabel")
        layout.addWidget(self.stats_label)
        
        # 取消按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.cancel_btn = PushButton("取消导入")
        self.cancel_btn.setFixedSize(100, 35)
        self.cancel_btn.clicked.connect(self.on_cancel)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
        
        self.is_cancelled = False
        self.imported_count = 0
        self.error_count = 0
    
    def set_total_rows(self, total: int):
        """设置总行数"""
        self.progress_bar.setMaximum(total)
        self.total_rows = total
    
    def update_progress(self, current: int, row_data: dict = None):
        """更新进度"""
        self.progress_bar.setValue(current)
        
        percentage = (current / self.total_rows * 100) if self.total_rows > 0 else 0
        self.status_label.setText(f"正在导入... {percentage:.1f}%")
        
        if row_data:
            sequence = row_data.get('sequence', '')
            name = row_data.get('name', '')
            self.current_row_label.setText(f"当前行: {current}/{self.total_rows} - {sequence} {name[:30]}")
            self.add_detail_line(current, row_data)
        
        self.stats_label.setText(f"已导入: {self.imported_count} 行 | 错误: {self.error_count} 行")
        QApplication.processEvents()
    
    def add_detail_line(self, row_num: int, row_data: dict):
        """添加明细行"""
        sequence = row_data.get('sequence', '')
        name = row_data.get('name', '')
        level = row_data.get('level', 1)
        
        indent = "  " * (level - 1)
        line = f"[{row_num:4d}] {indent}{sequence:8s} {name[:40]}"
        
        self.detail_text.append(line)
        scrollbar = self.detail_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def on_cancel(self):
        """取消导入"""
        self.is_cancelled = True
        self.status_label.setText("正在取消...")
        self.status_label.setProperty("status", "error")
        self.status_label.style().unpolish(self.status_label)
        self.status_label.style().polish(self.status_label)

    def set_completed(self, success_count: int, error_count: int):
        """设置完成状态"""
        self.progress_bar.setValue(self.progress_bar.maximum())
        self.status_label.setText("✓ 导入完成")
        self.status_label.setProperty("status", "success")
        self.status_label.style().unpolish(self.status_label)
        self.status_label.style().polish(self.status_label)
        self.stats_label.setText(f"成功: {success_count} 行 | 错误: {error_count} 行")
        self.cancel_btn.setText("关闭")
        self.cancel_btn.setEnabled(True)

    def set_error(self, error_msg: str):
        """设置错误状态"""
        self.status_label.setText(f"✗ 导入失败")
        self.status_label.setProperty("status", "error")
        self.status_label.style().unpolish(self.status_label)
        self.status_label.style().polish(self.status_label)
        self.detail_text.append(f"\n错误: {error_msg}")
        self.cancel_btn.setText("关闭")


class DetailImportDialog(QDialog):
    """投标明细导入对话框"""
    
    # 字段定义: (字段标识, 字段显示名称, 是否必填, 匹配关键词列表)
    FIELD_DEFINITIONS = [
        ('sequence', '序号', False, [
            '序号', '编号', 'no', 'no.', 'number', 'id'
        ]),
        ('name', '分部分项工程名称', True, [
            '分部分项工程名称', '工程名称', '项目名称', '名称', '项目', '工程', 'name', 'item'
        ]),
        ('specification', '规格型号', False, [
            '规格型号', '规格', '型号', 'spec', 'specification', 'type', 'model'
        ]),
        ('description', '项目特征描述', False, [
            '项目特征描述', '特征描述', '项目特征', '描述', '特征', 'description',
            '工作内容', '内容', '工作', 'work content', 'content'
        ]),
        ('unit', '单位', False, [
            '单位', '计量单位', 'unit', '计量'
        ]),
        ('quantity', '工程量', False, [
            '工程量', '数量', 'quantity', 'amount', '工程数量'
        ]),
        ('unit_price', '综合单价', False, [
            '综合单价', '单价', 'unit price', 'price', '单价(元)'
        ]),
        ('labor_unit_price', '人工单价', False, [
            '人工单价', '人工费单价', '人工', 'labor price', 'labor unit price'
        ]),
        ('material_unit_price', '主材单价', False, [
            '主材单价', '材料单价', '材料费单价'
        ]),
        ('material_loss_rate', '主材损耗%', False, [
            '主材损耗', '材料损耗', '损耗率', '损耗%', '损耗'
        ]),
        ('auxiliary_unit_price', '辅材单价', False, [
            '辅材单价', '辅助材料单价', '辅材', 'auxiliary price'
        ]),
        ('machine_unit_price', '机械单价', False, [
            '机械单价', '机械费单价', '机械台班单价', '机械', 'machine price'
        ]),
        ('other_unit_price', '其他单价', False, [
            '其他单价', '其他费单价', '其他', 'other price'
        ]),
        # 合价由计算得出，不进行导入识别
        # ('total_price', '合价', False, [
        #     '合价', '总价', '金额', 'total', '合价(元)', '金额(元)'
        # ]),
        ('remark', '备注', False, [
            '备注', '说明', 'note', 'remark', '注释'
        ]),
    ]
    
    def __init__(self, excel_path: str, bidding_name: str = "", summary_name: str = "", parent=None):
        super().__init__(parent)
        self.excel_path = excel_path
        self.bidding_name = bidding_name
        self.summary_name = summary_name  # 汇总项名称（工程项目及费用名称）
        self.sheet_names = []
        self.headers = []
        self.sample_data = []
        self.preview_rows = []
        self.header_row_index = 0
        self.header_row_count = 1  # 表头行数（支持多行表头）
        self.load_worker = None
        self.current_sheet_name = None
        self.drag_pos = None

        # 设置无边框窗口
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)

        self.setMinimumSize(1000, 800)

        self.setup_ui()
        self.load_excel_async()

    def setup_ui(self):
        """设置UI - 自定义标题栏"""
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # 创建内容容器
        self.content_frame = QFrame()
        self.content_frame.setObjectName("contentFrame")
        layout = QVBoxLayout(self.content_frame)
        layout.setSpacing(15)
        layout.setContentsMargins(0, 0, 0, 20)

        # ========== 自定义标题栏 ==========
        title_bar = QFrame()
        title_bar.setObjectName("titleBarFrame")
        title_bar.setFixedHeight(50)
        title_bar_layout = QHBoxLayout(title_bar)
        title_bar_layout.setContentsMargins(20, 0, 20, 0)
        title_bar_layout.setSpacing(10)

        # 标题
        title = QLabel("投标明细导入 - 列映射配置")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Bold))
        title.setObjectName("titleLabel")
        title_bar_layout.addWidget(title)

        title_bar_layout.addStretch()

        # 关闭按钮
        close_btn = QPushButton("×")
        close_btn.setObjectName("closeButton")
        close_btn.setProperty("class", "titlebar-button")
        close_btn.setFixedSize(46, 32)
        close_btn.setFont(QFont("Microsoft YaHei", 14))
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.clicked.connect(self.reject)
        title_bar_layout.addWidget(close_btn)

        layout.addWidget(title_bar)

        # ========== 表单内容区域 ==========
        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(10)
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        form_layout.setContentsMargins(30, 10, 30, 0)

        # 文件信息
        info_frame = QFrame()
        info_frame.setObjectName("infoFrame")
        info_layout = QHBoxLayout(info_frame)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_label = QLabel(f"📄 文件: {os.path.basename(self.excel_path)}")
        info_layout.addWidget(info_label)
        info_layout.addStretch()
        form_layout.addRow("文件:", info_frame)

        # 工作簿选择
        sheet_layout = QHBoxLayout()
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(180)
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)
        sheet_layout.addWidget(self.sheet_combo)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(100)
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.hide()
        sheet_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("")
        sheet_layout.addWidget(self.status_label)

        sheet_layout.addStretch()
        form_layout.addRow("工作簿:", sheet_layout)

        # 表头行选择
        header_row_layout = QHBoxLayout()
        self.header_row_combo = QComboBox()
        self.header_row_combo.setMinimumWidth(250)
        self.header_row_combo.setToolTip("选择哪一行作为表头")
        self.header_row_combo.currentIndexChanged.connect(self.on_header_row_changed)
        header_row_layout.addWidget(self.header_row_combo)

        header_row_hint = QLabel("💡 提示: 请选择包含列标题的行")
        header_row_hint.setObjectName("hintLabel")
        header_row_layout.addWidget(header_row_hint)

        header_row_layout.addStretch()
        form_layout.addRow("表头行:", header_row_layout)

        layout.addWidget(form_widget)

        # 分割器
        splitter = QSplitter(Qt.Vertical)

        # 上半部分：列映射（带滚动条）
        mapping_widget = QWidget()
        mapping_layout = QVBoxLayout(mapping_widget)
        mapping_layout.setContentsMargins(0, 0, 0, 0)

        # 创建滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setObjectName("mappingScrollArea")

        # 列映射容器（两列布局）
        mapping_container = QWidget()
        mapping_container.setObjectName("mappingContainer")
        mapping_inner_layout = QGridLayout(mapping_container)
        mapping_inner_layout.setContentsMargins(15, 15, 15, 15)
        mapping_inner_layout.setSpacing(10)
        mapping_inner_layout.setColumnStretch(0, 1)
        mapping_inner_layout.setColumnStretch(1, 1)

        # 创建列映射组件（两列布局）
        self.mapping_widgets = {}
        row = 0
        col = 0
        for field_key, field_name, is_required, _ in self.FIELD_DEFINITIONS:
            display_name = f"* {field_name}" if is_required else field_name
            widget = ColumnMappingWidget(field_key, display_name)
            widget.mapping_changed.connect(self.on_mapping_changed)
            self.mapping_widgets[field_key] = widget
            mapping_inner_layout.addWidget(widget, row, col)

            col += 1
            if col >= 2:  # 每行2个
                col = 0
                row += 1

        scroll_area.setWidget(mapping_container)
        mapping_layout.addWidget(scroll_area)

        # 按钮区域（固定居中）
        btn_frame = QFrame()
        btn_frame.setObjectName("buttonFrame")
        btn_layout = QHBoxLayout(btn_frame)
        btn_layout.setAlignment(Qt.AlignCenter)
        btn_layout.setSpacing(20)

        auto_btn = PushButton("🔍 智能识别列")
        auto_btn.setToolTip("根据表头名称自动匹配列")
        auto_btn.setFixedSize(120, 35)
        auto_btn.clicked.connect(self.auto_detect_columns)
        btn_layout.addWidget(auto_btn)

        clear_btn = PushButton("🗑️ 清除选择")
        clear_btn.setFixedSize(120, 35)
        clear_btn.clicked.connect(self.clear_mappings)
        btn_layout.addWidget(clear_btn)

        mapping_layout.addWidget(btn_frame)
        splitter.addWidget(mapping_widget)

        # 下半部分：原始数据预览（增加高度）
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)
        preview_layout.setContentsMargins(0, 0, 0, 0)

        self.preview_table = QTableWidget()
        self.preview_table.setMinimumHeight(250)

        preview_layout.addWidget(self.preview_table)
        splitter.addWidget(preview_widget)
        
        splitter.setSizes([350, 400])
        layout.addWidget(splitter)

        layout.addStretch()

        # 按钮区域（居中）
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        cancel_btn = PushButton("取消")
        cancel_btn.setObjectName("secondaryButton")
        cancel_btn.setFixedSize(100, 35)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        btn_layout.addSpacing(20)

        self.import_btn = PrimaryPushButton("确认导入")
        self.import_btn.setFixedSize(100, 35)
        self.import_btn.clicked.connect(self.on_import)
        btn_layout.addWidget(self.import_btn)

        btn_layout.addStretch()

        layout.addLayout(btn_layout)

        # 添加内容框架到主布局
        main_layout.addWidget(self.content_frame)
    
    def load_excel_async(self):
        """异步加载Excel文件"""
        if not OPENPYXL_AVAILABLE:
            MessageDialog.warning(self, "提示", "请先安装openpyxl: pip install openpyxl")
            self.reject()
            return
        
        self.progress_bar.show()
        self.status_label.setText("正在加载Excel文件...")
        self.import_btn.setEnabled(False)
        
        self.load_worker = ExcelLoadWorker(self.excel_path)
        self.load_worker.loaded.connect(self.on_excel_loaded)
        self.load_worker.sheet_data_loaded.connect(self.on_sheet_data_loaded)
        self.load_worker.error.connect(self.on_load_error)
        self.load_worker.start()
    
    def on_excel_loaded(self, sheet_names: list):
        """Excel文件加载完成回调 - 智能选择工作簿"""
        self.sheet_names = sheet_names
        
        self.sheet_combo.clear()
        for name in sheet_names:
            self.sheet_combo.addItem(name)
        
        if sheet_names:
            # 智能选择工作簿：优先选择包含"清单"、"明细"、"分部分项"等关键词的工作簿
            preferred_keywords = ['清单', '明细', '分部分项', '工程', '项目', 'data', 'sheet']
            selected_idx = 0
            best_score = 0
            
            for idx, name in enumerate(sheet_names):
                name_lower = name.lower()
                score = 0
                for keyword in preferred_keywords:
                    if keyword in name_lower:
                        score += 1
                # 默认工作簿名称（如Sheet1）得分较低
                if 'sheet' in name_lower and any(c.isdigit() for c in name):
                    score -= 0.5
                    
                if score > best_score:
                    best_score = score
                    selected_idx = idx
            
            self.current_sheet_name = sheet_names[selected_idx]
            self.sheet_combo.setCurrentIndex(selected_idx)
            # print(f"[工作簿选择] 自动选择: {self.current_sheet_name} (得分: {best_score})")
            self.load_worker.set_sheet(self.current_sheet_name)

    def on_sheet_data_loaded(self, preview_rows: list, _, row_count: int):
        """工作表数据加载完成回调"""
        # print(f"[明细导入] 加载完成，共 {row_count} 行预览数据")

        self.preview_rows = preview_rows

        # 填充表头行选择下拉框
        self.header_row_combo.clear()
        for i, row in enumerate(preview_rows):
            preview_text = " | ".join(str(cell)[:10] for cell in row[:3] if cell)
            if len(preview_text) > 30:
                preview_text = preview_text[:30] + "..."
            self.header_row_combo.addItem(f"第 {i+1} 行: {preview_text}", i)

        # 自动识别表头行
        detected_row = self._detect_header_row(preview_rows)
        self.header_row_index = detected_row
        self.header_row_combo.setCurrentIndex(detected_row)

        self.progress_bar.hide()
        self.import_btn.setEnabled(True)

        self._update_header_and_data()
        
        # 更新状态标签，显示表头行数
        self.status_label.setText(f"✓ 已加载 {row_count} 行预览数据 | 表头: {self.header_row_count} 行")

        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None

    def on_header_row_changed(self, index):
        """表头行选择改变"""
        if index >= 0 and self.preview_rows:
            self.header_row_index = index
            self._update_header_and_data()

    def _update_header_and_data(self):
        """根据选择的表头行更新数据 - 支持多行表头合并"""
        if not self.preview_rows or self.header_row_index >= len(self.preview_rows):
            return

        # 检查是否需要合并多行表头
        # 如果当前行和下一行都有内容，可能是多行表头
        merged_headers = []
        current_row = self.preview_rows[self.header_row_index]
        
        # 尝试合并后续2-3行作为完整表头
        rows_to_merge = [current_row]
        for i in range(1, 3):  # 最多合并后续2行
            if self.header_row_index + i < len(self.preview_rows):
                next_row = self.preview_rows[self.header_row_index + i]
                # 如果下一行有非空单元格，可能是多行表头的一部分
                if any(cell and str(cell).strip() for cell in next_row):
                    rows_to_merge.append(next_row)
                else:
                    break
        
        # 合并多行表头
        max_cols = max(len(row) for row in rows_to_merge)
        for col_idx in range(max_cols):
            header_parts = []
            for row in rows_to_merge:
                if col_idx < len(row) and row[col_idx]:
                    cell_val = str(row[col_idx]).strip()
                    if cell_val and cell_val not in header_parts:
                        header_parts.append(cell_val)
            
            # 合并表头部分（如"人工费" + "单价" = "人工费单价"）
            if header_parts:
                merged_headers.append(''.join(header_parts))
            else:
                merged_headers.append('')
        
        self.headers = merged_headers
        self.header_row_count = len(rows_to_merge)  # 记录表头行数
        self.sample_data = self.preview_rows[self.header_row_index + len(rows_to_merge):]
        
        # print(f"[表头识别] 合并了 {len(rows_to_merge)} 行表头: {self.headers[:5]}...")

        self.update_mapping_widgets()
        self.update_raw_preview()
        self.auto_detect_columns()
    
    def on_load_error(self, error_msg: str):
        """加载错误回调"""
        self.progress_bar.hide()
        self.status_label.setText("加载失败")
        self.import_btn.setEnabled(True)
        MessageDialog.critical(self, "错误", f"无法加载Excel文件:\n{error_msg}")
        
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
    
    def load_sheet_async(self, sheet_name: str):
        """异步加载指定工作表"""
        if not sheet_name:
            return
        
        self.current_sheet_name = sheet_name
        
        self.progress_bar.show()
        self.status_label.setText(f"正在加载工作表: {sheet_name}...")
        self.import_btn.setEnabled(False)
        
        self.load_worker = ExcelLoadWorker(self.excel_path)
        self.load_worker.sheet_data_loaded.connect(self.on_sheet_data_loaded)
        self.load_worker.error.connect(self.on_load_error)
        self.load_worker.set_sheet(sheet_name)
        self.load_worker.start()
    
    def update_mapping_widgets(self):
        """更新列映射组件"""
        for widget in self.mapping_widgets.values():
            widget.set_headers(self.headers, self.sample_data)
    
    def update_raw_preview(self):
        """更新原始数据预览"""
        if not self.headers:
            return
        
        self.preview_table.clear()
        self.preview_table.setColumnCount(len(self.headers))
        self.preview_table.setHorizontalHeaderLabels(self.headers)
        
        row_count = min(5, len(self.sample_data))
        self.preview_table.setRowCount(row_count)
        
        for row_idx, row_data in enumerate(self.sample_data[:5]):
            for col_idx, value in enumerate(row_data):
                if value is None:
                    value = ""
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.preview_table.setItem(row_idx, col_idx, item)
        
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
    
    def _normalize_text(self, text: str) -> str:
        """标准化文本"""
        if not text:
            return ""
        text = str(text).strip()
        text = re.sub(r'[：:\(\)（）\[\]【】\s\.\-_]', '', text)
        return text.lower()

    def _detect_header_row(self, preview_rows: List[List[str]]) -> int:
        """自动识别表头行 - 增强版"""
        if not preview_rows:
            return 0

        # 收集所有字段的匹配关键词
        all_keywords = []
        for _, _, _, keywords in self.FIELD_DEFINITIONS:
            all_keywords.extend(keywords)
        
        # 添加额外的表头识别关键词
        header_indicators = ['序号', '编号', '名称', '项目', '单位', '数量', '单价', '合价', '金额', '备注']
        all_keywords.extend(header_indicators)

        best_row_idx = 0
        best_score = 0.0
        
        # print(f"[自动识别表头] 开始识别，共 {len(preview_rows)} 行预览数据")

        for row_idx, row in enumerate(preview_rows):
            score = 0.0
            matched_keywords = []

            for cell in row:
                cell_norm = self._normalize_text(cell)
                if not cell_norm:
                    continue

                for keyword in all_keywords:
                    keyword_norm = self._normalize_text(keyword)
                    if not keyword_norm:
                        continue

                    if cell_norm == keyword_norm:
                        score += 1.0
                        matched_keywords.append(f"{cell}={keyword}")
                    elif keyword_norm in cell_norm:
                        score += 0.8
                        matched_keywords.append(f"{cell}~{keyword}")
                    # 部分匹配（至少2个字符）
                    elif len(keyword_norm) >= 2:
                        for i in range(len(keyword_norm)):
                            for j in range(i + 2, min(i + 5, len(keyword_norm) + 1)):
                                substr = keyword_norm[i:j]
                                if substr in cell_norm:
                                    score += 0.3 * (len(substr) / len(keyword_norm))

            # 归一化分数
            normalized_score = score / max(len(row), 1)

            # print(f"[自动识别表头] 第 {row_idx + 1} 行: 得分={normalized_score:.2f}, 匹配={matched_keywords[:5]}")

            if normalized_score > best_score:
                best_score = normalized_score
                best_row_idx = row_idx

        # print(f"[自动识别表头] 结果: 第 {best_row_idx + 1} 行是表头 (得分: {best_score:.2f})")
        return best_row_idx

    def _calculate_similarity(self, header: str, keyword: str) -> float:
        """计算表头与关键词的相似度 - 增强版，支持多行表头"""
        header_norm = self._normalize_text(header)
        keyword_norm = self._normalize_text(keyword)
        
        if not header_norm or not keyword_norm:
            return 0.0
        
        # 完全匹配
        if header_norm == keyword_norm:
            return 1.0
        
        # 包含匹配
        if keyword_norm in header_norm:
            if header_norm.startswith(keyword_norm):
                return 0.9
            elif header_norm.endswith(keyword_norm):
                return 0.8
            else:
                return 0.7
        
        # 处理多行表头情况（如"人工费单价"可以匹配"人工费"+"单价"）
        # 将关键词拆分为部分
        keyword_parts = keyword_norm.split('费') if '费' in keyword_norm else [keyword_norm]
        if len(keyword_parts) > 1:
            # 检查是否包含关键词的各个部分
            matched_parts = 0
            for part in keyword_parts:
                if part and part in header_norm:
                    matched_parts += 1
            if matched_parts > 0:
                return 0.6 * (matched_parts / len(keyword_parts))
        
        # 部分匹配（共同子串）
        max_common = 0
        for i in range(len(keyword_norm)):
            for j in range(i + 1, len(keyword_norm) + 1):
                substr = keyword_norm[i:j]
                if substr in header_norm and len(substr) > max_common:
                    max_common = len(substr)
        
        if max_common > 0:
            return 0.5 * (max_common / max(len(header_norm), len(keyword_norm)))
        
        return 0.0
    
    def auto_detect_columns(self):
        """智能识别列 - 带排除词检查"""
        if not self.headers:
            return

        assigned_columns = set()

        # 定义字段的排除词（如果表头包含这些词，则不应匹配）
        exclusion_keywords = {
            'material_unit_price': ['损耗', '损', 'loss', '损耗率'],
            'material_loss_rate': [],  # 损耗列没有排除词
        }

        for field_key, field_name, is_required, keywords in self.FIELD_DEFINITIONS:
            best_match_idx = None
            best_score = 0.0

            for idx, header in enumerate(self.headers):
                if idx in assigned_columns:
                    continue

                # 检查排除词
                header_lower = header.lower()
                excluded = False
                if field_key in exclusion_keywords:
                    for exclude_word in exclusion_keywords[field_key]:
                        if exclude_word in header_lower:
                            excluded = True
                            # print(f"[排除匹配] {field_key}: '{header}' 包含排除词 '{exclude_word}'")
                            break

                if excluded:
                    continue

                for keyword in keywords:
                    score = self._calculate_similarity(header, keyword)

                    if score > best_score:
                        best_score = score
                        best_match_idx = idx

            threshold = 0.6 if is_required else 0.5

            if best_match_idx is not None and best_score >= threshold:
                self.mapping_widgets[field_key].set_selected_column(best_match_idx)
                assigned_columns.add(best_match_idx)
                # print(f"[匹配成功] {field_key} -> 列{best_match_idx} '{self.headers[best_match_idx]}' (得分: {best_score:.2f})")

        self.on_mapping_changed()
    
    def clear_mappings(self):
        """清除所有列映射"""
        for widget in self.mapping_widgets.values():
            widget.combo.setCurrentIndex(0)
    
    def on_sheet_changed(self, index):
        """工作簿改变"""
        if index >= 0 and self.sheet_names:
            sheet_name = self.sheet_names[index]
            if sheet_name != self.current_sheet_name:
                self.load_sheet_async(sheet_name)
    
    def on_mapping_changed(self):
        """列映射改变"""
        pass
    
    def get_column_mapping(self) -> Dict[str, int]:
        """获取列映射配置"""
        mapping = {}
        for field_key, widget in self.mapping_widgets.items():
            col_idx = widget.get_selected_column()
            if col_idx is not None:
                mapping[field_key] = col_idx
        return mapping
    
    def parse_sequence_level(self, sequence: str) -> int:
        """解析序号层级"""
        if not sequence:
            return 1
        
        sequence = str(sequence).strip()
        
        # 中文数字 -> 层级1
        chinese_nums = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
        if any(c in sequence for c in chinese_nums):
            return 1
        
        # 计算小数点数量
        dots = sequence.count('.')
        return dots + 2
    
    def _is_header_row(self, row: tuple, mapping: Dict[str, int]) -> bool:
        """
        智能检测是否为表头行或需要跳过的行
        
        检测策略：
        1. 检查是否为重复的表头行（"分部分项工程和单价措施项目清单与计价表"、"工程名称"等）
        2. 检查是否为表头列名行（"序号|项目编码|项目名称|..."）
        3. 检查是否为"本页小计"、"合计"等汇总行
        4. 检查是否为空行
        """
        # 表头标题行关键词（整行匹配）
        header_title_patterns = [
            '分部分项工程和单价措施项目清单与计价表',
            '分部分项工程',
            '工程名称：',
            '标段：',
        ]
        
        # 表头列名行关键词（多列联合判断）
        header_column_keywords = ['序号', '项目编码', '项目名称', '项目特征', '计量单位', '工程量', '综合单价', '合价']
        
        # 汇总行关键词（需要跳过）
        summary_keywords = ['本页小计', '合计', '总计', '小计', '共', '页', '表—']
        
        # 获取关键列的索引
        name_col = mapping.get('name')
        desc_col = mapping.get('description')
        sequence_col = mapping.get('sequence')
        code_col = mapping.get('code')  # 项目编码列
        
        # 检查项目名称列
        if name_col is not None and name_col < len(row):
            name_val = str(row[name_col]).strip() if row[name_col] else ""
            
            # 检查是否为汇总行
            for keyword in summary_keywords:
                if keyword in name_val:
                    return True
            
            # 检查是否为表头标题行
            for pattern in header_title_patterns:
                if pattern in name_val:
                    return True
            
            # 检查是否为表头列名行（项目名称列包含"项目名称"且长度较短）
            if '项目名称' in name_val and len(name_val) < 15:
                return True
        
        # 检查项目特征描述列
        if desc_col is not None and desc_col < len(row):
            desc_val = str(row[desc_col]).strip() if row[desc_col] else ""
            
            # 检查是否为汇总行
            for keyword in summary_keywords:
                if keyword in desc_val:
                    return True
            
            # 检查是否为表头列名行
            if '项目特征描述' in desc_val and len(desc_val) < 15:
                return True
        
        # 检查整行是否为表头列名行（多列联合判断）
        # 如果序号列、项目编码列、项目名称列同时包含表头关键词，则认为是表头行
        header_col_count = 0
        for col_idx in [sequence_col, code_col, name_col, desc_col]:
            if col_idx is not None and col_idx < len(row):
                cell_val = str(row[col_idx]).strip() if row[col_idx] else ""
                for keyword in header_column_keywords:
                    if keyword in cell_val:
                        header_col_count += 1
                        break
        
        # 如果超过2个关键列包含表头关键词，认为是表头行
        if header_col_count >= 2:
            return True
        
        # 检查序号列 - 如果序号列包含表头关键词
        if sequence_col is not None and sequence_col < len(row):
            seq_val = str(row[sequence_col]).strip() if row[sequence_col] else ""
            if seq_val == '序号' or '序号' in seq_val:
                return True
        
        return False
    
    def _is_division_row(self, row: tuple, mapping: Dict[str, int]) -> bool:
        """
        检测是否为分部行（如"土石方工程"、"混凝土工程"、"钢结构"、"装饰装修"、"门窗"）

        特征：
        1. 项目名称列有值
        2. 序号列为空
        3. 项目编码列为空
        4. 项目名称通常较短（少于15个字符）
        """
        name_col = mapping.get('name')
        sequence_col = mapping.get('sequence')
        code_col = mapping.get('code')
        desc_col = mapping.get('description')

        if name_col is None or name_col >= len(row):
            return False

        name_val = str(row[name_col]).strip() if row[name_col] else ""
        if not name_val:
            return False

        # 检查其他列是否为空
        seq_val = str(row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(row) and row[sequence_col] else ""
        code_val = str(row[code_col]).strip() if code_col is not None and code_col < len(row) and row[code_col] else ""
        desc_val = str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) and row[desc_col] else ""

        # 分部行的核心特征：有项目名称，但无序号和项目编码
        if seq_val or code_val:
            return False

        # 检查是否为工程类分部名称（包含关键词）
        division_keywords = ['工程', '分部', '分项', '措施项目']
        has_keyword = any(keyword in name_val for keyword in division_keywords)

        if has_keyword:
            return True

        # 对于不包含关键词的分部（如"钢结构"、"装饰装修"、"门窗"）
        # 判断标准：无序号、无编码、名称较短（少于15个字符）、无项目特征描述
        if len(name_val) < 15 and not desc_val:
            return True

        return False
    
    def _should_merge_with_previous(self, current_row: tuple, previous_item: Dict,
                                     mapping: Dict[str, int]) -> bool:
        """
        判断当前行是否应该与前一行合并 - 增强版

        合并策略（必须满足前一行不完整的前提）：
        1. 前一行项目名称以不完整的方式结束（如"混凝"而不是"混凝土"）
        2. 前一行项目特征描述以不完整的方式结束
        3. 当前行没有序号，且前一行不完整，可能是前一行的延续
        4. 当前行项目名称异常短（少于3个字符），且前一行不完整
        5. 当前行项目名称看起来像前一行的延续（以"土"、"送"等结尾词开头）
        6. 当前行项目特征描述看起来像前一行描述的延续

        重要：独立的分部行（如"钢结构"、"装饰装修"）不应该被合并
        """
        if not previous_item:
            return False

        name_col = mapping.get('name')
        desc_col = mapping.get('description')
        sequence_col = mapping.get('sequence')
        code_col = mapping.get('code')  # 项目编码列

        # 获取当前行的值
        current_name = str(current_row[name_col]).strip() if name_col is not None and name_col < len(current_row) and current_row[name_col] else ""
        current_desc = str(current_row[desc_col]).strip() if desc_col is not None and desc_col < len(current_row) and current_row[desc_col] else ""
        current_seq = str(current_row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(current_row) and current_row[sequence_col] else ""
        current_code = str(current_row[code_col]).strip() if code_col is not None and code_col < len(current_row) and current_row[code_col] else ""

        # 获取前一行的值
        prev_name = str(previous_item.get('name', '')).strip()
        prev_desc = str(previous_item.get('description', '')).strip()
        prev_seq = str(previous_item.get('sequence', '')).strip()

        # 如果当前行有项目编码，说明是新的一条记录，不合并
        if current_code and re.match(r'^\d{9,}$', current_code.replace('-', '')):
            return False

        # 如果当前行有序号且是数字格式，说明是新的一条记录，不合并
        if current_seq and re.match(r'^\d+$', current_seq):
            return False

        # 检查前一行是否以不完整的方式结束（这是合并的前提条件）
        prev_is_incomplete = False

        # 检查前一行项目名称是否以不完整的方式结束
        incomplete_endings_name = ['混凝', '预拌', '泵送', '强度', '等级', '基础', '独立', '类型', '种类', '形式']
        for ending in incomplete_endings_name:
            if prev_name.endswith(ending):
                prev_is_incomplete = True
                break

        # 检查前一行项目特征描述是否以不完整的方式结束
        if not prev_is_incomplete:
            incomplete_endings_desc = ['混凝', '预拌', '泵送', '强度等级:', '基础类型:', '混凝土种类:', '种类:', '等级:', '类型:']
            for ending in incomplete_endings_desc:
                if prev_desc.endswith(ending):
                    prev_is_incomplete = True
                    break

        # 如果前一行是完整的（如"钢结构"、"装饰装修"等分部行），则不合并
        if not prev_is_incomplete:
            return False

        # 如果当前行没有序号，但有项目名称，且前一行不完整，可能是前一行的延续
        if not current_seq and current_name and not current_code:
            return True

        # 如果当前行项目名称非常短（少于3个字符），且前一行不完整
        if len(current_name) < 3 and len(prev_name) > 0:
            return True

        # 如果当前行项目名称看起来像前一行的延续
        continuation_starts = ['土', '送', '级', '凝', '础', '类', '型']
        if current_name and current_name[0] in continuation_starts:
            # 进一步检查：当前行内容应该能补全前一行的结尾
            for ending in ['混凝', '预拌', '种类:', '类型:', '等级:']:
                if prev_name.endswith(ending) or prev_desc.endswith(ending):
                    return True

        # 如果当前行项目特征描述看起来像前一行描述的延续
        if current_desc and prev_desc:
            # 检查描述是否有明显的延续特征（如以"土"开头补充"混凝"）
            desc_continuations = [
                ('混凝', '土'),  # 混凝土
                ('预拌', '混凝'),  # 预拌混凝土
                ('种类:', '预拌'),  # 种类:预拌混凝土
                ('等级:', 'C'),  # 等级:C30
            ]
            for prev_ending, curr_start in desc_continuations:
                if prev_desc.endswith(prev_ending) and current_desc.startswith(curr_start):
                    return True

        return False
    
    def _merge_rows(self, previous_item: Dict, current_row: tuple, mapping: Dict[str, int]) -> Dict:
        """
        合并两行数据
        
        合并规则：
        1. 项目名称：拼接前一行和当前行
        2. 项目特征描述：拼接前一行和当前行
        3. 其他字段：优先使用非空值
        """
        merged = previous_item.copy()
        
        name_col = mapping.get('name')
        desc_col = mapping.get('description')
        
        # 合并项目名称
        if name_col is not None and name_col < len(current_row):
            current_name = str(current_row[name_col]).strip() if current_row[name_col] else ""
            if current_name:
                prev_name = str(previous_item.get('name', '')).strip()
                # 智能拼接：避免重复字符
                if prev_name and current_name:
                    # 检查是否有重复字符
                    overlap = 0
                    for i in range(1, min(len(prev_name), len(current_name)) + 1):
                        if prev_name[-i:] == current_name[:i]:
                            overlap = i
                    if overlap > 0:
                        merged['name'] = prev_name + current_name[overlap:]
                    else:
                        merged['name'] = prev_name + current_name
        
        # 合并项目特征描述
        if desc_col is not None and desc_col < len(current_row):
            current_desc = str(current_row[desc_col]).strip() if current_row[desc_col] else ""
            if current_desc:
                prev_desc = str(previous_item.get('description', '')).strip()
                # 智能拼接
                if prev_desc and current_desc:
                    overlap = 0
                    for i in range(1, min(len(prev_desc), len(current_desc)) + 1):
                        if prev_desc[-i:] == current_desc[:i]:
                            overlap = i
                    if overlap > 0:
                        merged['description'] = prev_desc + current_desc[overlap:]
                    else:
                        merged['description'] = prev_desc + current_desc
        
        # 更新数值字段（如果当前行有值则使用当前行的值）
        numeric_fields = ['quantity', 'unit_price', 'labor_unit_price',
                         'material_unit_price', 'material_loss_rate',
                         'auxiliary_unit_price', 'machine_unit_price', 'other_unit_price']
        
        for field_key in numeric_fields:
            col_idx = mapping.get(field_key)
            if col_idx is not None and col_idx < len(current_row):
                value = current_row[col_idx]
                if value is not None:
                    try:
                        num_value = float(value)
                        if num_value != 0:  # 优先使用非零值
                            merged[field_key] = num_value
                    except (ValueError, TypeError):
                        pass
        
        return merged
    
    def _validate_imported_data(self, items: List[Dict]) -> tuple:
        """
        验证导入数据的完整性和准确性
        
        返回: (是否通过验证, 错误信息列表, 警告信息列表)
        """
        errors = []
        warnings = []
        
        if not items:
            errors.append("没有导入任何数据")
            return False, errors, warnings
        
        # 检查必填字段
        required_fields = ['name']
        for idx, item in enumerate(items):
            for field in required_fields:
                if not item.get(field):
                    errors.append(f"第 {idx + 1} 行: {field} 字段为空")
        
        # 检查数据完整性
        for idx, item in enumerate(items):
            name = str(item.get('name', '')).strip()
            
            # 检查项目名称是否异常短（可能被错误截断）
            if len(name) < 3:
                warnings.append(f"第 {idx + 1} 行: 项目名称 '{name}' 异常短，可能被截断")
            
            # 检查项目特征描述是否包含不完整的句子
            desc = str(item.get('description', '')).strip()
            incomplete_markers = ['混凝', '预拌', '泵送', '强度等级:', '基础类型:']
            for marker in incomplete_markers:
                if desc.endswith(marker):
                    warnings.append(f"第 {idx + 1} 行: 项目特征描述可能不完整，以 '{marker}' 结尾")
            
            # 检查数值字段的合理性
            quantity = item.get('quantity', 0)
            unit_price = item.get('unit_price', 0)
            
            if quantity < 0:
                warnings.append(f"第 {idx + 1} 行: 工程量 {quantity} 为负数")
            if unit_price < 0:
                warnings.append(f"第 {idx + 1} 行: 单价 {unit_price} 为负数")
        
        # 检查是否有重复的项目名称
        name_counts = {}
        for item in items:
            name = str(item.get('name', '')).strip()
            if name:
                name_counts[name] = name_counts.get(name, 0) + 1
        
        for name, count in name_counts.items():
            if count > 1:
                warnings.append(f"项目名称 '{name}' 重复出现 {count} 次")
        
        # 检查层级结构是否合理
        levels = [item.get('level', 1) for item in items]
        if levels:
            max_level = max(levels)
            min_level = min(levels)
            if max_level - min_level > 5:
                warnings.append(f"层级跨度较大（{min_level} 到 {max_level}），请检查数据结构")
        
        return len(errors) == 0, errors, warnings
    
    def _clean_imported_data(self, items: List[Dict]) -> List[Dict]:
        """
        清理导入的数据
        
        清理规则：
        1. 去除项目名称和描述中的多余空格
        2. 标准化单位表示
        3. 修复常见的 OCR/导入错误
        """
        cleaned_items = []
        
        for item in items:
            cleaned = item.copy()
            
            # 清理项目名称
            if 'name' in cleaned:
                name = str(cleaned['name']).strip()
                # 去除多余空格
                name = re.sub(r'\s+', ' ', name)
                # 修复常见的拼接错误
                name = name.replace('混凝 土', '混凝土')
                name = name.replace('独立 基础', '独立基础')
                cleaned['name'] = name
            
            # 清理项目特征描述
            if 'description' in cleaned:
                desc = str(cleaned['description']).strip()
                # 去除多余空格和换行
                desc = re.sub(r'\s+', ' ', desc)
                desc = re.sub(r'\n\s*\n', '\n', desc)
                # 标准化编号格式
                desc = re.sub(r'(\d+)\s*[.．]\s*', r'\1.', desc)
                cleaned['description'] = desc
            
            # 标准化单位
            if 'unit' in cleaned:
                unit = str(cleaned['unit']).strip().lower()
                unit_mapping = {
                    'm3': 'm³',
                    'm2': 'm²',
                    'm': 'm',
                    'kg': 'kg',
                    't': 't',
                    '个': '个',
                    '套': '套',
                    '台': '台',
                    '组': '组',
                    '米': 'm',
                    '平方米': 'm²',
                    '立方米': 'm³',
                }
                if unit in unit_mapping:
                    cleaned['unit'] = unit_mapping[unit]
            
            cleaned_items.append(cleaned)
        
        return cleaned_items

    def import_data_with_progress(self, progress_dialog: ImportProgressDialog) -> List[Dict]:
        """导入数据（带进度显示）- 增强版，支持智能行识别和中间表头处理，支持用户手动标记行类型"""
        if not self.current_sheet_name or not self.excel_path:
            return []

        mapping = self.get_column_mapping()

        # 检查必填字段
        for field_key, field_name, is_required, _ in self.FIELD_DEFINITIONS:
            if is_required and field_key not in mapping:
                MessageDialog.warning(self, "提示", f"必须配置 {field_name} 列")
                return []

        # 获取用户手动标记的行类型（如果有）
        user_row_types = getattr(self, '_user_row_types', None)
        use_manual_types = user_row_types is not None and len(user_row_types) > 0

        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook[self.current_sheet_name]

            # 数据起始行 = 表头起始行 + 表头行数 + 1
            start_row = self.header_row_index + self.header_row_count + 1
            
            # 计算总行数
            total_rows = sheet.max_row - start_row + 1
            progress_dialog.set_total_rows(total_rows)

            imported_items = []
            error_count = 0
            first_row_name = None  # 存储第一行（顶级）的名称
            pending_item = None  # 待合并的临时项
            current_division = None  # 当前分部名称
            division_stack = []  # 分部层级栈
            last_data_row_idx = 0  # 最后一条数据行的索引
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=1):
                # 检查是否取消
                if progress_dialog.is_cancelled:
                    break
                
                # 计算实际Excel行号
                actual_row = start_row + row_idx - 1
                
                try:
                    # 如果使用用户手动标记的类型
                    if use_manual_types:
                        row_type = user_row_types.get(actual_row)
                        
                        # 无效行 - 跳过
                        if row_type and row_type.value == "无效行":
                            continue
                        
                        # 合并行 - 与前一行合并
                        if row_type and row_type.value == "合并行":
                            if pending_item:
                                pending_item = self._merge_rows(pending_item, row, mapping)
                                last_data_row_idx = row_idx
                            continue
                        
                        # 父级行/分部行
                        if row_type and row_type.value == "父级行":
                            name_col = mapping.get('name')
                            if name_col is not None and name_col < len(row) and row[name_col]:
                                division_name = str(row[name_col]).strip()
                                current_division = division_name
                                division_item = {
                                    'sequence': '',
                                    'name': division_name,
                                    'specification': '',
                                    'description': '',
                                    'unit': '',
                                    'quantity': 0.0,
                                    'unit_price': 0.0,
                                    'labor_unit_price': 0.0,
                                    'material_unit_price': 0.0,
                                    'material_loss_rate': 0.0,
                                    'auxiliary_unit_price': 0.0,
                                    'machine_unit_price': 0.0,
                                    'other_unit_price': 0.0,
                                    'total_price': 0.0,
                                    'level': 2,
                                    'remark': '',
                                    'is_division': True
                                }
                                if pending_item:
                                    imported_items.append(pending_item)
                                    progress_dialog.imported_count += 1
                                    pending_item = None
                                imported_items.append(division_item)
                                progress_dialog.imported_count += 1
                            continue
                        
                        # 清单行或未知类型 - 正常处理
                        # 继续执行下面的数据提取逻辑
                    else:
                        # 使用自动识别逻辑
                        # 检测是否为表头行或汇总行，如果是则跳过
                        if self._is_header_row(row, mapping):
                            continue

                        # 如果有待合并的项，先检查当前行是否应该与它合并
                        if pending_item and self._should_merge_with_previous(row, pending_item, mapping):
                            pending_item = self._merge_rows(pending_item, row, mapping)
                            last_data_row_idx = row_idx
                            continue

                        # 检测是否为分部行
                        if self._is_division_row(row, mapping):
                            name_col = mapping.get('name')
                            if name_col is not None and name_col < len(row) and row[name_col]:
                                division_name = str(row[name_col]).strip()
                                current_division = division_name
                                division_item = {
                                    'sequence': '',
                                    'name': division_name,
                                    'specification': '',
                                    'description': '',
                                    'unit': '',
                                    'quantity': 0.0,
                                    'unit_price': 0.0,
                                    'labor_unit_price': 0.0,
                                    'material_unit_price': 0.0,
                                    'material_loss_rate': 0.0,
                                    'auxiliary_unit_price': 0.0,
                                    'machine_unit_price': 0.0,
                                    'other_unit_price': 0.0,
                                    'total_price': 0.0,
                                    'level': 2,
                                    'remark': '',
                                    'is_division': True
                                }
                                if pending_item:
                                    imported_items.append(pending_item)
                                    progress_dialog.imported_count += 1
                                    pending_item = None
                                imported_items.append(division_item)
                                progress_dialog.imported_count += 1
                            continue
                    
                    # 提取当前行数据
                    item_data = {}
                    for field_key, col_idx in mapping.items():
                        if col_idx < len(row):
                            value = row[col_idx]

                            # 数值字段处理（合价由计算得出，不导入）
                            numeric_fields = ['quantity', 'unit_price', 'labor_unit_price',
                                            'material_unit_price', 'material_loss_rate',
                                            'auxiliary_unit_price',
                                            'machine_unit_price', 'other_unit_price']
                            if field_key in numeric_fields:
                                try:
                                    value = float(value) if value is not None else 0.0
                                except (ValueError, TypeError):
                                    value = 0.0
                            else:
                                value = str(value).strip() if value is not None else ""

                            item_data[field_key] = value
                        else:
                            # 数值字段默认值
                            numeric_fields = ['quantity', 'unit_price', 'labor_unit_price',
                                            'material_unit_price', 'material_loss_rate',
                                            'auxiliary_unit_price',
                                            'machine_unit_price', 'other_unit_price']
                            if field_key in numeric_fields:
                                item_data[field_key] = 0.0
                            else:
                                item_data[field_key] = ""

                    # 获取分部分项工程名称
                    name = str(item_data.get('name', '')).strip()
                    
                    # 过滤空行（分部分项工程名称为空的行）
                    if not name:
                        continue
                    
                    # 存储第一行（顶级）的名称
                    if first_row_name is None:
                        first_row_name = name
                    
                    # 计算层级
                    sequence = str(item_data.get('sequence', ''))
                    if sequence and sequence.strip():
                        base_level = self.parse_sequence_level(sequence)
                        # 如果有当前分部，则层级加2（顶级1 + 分部2）
                        if current_division:
                            item_data['level'] = base_level + 2
                        else:
                            item_data['level'] = base_level + 1
                    else:
                        # 如果没有序号，但有项目编码，视为明细项
                        code = str(item_data.get('code', '')).strip()
                        if code and re.match(r'^\d{9,}$', code.replace('-', '')):
                            item_data['level'] = 3 if current_division else 2
                        else:
                            item_data['level'] = 1
                    
                    # 检查是否需要与前一行合并（仅在非手动模式下）
                    if not use_manual_types and pending_item and self._should_merge_with_previous(row, pending_item, mapping):
                        # 合并数据
                        pending_item = self._merge_rows(pending_item, row, mapping)
                        last_data_row_idx = row_idx  # 更新最后数据行索引
                    else:
                        # 保存前一行（如果有）
                        if pending_item:
                            imported_items.append(pending_item)
                            progress_dialog.imported_count += 1
                        
                        # 开始新的项
                        pending_item = item_data
                        last_data_row_idx = row_idx  # 更新最后数据行索引
                    
                except Exception as e:
                    error_count += 1
                    progress_dialog.error_count += 1
                
                # 更新进度（每5行更新一次）
                if row_idx % 5 == 0 or row_idx == total_rows:
                    display_item = pending_item if pending_item else (imported_items[-1] if imported_items else None)
                    progress_dialog.update_progress(row_idx, display_item)

            # 处理最后一行
            if pending_item:
                imported_items.append(pending_item)
                progress_dialog.imported_count += 1

            workbook.close()
            
            # 使用汇总项名称作为顶级行名称
            if self.summary_name and imported_items:
                # 检查第一行是否已经是顶级（level=1）
                if imported_items[0].get('level', 1) == 1:
                    # 第一行已经是顶级，替换其名称
                    imported_items[0]['name'] = self.summary_name
                else:
                    # 需要创建一个顶级行
                    top_level_item = {
                        'sequence': '1',
                        'name': self.summary_name,
                        'specification': '',
                        'description': '',
                        'unit': '',
                        'quantity': 0.0,
                        'unit_price': 0.0,
                        'labor_unit_price': 0.0,
                        'material_unit_price': 0.0,
                        'material_loss_rate': 0.0,
                        'auxiliary_unit_price': 0.0,
                        'machine_unit_price': 0.0,
                        'other_unit_price': 0.0,
                        'total_price': 0.0,
                        'level': 1,
                        'remark': ''
                    }
                    # 将其他所有行的层级加1
                    for item in imported_items:
                        item['level'] = item.get('level', 1) + 1
                    # 在开头插入顶级行
                    imported_items.insert(0, top_level_item)
            
            # 清理导入的数据
            imported_items = self._clean_imported_data(imported_items)
            
            # 验证导入的数据
            is_valid, errors, warnings = self._validate_imported_data(imported_items)
            
            # 显示验证结果
            if warnings:
                warning_msg = "导入完成，但发现以下警告：\n\n"
                warning_msg += "\n".join([f"⚠ {w}" for w in warnings[:10]])
                if len(warnings) > 10:
                    warning_msg += f"\n... 还有 {len(warnings) - 10} 条警告"
                MessageDialog.warning(self, "数据验证警告", warning_msg)
            
            if not is_valid:
                error_msg = "导入数据存在错误：\n\n"
                error_msg += "\n".join([f"✗ {e}" for e in errors[:10]])
                if len(errors) > 10:
                    error_msg += f"\n... 还有 {len(errors) - 10} 条错误"
                MessageDialog.critical(self, "数据验证失败", error_msg)
                progress_dialog.set_error("数据验证失败")
                return []
            
            progress_dialog.set_completed(len(imported_items), error_count)
            return imported_items

        except Exception as e:
            progress_dialog.set_error(str(e))
            MessageDialog.critical(self, "错误", f"导入数据失败:\n{e}")
            return []
    
    def on_import(self):
        """确认导入"""
        # 创建进度对话框
        progress_dialog = ImportProgressDialog(self)
        
        # 显示进度对话框
        progress_dialog.show()
        QApplication.processEvents()
        
        # 执行导入
        data = self.import_data_with_progress(progress_dialog)
        
        # 存储导入的数据
        if data and len(data) > 0:
            self._imported_data = data
            # 导入完成，显示成功信息
            MessageDialog.information(self, "导入成功", f"成功导入 {len(data)} 条数据")
            progress_dialog.close()
            self.accept()
        elif data is not None and len(data) == 0:
            # 没有数据导入
            MessageDialog.warning(self, "提示", "没有导入任何数据，请检查Excel文件")
            progress_dialog.close()
        else:
            # 导入失败
            progress_dialog.close()
    
    def set_row_types(self, row_types: Dict[int, 'RowType']):
        """设置用户手动标记的行类型（从可视化导入对话框传入）"""
        self._user_row_types = row_types
    
    def get_imported_data(self) -> List[Dict]:
        """获取导入的数据"""
        return getattr(self, '_imported_data', [])
    
    def closeEvent(self, event):
        """关闭事件"""
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
        event.accept()

    # ========== 鼠标拖动支持 ==========
    def mousePressEvent(self, event):
        """鼠标按下 - 开始拖动"""
        if event.button() == Qt.LeftButton:
            self.drag_pos = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        """鼠标移动 - 拖动窗口"""
        if event.buttons() == Qt.LeftButton and self.drag_pos is not None:
            self.move(event.globalPos() - self.drag_pos)
            event.accept()
