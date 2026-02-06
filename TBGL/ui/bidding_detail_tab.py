"""
投标明细页签
显示投标明细信息，支持双击汇总表行后显示对应明细
使用树形表格展示层级结构，支持Excel导入
"""
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QMenu,
    QDialog,
    QCheckBox,
    QScrollArea,
    QGridLayout,
    QFrame,
    QApplication,
    QFileDialog
)
from PySide6.QtCore import QTimer

from ui.message_dialog import MessageDialog
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QAction

from ui.fluent_widgets import PushButton, PrimaryPushButton
from .visual_import_dialog import VisualImportDialog
from .excel_import_dialog import ExcelFileSelector

from ..logic.detail_manager import DetailManager
from ..data.detail_model import BiddingDetail, DetailItem as DetailItemModel


class DetailItem:
    """明细项数据类"""
    def __init__(self):
        self.id = 0
        self.parent_id = None
        self.sequence = ""           # 序号(ID字段，显示为序号)
        self.name = ""              # 分部分项工程名称
        self.specification = ""     # 规格型号
        self.description = ""       # 项目特征描述
        self.unit = ""              # 单位
        self.quantity = 0.0         # 工程量
        self.unit_price = 0.0       # 综合单价
        self.labor_unit_price = 0.0     # 人工单价
        self.material_unit_price = 0.0  # 主材单价
        self.material_loss_rate = 0.0   # 主材损耗%
        self.auxiliary_unit_price = 0.0 # 辅材单价
        self.machine_unit_price = 0.0   # 机械单价
        self.other_unit_price = 0.0     # 其他单价
        self.total_price = 0.0          # 合价(净)
        self.labor_total = 0.0          # 人工合价
        self.material_total = 0.0       # 主材合价
        self.auxiliary_total = 0.0      # 辅材合价
        self.machine_total = 0.0        # 机械合价
        self.other_total = 0.0          # 其他合价
        self.management_total = 0.0     # 管理费合计
        self.tax_total = 0.0            # 税金合价
        self.comprehensive_total = 0.0  # 综合合价
        self.remark = ""                # 备注
        self.level = 1                  # 层级
        
    def calculate_total(self):
        """计算合价"""
        self.total_price = self.quantity * self.unit_price
        self.labor_total = self.quantity * self.labor_unit_price
        self.material_total = self.quantity * self.material_unit_price
        self.auxiliary_total = self.quantity * self.auxiliary_unit_price
        self.machine_total = self.quantity * self.machine_unit_price
        self.other_total = self.quantity * self.other_unit_price
        self.comprehensive_total = self.total_price + self.management_total + self.tax_total
        return self.total_price


class ColumnSettingsDialog(QDialog):
    """列显示设置对话框"""
    
    # 所有可用的列定义: (列索引, 字段名, 显示名称, 默认是否显示)
    # 第0列是层级缩进列（始终显示，不在选择列表中）
    ALL_COLUMNS = [
        (0, 'indent', '', True),  # 层级缩进列（始终显示）
        (1, 'sequence', '序号', True),
        (2, 'name', '分部分项工程名称', True),
        (3, 'specification', '规格型号', True),
        (4, 'description', '项目特征描述', True),
        (5, 'unit', '单位', True),
        (6, 'quantity', '工程量', True),
        (7, 'unit_price', '综合单价', True),
        (8, 'labor_unit_price', '人工单价', True),
        (9, 'material_unit_price', '主材单价', True),
        (10, 'material_loss_rate', '主材损耗%', True),
        (11, 'auxiliary_unit_price', '辅材单价', True),
        (12, 'machine_unit_price', '机械单价', True),
        (13, 'other_unit_price', '其他单价', True),
        (14, 'total_price', '合价(净)', True),
        (15, 'labor_total', '人工合价', False),
        (16, 'material_total', '主材合价', False),
        (17, 'auxiliary_total', '辅材合价', False),
        (18, 'machine_total', '机械合价', False),
        (19, 'other_total', '其他合价', False),
        (20, 'management_total', '管理费合计', False),
        (21, 'tax_total', '税金合价', False),
        (22, 'comprehensive_total', '综合合价', False),
        (23, 'remark', '备注', False),
    ]
    
    def __init__(self, current_visible_columns=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("列显示设置")
        self.setMinimumSize(400, 500)
        self.current_visible_columns = current_visible_columns or [col[0] for col in self.ALL_COLUMNS if col[3]]
        self.checkboxes = {}
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 设置对话框样式
        self.setObjectName("columnSelectDialog")
        
        # 标题
        title_label = QLabel("请选择要显示的列：")
        title_label.setObjectName("titleLabel")
        layout.addWidget(title_label)

        # 说明文字
        hint_label = QLabel("💡 提示：分部分项工程名称、规格型号、项目特征描述会自动调整列宽显示全部内容")
        hint_label.setObjectName("hintLabel")
        layout.addWidget(hint_label)
        
        # 缩进列和序号列说明（始终显示）
        indent_hint = QLabel("ℹ️ 层级缩进列和序号列始终显示")
        indent_hint.setObjectName("indentHintLabel")
        layout.addWidget(indent_hint)
        
        # 滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # 复选框容器
        checkbox_container = QFrame()
        checkbox_layout = QGridLayout(checkbox_container)
        checkbox_layout.setSpacing(10)
        checkbox_layout.setContentsMargins(15, 15, 15, 15)
        
        # 创建复选框（跳过缩进列和序号列，因为它们始终显示）
        row = 0
        col = 0
        for col_idx, field_name, display_name, default_visible in self.ALL_COLUMNS:
            # 跳过缩进列（第0列）和序号列（第1列），因为它们始终显示
            if col_idx == 0 or col_idx == 1:
                continue
                
            checkbox = QCheckBox(display_name)
            checkbox.setChecked(col_idx in self.current_visible_columns)
            self.checkboxes[col_idx] = checkbox
            checkbox_layout.addWidget(checkbox, row, col)
            
            col += 1
            if col >= 2:  # 每行2个复选框
                col = 0
                row += 1
        
        scroll_area.setWidget(checkbox_container)
        layout.addWidget(scroll_area)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        # 全选/取消全选按钮
        select_all_btn = PushButton("全选")
        select_all_btn.clicked.connect(self.select_all)
        btn_layout.addWidget(select_all_btn)
        
        select_none_btn = PushButton("取消全选")
        select_none_btn.clicked.connect(self.select_none)
        btn_layout.addWidget(select_none_btn)
        
        # 恢复默认按钮
        default_btn = PushButton("恢复默认")
        default_btn.clicked.connect(self.restore_default)
        btn_layout.addWidget(default_btn)
        
        btn_layout.addSpacing(20)
        
        # 取消按钮
        cancel_btn = PushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        # 确定按钮
        ok_btn = PrimaryPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)
        
        layout.addLayout(btn_layout)
    
    def select_all(self):
        """全选"""
        for checkbox in self.checkboxes.values():
            checkbox.setChecked(True)
    
    def select_none(self):
        """取消全选"""
        for checkbox in self.checkboxes.values():
            checkbox.setChecked(False)
    
    def restore_default(self):
        """恢复默认设置"""
        for col_idx, checkbox in self.checkboxes.items():
            # 查找默认设置
            for col in self.ALL_COLUMNS:
                if col[0] == col_idx:
                    checkbox.setChecked(col[3])
                    break
    
    def get_visible_columns(self):
        """获取选中的列索引列表（始终包含缩进列和序号列）"""
        visible_cols = [0, 1]  # 缩进列和序号列始终显示
        visible_cols.extend([col_idx for col_idx, checkbox in self.checkboxes.items() if checkbox.isChecked()])
        return visible_cols


class BiddingDetailTab(QWidget):
    """投标明细页签 - 树形表格展示层级结构"""

    def __init__(self, parent_page):
        super().__init__()
        self.parent_page = parent_page
        self.bidding_manager = parent_page.bidding_manager
        self.detail_manager = DetailManager(parent_page.bidding_manager.db_manager)
        self.current_bidding_id = None
        self.current_bidding_code = None
        self.current_bidding_name = None
        self.current_summary_item = None
        self.current_detail_id = None  # 当前明细表ID
        # 获取WD基础路径 - 使用固定路径，与投标汇总表保持一致
        import os
        self.wd_base_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "WD")

        # 初始化可见列（默认显示第2-14列，第0列缩进列和第1列序号始终显示）
        self.visible_columns = list(range(2, 15))  # 2-14列默认显示（缩进列0和序号列1始终显示，不在此列表中）

        # 自动保存定时器（防抖，延迟500ms保存）
        self.auto_save_timer = QTimer(self)
        self.auto_save_timer.setSingleShot(True)
        self.auto_save_timer.timeout.connect(self._auto_save)
        self._pending_save = False  # 标记是否有待保存的数据

        self.setup_ui()
        
        # 应用初始列可见性
        self.apply_column_visibility()

    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # 提示信息
        self.tip_label = QLabel("请先选择投标项目，或在汇总表中双击行查看明细")
        self.tip_label.setAlignment(Qt.AlignCenter)
        self.tip_label.setObjectName("tipLabel")
        layout.addWidget(self.tip_label)

        # 明细内容区域（初始隐藏）
        self.content_widget = QWidget()
        content_layout = QVBoxLayout(self.content_widget)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(15)

        # 顶部信息栏 - 显示当前投标信息
        info_frame = QWidget()
        info_layout = QVBoxLayout(info_frame)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_layout.setSpacing(10)

        # 投标编码和名称
        bidding_info_layout = QHBoxLayout()
        bidding_info_layout.addWidget(QLabel("投标编码:"))
        self.bidding_code_label = QLabel("")
        self.bidding_code_label.setObjectName("biddingCodeLabel")
        bidding_info_layout.addWidget(self.bidding_code_label)

        bidding_info_layout.addSpacing(30)

        bidding_info_layout.addWidget(QLabel("投标名称:"))
        self.bidding_name_label = QLabel("")
        self.bidding_name_label.setObjectName("biddingNameLabel")
        bidding_info_layout.addWidget(self.bidding_name_label)

        bidding_info_layout.addStretch()
        info_layout.addLayout(bidding_info_layout)

        # 工程项目及费用名称
        project_info_layout = QHBoxLayout()
        project_info_layout.addWidget(QLabel("工程项目及费用名称:"))
        self.project_item_label = QLabel("")
        self.project_item_label.setObjectName("projectItemLabel")
        project_info_layout.addWidget(self.project_item_label)
        project_info_layout.addStretch()
        info_layout.addLayout(project_info_layout)

        content_layout.addWidget(info_frame)

        # Excel导入工具栏
        excel_toolbar = QHBoxLayout()
        excel_toolbar.setSpacing(10)

        # Excel文件选择器
        self.excel_selector = ExcelFileSelector()
        excel_toolbar.addWidget(self.excel_selector)

        # 导入按钮
        self.import_excel_btn = PrimaryPushButton("📥 导入明细Excel")
        self.import_excel_btn.clicked.connect(self.on_import_excel)
        excel_toolbar.addWidget(self.import_excel_btn)

        # 重置级别按钮
        self.reset_level_btn = PushButton("🔄 重置为级别1")
        self.reset_level_btn.setToolTip("将所有导入的数据重置为级别1，保持原有顺序")
        self.reset_level_btn.clicked.connect(self.reset_all_to_level_one)
        excel_toolbar.addWidget(self.reset_level_btn)

        # 导出按钮
        self.export_excel_btn = PushButton("📤 导出Excel")
        self.export_excel_btn.setToolTip("将当前明细数据导出为Excel文件")
        self.export_excel_btn.clicked.connect(self.on_export_excel)
        excel_toolbar.addWidget(self.export_excel_btn)

        excel_toolbar.addStretch()

        # 合计金额显示
        excel_toolbar.addWidget(QLabel("合计:"))
        self.total_label = QLabel("¥ 0.00")
        self.total_label.setObjectName("totalLabel")
        excel_toolbar.addWidget(self.total_label)

        content_layout.addLayout(excel_toolbar)

        # 树形表格区域
        table_frame = QWidget()
        table_layout = QVBoxLayout(table_frame)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(10)

        # 工具栏
        toolbar_layout = QHBoxLayout()
        toolbar_layout.setSpacing(10)
        
        # 展开/折叠按钮
        self.expand_btn = PushButton("📂 展开")
        self.expand_btn.clicked.connect(self.expand_all)
        toolbar_layout.addWidget(self.expand_btn)

        self.collapse_btn = PushButton("📁 折叠")
        self.collapse_btn.clicked.connect(self.collapse_all)
        toolbar_layout.addWidget(self.collapse_btn)

        toolbar_layout.addSpacing(20)

        # 添加明细按钮
        self.add_detail_btn = PrimaryPushButton("➕ 添加明细")
        self.add_detail_btn.clicked.connect(self.on_add_detail)
        toolbar_layout.addWidget(self.add_detail_btn)

        # 删除明细按钮
        self.delete_detail_btn = PushButton("🗑️ 删除明细")
        self.delete_detail_btn.setObjectName("secondaryButton")
        self.delete_detail_btn.clicked.connect(self.on_delete_detail)
        toolbar_layout.addWidget(self.delete_detail_btn)

        toolbar_layout.addSpacing(20)

        # 显示列设置按钮
        self.column_settings_btn = PushButton("⚙️ 显示列")
        self.column_settings_btn.setToolTip("设置要显示的列")
        self.column_settings_btn.clicked.connect(self.on_column_settings)
        toolbar_layout.addWidget(self.column_settings_btn)

        toolbar_layout.addStretch()

        # 保存明细按钮（已改为自动保存，隐藏按钮）
        # self.save_btn = PrimaryPushButton("💾 保存明细")
        # self.save_btn.clicked.connect(self.on_save_detail)
        # toolbar_layout.addWidget(self.save_btn)

        # 自动保存状态标签
        self.auto_save_label = QLabel("✓ 自动保存已开启")
        self.auto_save_label.setStyleSheet("color: #28a745; font-size: 12px;")
        toolbar_layout.addWidget(self.auto_save_label)

        table_layout.addLayout(toolbar_layout)

        # 树形表格 - 使用QTreeWidget替代QTableWidget以支持层级
        self.detail_tree = QTreeWidget()
        self.detail_tree.setColumnCount(24)  # 增加1列专门用于层级缩进，1列主材损耗%
        self.detail_tree.setHeaderLabels([
            "", "序号", "分部分项工程名称", "规格型号", "项目特征描述", "单位",
            "工程量", "综合单价", "人工单价", "主材单价", "主材损耗%", "辅材单价", "机械单价", "其他单价",
            "合价(净)", "人工合价", "主材合价", "辅材合价", "机械合价", "其他合价",
            "管理费合计", "税金合价", "综合合价", "备注"
        ])
        
        # 启用原生树形连接线
        self.detail_tree.setRootIsDecorated(True)
        self.detail_tree.setItemsExpandable(True)
        self.detail_tree.setExpandsOnDoubleClick(True)
        self.detail_tree.setIndentation(20)  # 每级缩进20像素
        self.detail_tree.setUniformRowHeights(True)  # 统一行高

        # 设置样式
        self.detail_tree.setObjectName("detailTree")

        self.detail_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.detail_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.detail_tree.setAlternatingRowColors(True)

        # 列宽设置将在apply_column_visibility中统一处理

        # 启用编辑
        self.detail_tree.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)

        # 连接信号
        self.detail_tree.itemChanged.connect(self.on_item_changed)

        # 启用右键菜单
        self.detail_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.detail_tree.customContextMenuRequested.connect(self.show_context_menu)

        # 双击单元格显示详情
        self.detail_tree.itemDoubleClicked.connect(self.on_item_double_clicked)

        table_layout.addWidget(self.detail_tree)
        content_layout.addWidget(table_frame)

        layout.addWidget(self.content_widget)
        self.content_widget.hide()

    def get_excel_directory(self) -> str:
        """获取Excel文件目录"""
        if not self.current_bidding_id:
            return ""

        if not self.wd_base_path:
            return ""

        # 获取项目名称
        bidding = self.bidding_manager.get_bidding(self.current_bidding_id)
        if not bidding:
            return ""

        project = self.bidding_manager.project_manager.get_project(bidding.project_id) if hasattr(self.bidding_manager, 'project_manager') else None
        project_name = project.name if project else "未知项目"

        # 构建路径: WD/项目名/投标附件/
        import os
        folder_path = os.path.join(self.wd_base_path, project_name, "投标附件")
        try:
            os.makedirs(folder_path, exist_ok=True)
        except Exception as e:
            pass  # print(f"创建投标附件目录失败: {e}")
        return folder_path

    def on_bidding_selected(self, bidding_id: int, bidding_code: str = None):
        """从父页面接收选中的投标"""
        if bidding_id:
            self.current_bidding_id = bidding_id
            self.current_bidding_code = bidding_code
            
            # 获取投标名称
            bidding = self.bidding_manager.get_bidding(bidding_id)
            if bidding:
                self.current_bidding_name = bidding.bidding_name
            
            # 更新Excel文件选择器目录
            excel_dir = self.get_excel_directory()
            self.excel_selector.set_directory(excel_dir)
            
            # 如果没有选中汇总项，只显示提示
            if not self.current_summary_item:
                self.show_tip()
        else:
            self.show_tip()
            self.current_bidding_id = None
            self.current_bidding_code = None
            self.current_bidding_name = None
            self.current_summary_item = None

    def load_detail_data(self, bidding_id: int, bidding_code: str, bidding_name: str, 
                         project_item_name: str, summary_item=None):
        """加载明细数据 - 从汇总表双击调用"""
        self.current_bidding_id = bidding_id
        self.current_bidding_code = bidding_code
        self.current_bidding_name = bidding_name
        self.current_summary_item = summary_item

        # 显示内容区域
        self.show_content()

        # 更新投标信息标签
        self.bidding_code_label.setText(bidding_code or "")
        self.bidding_name_label.setText(bidding_name or "")
        self.project_item_label.setText(project_item_name or "")

        # 更新Excel文件选择器目录
        excel_dir = self.get_excel_directory()
        self.excel_selector.set_directory(excel_dir)

        # 加载明细数据
        self.load_detail_tree_data()

    def load_detail_tree_data(self):
        """加载明细树形数据 - 支持大数据量进度显示"""
        self.detail_tree.clear()
        self.current_detail_id = None
        
        if not self.current_summary_item:
            print("[加载明细] 当前汇总项为空")
            return
        
        # 尝试从数据库加载明细数据
        summary_item_id = getattr(self.current_summary_item, 'id', 0)
        summary_item_name = getattr(self.current_summary_item, 'name', 'Unknown')
        print(f"[加载明细] 投标ID: {self.current_bidding_id}, 汇总项ID: {summary_item_id}, 名称: {summary_item_name}")
        
        if summary_item_id and self.current_bidding_id:
            detail = self.detail_manager.get_detail_by_summary_item(
                self.current_bidding_id, summary_item_id
            )
            print(f"[加载明细] 查询结果: {detail}")
            if detail:
                print(f"[加载明细] 明细ID: {detail.id}, 项目数: {len(detail.items)}")
                if detail.items:
                    # 计算总项目数（包括子节点）
                    total_count = self._count_detail_items(detail.items)
                    print(f"[加载明细] 总项目数（含子节点）: {total_count}")
                    
                    # 数据量大时显示进度条
                    if total_count > 50:
                        from PySide6.QtWidgets import QProgressDialog
                        progress = QProgressDialog("正在加载明细数据...", "取消", 0, total_count, self)
                        progress.setWindowModality(Qt.WindowModal)
                        progress.setWindowTitle("加载进度")
                        progress.setMinimumDuration(0)
                        progress.setValue(0)
                        progress.setMinimumSize(500, 150)

                        self.current_detail_id = detail.id
                        self._build_tree_from_detail_items(detail.items, None, progress, total_count)
                        progress.setValue(total_count)
                        progress.close()
                        # 加载完成后，从叶子节点向上计算所有父级节点的汇总值
                        self._update_all_tree_items_display_with_calc()
                    else:
                        self.current_detail_id = detail.id
                        self._build_tree_from_detail_items(detail.items)
                        # 加载完成后，从叶子节点向上计算所有父级节点的汇总值
                        self._update_all_tree_items_display_with_calc()

                    self.calculate_total()
                    print(f"[加载明细] 成功加载 {len(detail.items)} 个顶层项目")
                    return
            else:
                print("[加载明细] 数据库中没有找到明细数据")
        else:
            print(f"[加载明细] 参数无效: bidding_id={self.current_bidding_id}, summary_item_id={summary_item_id}")
        
        # 数据库中没有数据，显示空表（新建汇总行）
        print("[加载明细] 新建汇总行，显示空表")
        # 不加载任何数据，保持空表状态
        self.calculate_total()
    
    def _count_detail_items(self, items: list) -> int:
        """递归计算明细项目总数（包括子节点）"""
        count = 0
        for item in items:
            count += 1
            if item.children:
                count += self._count_detail_items(item.children)
        return count

    def _build_tree_from_detail_items(self, items: list, parent_tree_item=None, progress=None, total_count=None, current_count=None):
        """从DetailItem模型构建树形控件 - 支持进度显示"""
        if current_count is None:
            current_count = [0]
        
        for item in items:
            # 更新进度
            if progress and total_count:
                current_count[0] += 1
                if current_count[0] % 10 == 0:
                    progress.setValue(current_count[0])
                    progress.setLabelText(f"正在加载明细数据... ({current_count[0]}/{total_count})")
                    QApplication.processEvents()
                    
                    if progress.wasCanceled():
                        return
            # 创建树节点
            if parent_tree_item:
                tree_item = QTreeWidgetItem(parent_tree_item)
            else:
                tree_item = QTreeWidgetItem(self.detail_tree)

            # 创建UI用的DetailItem
            ui_item = DetailItem()
            ui_item.sequence = item.sequence
            ui_item.name = item.name
            ui_item.specification = item.specification
            ui_item.description = item.description
            ui_item.unit = item.unit
            ui_item.quantity = item.quantity
            ui_item.unit_price = item.unit_price
            ui_item.labor_unit_price = item.labor_price
            ui_item.material_unit_price = item.main_material_price
            ui_item.material_loss_rate = item.main_material_loss_rate
            ui_item.auxiliary_unit_price = item.aux_material_price
            ui_item.machine_unit_price = item.machinery_price
            ui_item.other_unit_price = item.other_price
            ui_item.total_price = item.total_price
            ui_item.labor_total = item.labor_total
            ui_item.material_total = item.material_total
            ui_item.auxiliary_total = item.auxiliary_total
            ui_item.machine_total = item.machine_total
            ui_item.other_total = item.other_total
            ui_item.management_total = item.management_total
            ui_item.tax_total = item.tax_total
            ui_item.comprehensive_total = item.comprehensive_total
            ui_item.remark = item.remark
            ui_item.level = item.level

            # 递归处理子节点（先处理子节点以计算父级汇总）
            if item.children:
                self._build_tree_from_detail_items(item.children, tree_item, progress, total_count, current_count)
                tree_item.setExpanded(True)
                # 计算父级节点的汇总值（从子节点重新计算，确保数据正确）
                self._calculate_parent_totals(tree_item, ui_item)

            # 设置显示文本（值为0的显示为空）
            tree_item.setText(0, "")
            tree_item.setText(1, ui_item.sequence)
            tree_item.setText(2, ui_item.name)
            tree_item.setText(3, ui_item.specification)
            tree_item.setText(4, ui_item.description)
            tree_item.setText(5, ui_item.unit)
            
            # 如果有子节点，父节点只显示合价，不显示工程量和单价
            if item.children:
                tree_item.setText(6, "")  # 工程量
                tree_item.setText(7, "")  # 综合单价
                tree_item.setText(8, "")  # 人工单价
                tree_item.setText(9, "")  # 主材单价
                tree_item.setText(10, "")  # 损耗率
                tree_item.setText(11, "")  # 辅材单价
                tree_item.setText(12, "")  # 机械单价
                tree_item.setText(13, "")  # 其他单价
            else:
                # 叶子节点显示所有值
                tree_item.setText(6, self._format_number(ui_item.quantity))
                tree_item.setText(7, self._format_number(ui_item.unit_price))
                tree_item.setText(8, self._format_number(ui_item.labor_unit_price))
                tree_item.setText(9, self._format_number(ui_item.material_unit_price))
                tree_item.setText(10, self._format_number(ui_item.material_loss_rate))
                tree_item.setText(11, self._format_number(ui_item.auxiliary_unit_price))
                tree_item.setText(12, self._format_number(ui_item.machine_unit_price))
                tree_item.setText(13, self._format_number(ui_item.other_unit_price))
            
            # 所有节点都显示合价
            tree_item.setText(14, self._format_number(ui_item.total_price))
            tree_item.setText(15, self._format_number(ui_item.labor_total))
            tree_item.setText(16, self._format_number(ui_item.material_total))
            tree_item.setText(17, self._format_number(ui_item.auxiliary_total))
            tree_item.setText(18, self._format_number(ui_item.machine_total))
            tree_item.setText(19, self._format_number(ui_item.other_total))
            tree_item.setText(20, self._format_number(ui_item.management_total))
            tree_item.setText(21, self._format_number(ui_item.tax_total))
            tree_item.setText(22, self._format_number(ui_item.comprehensive_total))
            tree_item.setText(23, ui_item.remark)

            # 存储数据
            tree_item.setData(0, Qt.UserRole, ui_item)

            # 设置对齐
            tree_item.setTextAlignment(1, Qt.AlignCenter)
            for col in range(6, 23):
                tree_item.setTextAlignment(col, Qt.AlignRight | Qt.AlignVCenter)

            # 一级节点加粗
            if ui_item.level == 1:
                font = QFont("Microsoft YaHei", 10, QFont.Bold)
                for col in range(24):
                    tree_item.setFont(col, font)

    def _format_number(self, value: float) -> str:
        """格式化数字，0值显示为空"""
        if value is None or value == 0:
            return ""
        return f"{value:,.2f}"

    def _calculate_parent_totals(self, tree_item: QTreeWidgetItem, ui_item: DetailItem):
        """计算父级节点的汇总值（从子节点汇总）"""
        # 重置汇总值
        total_price = 0.0
        labor_total = 0.0
        material_total = 0.0
        auxiliary_total = 0.0
        machine_total = 0.0
        other_total = 0.0
        management_total = 0.0
        tax_total = 0.0
        comprehensive_total = 0.0

        # 累加所有子节点的值
        for i in range(tree_item.childCount()):
            child_item = tree_item.child(i)
            child_data = child_item.data(0, Qt.UserRole)
            if child_data:
                total_price += child_data.total_price
                labor_total += child_data.labor_total
                material_total += child_data.material_total
                auxiliary_total += child_data.auxiliary_total
                machine_total += child_data.machine_total
                other_total += child_data.other_total
                management_total += child_data.management_total
                tax_total += child_data.tax_total
                comprehensive_total += child_data.comprehensive_total

        # 更新父级节点的值
        ui_item.total_price = total_price
        ui_item.labor_total = labor_total
        ui_item.material_total = material_total
        ui_item.auxiliary_total = auxiliary_total
        ui_item.machine_total = machine_total
        ui_item.other_total = other_total
        ui_item.management_total = management_total
        ui_item.tax_total = tax_total
        ui_item.comprehensive_total = comprehensive_total

    def _generate_sample_data(self, summary_name: str) -> list:
        """根据汇总项名称生成示例明细数据"""
        if "人工" in summary_name:
            return [
                {"sequence": "1", "name": "人工费小计", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 34000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "普通工", "specification": "", "description": "土方工程用工", "unit": "工日", "quantity": 100, "unit_price": 150.00, "total_price": 15000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "技术工", "specification": "", "description": "砌筑工程用工", "unit": "工日", "quantity": 50, "unit_price": 200.00, "total_price": 10000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "高级技工", "specification": "", "description": "安装工程用工", "unit": "工日", "quantity": 30, "unit_price": 300.00, "total_price": 9000, "remark": "", "level": 2},
                ]},
            ]
        elif "材料" in summary_name or "主材" in summary_name:
            return [
                {"sequence": "1", "name": "材料费小计", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 299000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "水泥", "specification": "P.O 42.5", "description": "普通硅酸盐水泥", "unit": "t", "quantity": 100, "unit_price": 500.00, "total_price": 50000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "钢筋", "specification": "HRB400", "description": "热轧带肋钢筋", "unit": "t", "quantity": 50, "unit_price": 4500.00, "total_price": 225000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "砂石", "specification": "中砂", "description": "建筑用砂", "unit": "m³", "quantity": 200, "unit_price": 120.00, "total_price": 24000, "remark": "", "level": 2},
                ]},
            ]
        elif "机械" in summary_name:
            return [
                {"sequence": "1", "name": "机械费小计", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 56000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "挖掘机", "specification": "斗容量1m³", "description": "液压挖掘机", "unit": "台班", "quantity": 20, "unit_price": 1200.00, "total_price": 24000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "起重机", "specification": "25吨", "description": "汽车起重机", "unit": "台班", "quantity": 10, "unit_price": 2000.00, "total_price": 20000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "混凝土搅拌机", "specification": "500L", "description": "强制式搅拌机", "unit": "台班", "quantity": 15, "unit_price": 800.00, "total_price": 12000, "remark": "", "level": 2},
                ]},
            ]
        elif "土方" in summary_name or "开挖" in summary_name:
            return [
                {"sequence": "1", "name": "土方工程", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 52500, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "土方开挖", "specification": "", "description": "机械挖土，运距1km", "unit": "m³", "quantity": 1000, "unit_price": 25.00, "total_price": 25000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "土方回填", "specification": "", "description": "机械回填夯实", "unit": "m³", "quantity": 500, "unit_price": 20.00, "total_price": 10000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "余土外运", "specification": "", "description": "运距5km", "unit": "m³", "quantity": 500, "unit_price": 35.00, "total_price": 17500, "remark": "", "level": 2},
                ]},
            ]
        elif "混凝土" in summary_name:
            return [
                {"sequence": "1", "name": "混凝土工程", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 235000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "混凝土浇筑", "specification": "C30", "description": "C30混凝土", "unit": "m³", "quantity": 500, "unit_price": 450.00, "total_price": 225000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "混凝土养护", "specification": "", "description": "洒水养护", "unit": "m²", "quantity": 2000, "unit_price": 5.00, "total_price": 10000, "remark": "", "level": 2},
                ]},
            ]
        elif "模板" in summary_name:
            return [
                {"sequence": "1", "name": "模板工程", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 155000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "模板制作", "specification": "木模板", "description": "木质模板制作", "unit": "m²", "quantity": 2000, "unit_price": 45.00, "total_price": 90000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "模板安装", "specification": "钢模板", "description": "钢质模板安装", "unit": "m²", "quantity": 1000, "unit_price": 35.00, "total_price": 35000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "模板拆除", "specification": "", "description": "模板拆除清理", "unit": "m²", "quantity": 3000, "unit_price": 10.00, "total_price": 30000, "remark": "", "level": 2},
                ]},
            ]
        elif "脚手" in summary_name:
            return [
                {"sequence": "1", "name": "脚手架工程", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 64500, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "钢管脚手架", "specification": "双排20m", "description": "双排，高度20m", "unit": "m²", "quantity": 1500, "unit_price": 35.00, "total_price": 52500, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "脚手架拆除", "specification": "", "description": "脚手架拆除", "unit": "m²", "quantity": 1500, "unit_price": 8.00, "total_price": 12000, "remark": "", "level": 2},
                ]},
            ]
        elif "钢筋" in summary_name:
            return [
                {"sequence": "1", "name": "钢筋工程", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 300000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "钢筋制作", "specification": "HRB400≤10", "description": "HRB400，直径≤10mm", "unit": "t", "quantity": 20, "unit_price": 5500.00, "total_price": 110000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "钢筋制作", "specification": "HRB400>10", "description": "HRB400，直径>10mm", "unit": "t", "quantity": 30, "unit_price": 5000.00, "total_price": 150000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "钢筋安装", "specification": "", "description": "钢筋安装绑扎", "unit": "t", "quantity": 50, "unit_price": 800.00, "total_price": 40000, "remark": "", "level": 2},
                ]},
            ]
        else:
            return [
                {"sequence": "1", "name": f"{summary_name}明细", "specification": "", "description": "", "unit": "", "quantity": 0, "unit_price": 0, "total_price": 30000, "remark": "", "level": 1, "children": [
                    {"sequence": "1.1", "name": "分部分项1", "specification": "", "description": f"{summary_name}相关", "unit": "m²", "quantity": 100, "unit_price": 100.00, "total_price": 10000, "remark": "", "level": 2},
                    {"sequence": "1.2", "name": "分部分项2", "specification": "", "description": f"{summary_name}相关", "unit": "m³", "quantity": 50, "unit_price": 200.00, "total_price": 10000, "remark": "", "level": 2},
                    {"sequence": "1.3", "name": "分部分项3", "specification": "", "description": f"{summary_name}相关", "unit": "t", "quantity": 10, "unit_price": 1000.00, "total_price": 10000, "remark": "", "level": 2},
                ]},
            ]

    def build_tree_from_data(self, data: list):
        """从数据构建树形结构"""
        self.detail_tree.clear()
        
        if not data:
            return
        
        for item_data in data:
            self._add_tree_item_recursive(None, item_data)
        
        self.detail_tree.expandAll()

    def _add_tree_item_recursive(self, parent_item, item_data: dict):
        """递归添加树节点"""
        tree_item = QTreeWidgetItem(parent_item if parent_item else self.detail_tree)

        # 创建DetailItem对象存储数据
        detail_item = DetailItem()
        detail_item.sequence = item_data.get('sequence', '')
        detail_item.name = item_data.get('name', '')
        detail_item.specification = item_data.get('specification', '')
        detail_item.description = item_data.get('description', '')
        detail_item.unit = item_data.get('unit', '')
        detail_item.quantity = item_data.get('quantity', 0)
        detail_item.unit_price = item_data.get('unit_price', 0)
        detail_item.labor_unit_price = item_data.get('labor_unit_price', 0)
        detail_item.material_unit_price = item_data.get('material_unit_price', 0)
        detail_item.material_loss_rate = item_data.get('material_loss_rate', 0)
        detail_item.auxiliary_unit_price = item_data.get('auxiliary_unit_price', 0)
        detail_item.machine_unit_price = item_data.get('machine_unit_price', 0)
        detail_item.other_unit_price = item_data.get('other_unit_price', 0)
        detail_item.total_price = item_data.get('total_price', 0)
        detail_item.labor_total = item_data.get('labor_total', 0)
        detail_item.material_total = item_data.get('material_total', 0)
        detail_item.auxiliary_total = item_data.get('auxiliary_total', 0)
        detail_item.machine_total = item_data.get('machine_total', 0)
        detail_item.other_total = item_data.get('other_total', 0)
        detail_item.management_total = item_data.get('management_total', 0)
        detail_item.tax_total = item_data.get('tax_total', 0)
        detail_item.comprehensive_total = item_data.get('comprehensive_total', 0)
        detail_item.remark = item_data.get('remark', '')
        detail_item.level = item_data.get('level', 1)

        # 递归添加子节点（先处理子节点以计算父级汇总）
        children = item_data.get('children', [])
        for child_data in children:
            self._add_tree_item_recursive(tree_item, child_data)

        if children:
            tree_item.setExpanded(True)
            # 计算父级节点的汇总值
            self._calculate_parent_totals_from_tree(tree_item, detail_item)

        # 设置显示文本（值为0的显示为空）
        tree_item.setText(0, "")  # 缩进列（空）
        tree_item.setText(1, detail_item.sequence)
        tree_item.setText(2, detail_item.name)
        tree_item.setText(3, detail_item.specification)
        tree_item.setText(4, detail_item.description)  # 项目特征描述
        tree_item.setText(5, detail_item.unit)
        
        # 如果有子节点，父节点只显示合价，不显示工程量和单价
        if children:
            tree_item.setText(6, "")  # 工程量
            tree_item.setText(7, "")  # 综合单价
            tree_item.setText(8, "")  # 人工单价
            tree_item.setText(9, "")  # 主材单价
            tree_item.setText(10, "")  # 损耗率
            tree_item.setText(11, "")  # 辅材单价
            tree_item.setText(12, "")  # 机械单价
            tree_item.setText(13, "")  # 其他单价
        else:
            # 叶子节点显示所有值
            tree_item.setText(6, self._format_number(detail_item.quantity))
            tree_item.setText(7, self._format_number(detail_item.unit_price))
            tree_item.setText(8, self._format_number(detail_item.labor_unit_price))
            tree_item.setText(9, self._format_number(detail_item.material_unit_price))
            tree_item.setText(10, self._format_number(detail_item.material_loss_rate))
            tree_item.setText(11, self._format_number(detail_item.auxiliary_unit_price))
            tree_item.setText(12, self._format_number(detail_item.machine_unit_price))
            tree_item.setText(13, self._format_number(detail_item.other_unit_price))
        
        # 所有节点都显示合价
        tree_item.setText(14, self._format_number(detail_item.total_price))
        tree_item.setText(15, self._format_number(detail_item.labor_total))
        tree_item.setText(16, self._format_number(detail_item.material_total))
        tree_item.setText(17, self._format_number(detail_item.auxiliary_total))
        tree_item.setText(18, self._format_number(detail_item.machine_total))
        tree_item.setText(19, self._format_number(detail_item.other_total))
        tree_item.setText(20, self._format_number(detail_item.management_total))
        tree_item.setText(21, self._format_number(detail_item.tax_total))
        tree_item.setText(22, self._format_number(detail_item.comprehensive_total))
        tree_item.setText(23, detail_item.remark)

        # 存储数据
        tree_item.setData(0, Qt.UserRole, detail_item)

        # 设置文本对齐
        tree_item.setTextAlignment(1, Qt.AlignCenter)  # 序号列居中
        for col in range(6, 23):  # 数值列右对齐
            tree_item.setTextAlignment(col, Qt.AlignRight | Qt.AlignVCenter)

        # 一级节点加粗
        if detail_item.level == 1:
            font = QFont("Microsoft YaHei", 10, QFont.Bold)
            for col in range(24):
                tree_item.setFont(col, font)

    def _calculate_parent_totals_from_tree(self, tree_item: QTreeWidgetItem, detail_item: DetailItem):
        """从树形控件计算父级节点的汇总值"""
        total_price = 0.0
        labor_total = 0.0
        material_total = 0.0
        auxiliary_total = 0.0
        machine_total = 0.0
        other_total = 0.0
        management_total = 0.0
        tax_total = 0.0
        comprehensive_total = 0.0

        # 累加所有子节点的值
        for i in range(tree_item.childCount()):
            child_item = tree_item.child(i)
            child_data = child_item.data(0, Qt.UserRole)
            if child_data:
                total_price += child_data.total_price
                labor_total += child_data.labor_total
                material_total += child_data.material_total
                auxiliary_total += child_data.auxiliary_total
                machine_total += child_data.machine_total
                other_total += child_data.other_total
                management_total += child_data.management_total
                tax_total += child_data.tax_total
                comprehensive_total += child_data.comprehensive_total

        # 更新父级节点的值
        detail_item.total_price = total_price
        detail_item.labor_total = labor_total
        detail_item.material_total = material_total
        detail_item.auxiliary_total = auxiliary_total
        detail_item.machine_total = machine_total
        detail_item.other_total = other_total
        detail_item.management_total = management_total
        detail_item.tax_total = tax_total
        detail_item.comprehensive_total = comprehensive_total

    def build_tree_from_imported_data(self, data: list):
        """从导入的数据构建树形结构 - 带进度显示"""
        from PySide6.QtWidgets import QProgressDialog
        from PySide6.QtCore import Qt
        
        self.detail_tree.clear()
        
        if not data:
            return
        
        # 创建进度对话框
        progress = QProgressDialog("正在构建明细数据...", "取消", 0, len(data), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setWindowTitle("导入进度")
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.setMinimumSize(600, 200)

        # 设置样式
        progress.setObjectName("importProgressDialog")
        
        # 使用栈构建层级结构
        stack = [(0, None)]  # (层级, 父节点)
        
        for idx, item_data in enumerate(data):
            # 更新进度
            if idx % 10 == 0:
                progress.setValue(idx)
                progress.setLabelText(f"正在构建明细数据... ({idx}/{len(data)})\n当前: {item_data.get('sequence', '')} {item_data.get('name', '')[:20]}")
                QApplication.processEvents()
                
                if progress.wasCanceled():
                    break
            
            level = item_data.get('level', 2)
            
            # 找到正确的父节点
            while stack and stack[-1][0] >= level:
                stack.pop()
            
            parent_tree_item = stack[-1][1] if stack else None
            
            # 创建树节点
            tree_item = QTreeWidgetItem(parent_tree_item if parent_tree_item else self.detail_tree)
            
            detail_item = DetailItem()
            detail_item.sequence = str(item_data.get('sequence', ''))
            detail_item.name = str(item_data.get('name', ''))
            detail_item.specification = str(item_data.get('specification', ''))
            detail_item.description = str(item_data.get('description', ''))
            detail_item.unit = str(item_data.get('unit', ''))
            detail_item.quantity = float(item_data.get('quantity', 0))
            detail_item.unit_price = float(item_data.get('unit_price', 0))
            detail_item.labor_unit_price = float(item_data.get('labor_unit_price', 0))
            detail_item.material_unit_price = float(item_data.get('material_unit_price', 0))
            detail_item.material_loss_rate = float(item_data.get('material_loss_rate', 0))
            detail_item.auxiliary_unit_price = float(item_data.get('auxiliary_unit_price', 0))
            detail_item.machine_unit_price = float(item_data.get('machine_unit_price', 0))
            detail_item.other_unit_price = float(item_data.get('other_unit_price', 0))
            detail_item.remark = str(item_data.get('remark', ''))
            detail_item.level = level
            
            # 调试输出
            print(f"[构建树调试] name={detail_item.name}, quantity={detail_item.quantity}, unit_price={detail_item.unit_price}")
            
            # 计算合价
            detail_item.calculate_total()
            
            # 设置基本文本信息
            tree_item.setText(0, "")
            tree_item.setText(1, detail_item.sequence)
            tree_item.setText(2, detail_item.name)
            tree_item.setText(3, detail_item.specification)
            tree_item.setText(4, detail_item.description)
            tree_item.setText(5, detail_item.unit)
            tree_item.setText(23, detail_item.remark)
            
            # 设置工程量和单价显示
            tree_item.setText(6, self._format_number(detail_item.quantity))
            tree_item.setText(7, self._format_number(detail_item.unit_price))
            tree_item.setText(8, self._format_number(detail_item.labor_unit_price))
            tree_item.setText(9, self._format_number(detail_item.material_unit_price))
            tree_item.setText(10, self._format_number(detail_item.material_loss_rate))
            tree_item.setText(11, self._format_number(detail_item.auxiliary_unit_price))
            tree_item.setText(12, self._format_number(detail_item.machine_unit_price))
            tree_item.setText(13, self._format_number(detail_item.other_unit_price))
            tree_item.setText(14, self._format_number(detail_item.total_price))
            tree_item.setText(15, self._format_number(detail_item.labor_total))
            tree_item.setText(16, self._format_number(detail_item.material_total))
            tree_item.setText(17, self._format_number(detail_item.auxiliary_total))
            tree_item.setText(18, self._format_number(detail_item.machine_total))
            tree_item.setText(19, self._format_number(detail_item.other_total))
            tree_item.setText(20, self._format_number(detail_item.management_total))
            tree_item.setText(21, self._format_number(detail_item.tax_total))
            tree_item.setText(22, self._format_number(detail_item.comprehensive_total))
            
            # 存储数据
            tree_item.setData(0, Qt.UserRole, detail_item)
            
            # 设置对齐
            tree_item.setTextAlignment(1, Qt.AlignCenter)
            for col in range(6, 23):
                tree_item.setTextAlignment(col, Qt.AlignRight | Qt.AlignVCenter)
            
            # 一级节点加粗
            if level == 1:
                font = QFont("Microsoft YaHei", 10, QFont.Bold)
                for col in range(24):
                    tree_item.setFont(col, font)
            
            stack.append((level, tree_item))
        
        progress.setValue(len(data))
        progress.close()

        # 计算父级节点的汇总值（从叶子节点向上汇总）
        self._update_all_tree_items_display_with_calc()

        self.detail_tree.expandAll()
        self.calculate_total()

    def _update_all_tree_items_display(self):
        """更新所有树节点的显示（从叶子节点开始向上更新）"""
        self._update_all_tree_items_display_with_calc()

    def _update_all_tree_items_display_with_calc(self):
        """更新所有树节点的显示，并计算父级节点的汇总值（从叶子节点开始向上更新）"""
        def update_item_display(item: QTreeWidgetItem):
            # 先递归处理子节点
            for i in range(item.childCount()):
                update_item_display(item.child(i))
            
            # 更新当前节点的显示
            detail_item = item.data(0, Qt.UserRole)
            if detail_item:
                print(f"[更新显示] {detail_item.name}, childCount={item.childCount()}")
                print(f"[更新显示]   quantity={detail_item.quantity}, unit_price={detail_item.unit_price}")
                print(f"[更新显示]   total_price={detail_item.total_price}, labor_total={detail_item.labor_total}")
                # 如果有子节点，计算汇总值并只显示合价
                if item.childCount() > 0:
                    # 计算父级节点的汇总值
                    total_price = 0.0
                    labor_total = 0.0
                    material_total = 0.0
                    auxiliary_total = 0.0
                    machine_total = 0.0
                    other_total = 0.0
                    management_total = 0.0
                    tax_total = 0.0
                    comprehensive_total = 0.0
                    
                    for i in range(item.childCount()):
                        child_item = item.child(i)
                        child_data = child_item.data(0, Qt.UserRole)
                        if child_data:
                            total_price += child_data.total_price
                            labor_total += child_data.labor_total
                            material_total += child_data.material_total
                            auxiliary_total += child_data.auxiliary_total
                            machine_total += child_data.machine_total
                            other_total += child_data.other_total
                            management_total += child_data.management_total
                            tax_total += child_data.tax_total
                            comprehensive_total += child_data.comprehensive_total
                    
                    # 更新父节点的数据
                    detail_item.total_price = total_price
                    detail_item.labor_total = labor_total
                    detail_item.material_total = material_total
                    detail_item.auxiliary_total = auxiliary_total
                    detail_item.machine_total = machine_total
                    detail_item.other_total = other_total
                    detail_item.management_total = management_total
                    detail_item.tax_total = tax_total
                    detail_item.comprehensive_total = comprehensive_total
                    
                    # 父节点只显示合价，不显示工程量和单价
                    item.setText(6, "")  # 工程量
                    item.setText(7, "")  # 综合单价
                    item.setText(8, "")  # 人工单价
                    item.setText(9, "")  # 主材单价
                    item.setText(10, "")  # 损耗率
                    item.setText(11, "")  # 辅材单价
                    item.setText(12, "")  # 机械单价
                    item.setText(13, "")  # 其他单价
                else:
                    # 叶子节点显示所有值
                    item.setText(6, self._format_number(detail_item.quantity))
                    item.setText(7, self._format_number(detail_item.unit_price))
                    item.setText(8, self._format_number(detail_item.labor_unit_price))
                    item.setText(9, self._format_number(detail_item.material_unit_price))
                    item.setText(10, self._format_number(detail_item.material_loss_rate))
                    item.setText(11, self._format_number(detail_item.auxiliary_unit_price))
                    item.setText(12, self._format_number(detail_item.machine_unit_price))
                    item.setText(13, self._format_number(detail_item.other_unit_price))
                
                # 所有节点都显示合价
                item.setText(14, self._format_number(detail_item.total_price))
                item.setText(15, self._format_number(detail_item.labor_total))
                item.setText(16, self._format_number(detail_item.material_total))
                item.setText(17, self._format_number(detail_item.auxiliary_total))
                item.setText(18, self._format_number(detail_item.machine_total))
                item.setText(19, self._format_number(detail_item.other_total))
                item.setText(20, self._format_number(detail_item.management_total))
                item.setText(21, self._format_number(detail_item.tax_total))
                item.setText(22, self._format_number(detail_item.comprehensive_total))
        
        # 从顶层节点开始更新
        for i in range(self.detail_tree.topLevelItemCount()):
            update_item_display(self.detail_tree.topLevelItem(i))

    def on_import_excel(self):
        """导入Excel明细"""
        if not self.current_bidding_id:
            MessageDialog.warning(self, "提示", "请先选择投标")
            return

        excel_path = self.excel_selector.get_selected_file()
        if not excel_path:
            MessageDialog.warning(self, "提示", "请先选择Excel文件")
            return

        # 打开可视化导入对话框，传入汇总项名称作为顶级行名称
        summary_name = self.current_summary_item.name if self.current_summary_item else ""
        # 检测当前主题
        is_dark = self._is_dark_theme()
        dialog = VisualImportDialog(excel_path, self.current_bidding_name, summary_name, self, is_dark)
        if dialog.exec() == VisualImportDialog.Accepted:
            imported_data = dialog.get_imported_data()
            if imported_data:
                self.build_tree_from_imported_data(imported_data)
                # 不再显示中间弹窗，进度条和成功提示已在导入对话框中完成

    def on_export_excel(self):
        """导出明细到Excel"""
        if not self.current_bidding_id:
            MessageDialog.warning(self, "提示", "请先选择投标")
            return

        # 收集所有节点数据
        all_items_data = []

        def collect_items(item: QTreeWidgetItem, level: int = 1):
            detail_item = item.data(0, Qt.UserRole)
            if detail_item:
                item_data = {
                    'level': level,
                    'sequence': detail_item.sequence,
                    'name': detail_item.name,
                    'specification': detail_item.specification,
                    'description': detail_item.description,
                    'unit': detail_item.unit,
                    'quantity': detail_item.quantity,
                    'unit_price': detail_item.unit_price,
                    'labor_unit_price': detail_item.labor_unit_price,
                    'material_unit_price': detail_item.material_unit_price,
                    'material_loss_rate': detail_item.material_loss_rate,
                    'auxiliary_unit_price': detail_item.auxiliary_unit_price,
                    'machine_unit_price': detail_item.machine_unit_price,
                    'other_unit_price': detail_item.other_unit_price,
                    'total_price': detail_item.total_price,
                    'labor_total': detail_item.labor_total,
                    'material_total': detail_item.material_total,
                    'auxiliary_total': detail_item.auxiliary_total,
                    'machine_total': detail_item.machine_total,
                    'other_total': detail_item.other_total,
                    'management_total': detail_item.management_total,
                    'tax_total': detail_item.tax_total,
                    'comprehensive_total': detail_item.comprehensive_total,
                    'remark': detail_item.remark,
                }
                all_items_data.append(item_data)

            # 递归处理子节点
            for i in range(item.childCount()):
                collect_items(item.child(i), level + 1)

        # 从顶层节点开始收集
        for i in range(self.detail_tree.topLevelItemCount()):
            collect_items(self.detail_tree.topLevelItem(i))

        if not all_items_data:
            MessageDialog.warning(self, "提示", "没有数据可导出")
            return

        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出Excel",
            f"{self.current_bidding_name}_{self.current_summary_item.name if self.current_summary_item else '明细'}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "投标明细"

            # 定义表头
            headers = ['序号', '分部分项工程名称', '规格型号', '项目特征描述', '单位',
                      '工程量', '综合单价', '人工单价', '主材单价', '主材损耗率%',
                      '辅材单价', '机械单价', '其他单价', '合价', '人工合价',
                      '主材合价', '辅材合价', '机械合价', '其他合价', '管理费',
                      '税金', '综合合价', '备注']

            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF', size=11)

            # 写入数据
            for row_idx, item_data in enumerate(all_items_data, 2):
                ws.cell(row=row_idx, column=1, value=item_data['sequence'])
                ws.cell(row=row_idx, column=2, value=item_data['name'])
                ws.cell(row=row_idx, column=3, value=item_data['specification'])
                ws.cell(row=row_idx, column=4, value=item_data['description'])
                ws.cell(row=row_idx, column=5, value=item_data['unit'])
                ws.cell(row=row_idx, column=6, value=item_data['quantity'])
                ws.cell(row=row_idx, column=7, value=item_data['unit_price'])
                ws.cell(row=row_idx, column=8, value=item_data['labor_unit_price'])
                ws.cell(row=row_idx, column=9, value=item_data['material_unit_price'])
                ws.cell(row=row_idx, column=10, value=item_data['material_loss_rate'])
                ws.cell(row=row_idx, column=11, value=item_data['auxiliary_unit_price'])
                ws.cell(row=row_idx, column=12, value=item_data['machine_unit_price'])
                ws.cell(row=row_idx, column=13, value=item_data['other_unit_price'])
                ws.cell(row=row_idx, column=14, value=item_data['total_price'])
                ws.cell(row=row_idx, column=15, value=item_data['labor_total'])
                ws.cell(row=row_idx, column=16, value=item_data['material_total'])
                ws.cell(row=row_idx, column=17, value=item_data['auxiliary_total'])
                ws.cell(row=row_idx, column=18, value=item_data['machine_total'])
                ws.cell(row=row_idx, column=19, value=item_data['other_total'])
                ws.cell(row=row_idx, column=20, value=item_data['management_total'])
                ws.cell(row=row_idx, column=21, value=item_data['tax_total'])
                ws.cell(row=row_idx, column=22, value=item_data['comprehensive_total'])
                ws.cell(row=row_idx, column=23, value=item_data['remark'])

                # 设置数值格式和对齐
                for col in range(6, 23):  # 数值列
                    cell = ws.cell(row=row_idx, column=col)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')

                # 序号列居中
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')

            # 调整列宽
            ws.column_dimensions['A'].width = 10  # 序号
            ws.column_dimensions['B'].width = 40  # 名称
            ws.column_dimensions['C'].width = 20  # 规格型号
            ws.column_dimensions['D'].width = 40  # 描述
            ws.column_dimensions['E'].width = 10  # 单位
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
                ws.column_dimensions[col].width = 12

            # 保存文件
            wb.save(file_path)
            MessageDialog.information(self, "成功", f"成功导出 {len(all_items_data)} 条明细数据到:\n{file_path}")

        except Exception as e:
            MessageDialog.error(self, "导出失败", f"导出Excel失败:\n{str(e)}")

    def reset_all_to_level_one(self):
        """将所有节点重置为级别1"""
        if not self.current_bidding_id:
            MessageDialog.warning(self, "提示", "请先选择投标")
            return

        # 收集所有节点数据
        all_items_data = []

        def collect_items(item: QTreeWidgetItem):
            detail_item = item.data(0, Qt.UserRole)
            if detail_item:
                item_data = {
                    'sequence': detail_item.sequence,
                    'name': detail_item.name,
                    'description': detail_item.description,
                    'unit': detail_item.unit,
                    'quantity': detail_item.quantity,
                    'unit_price': detail_item.unit_price,
                    'total_price': detail_item.total_price,
                    'remark': detail_item.remark,
                }
                all_items_data.append(item_data)

            for i in range(item.childCount()):
                collect_items(item.child(i))

        for i in range(self.detail_tree.topLevelItemCount()):
            collect_items(self.detail_tree.topLevelItem(i))

        if not all_items_data:
            MessageDialog.information(self, "提示", "没有数据需要重置")
            return

        reply = MessageDialog.question(
            self, "确认重置",
            f"确定要将所有 {len(all_items_data)} 个节点重置为级别1吗？",
            MessageDialog.Yes | MessageDialog.No,
            MessageDialog.No
        )

        if reply != MessageDialog.Yes:
            return

        self.detail_tree.clear()

        for i, data in enumerate(all_items_data):
            new_item = DetailItem()
            new_item.sequence = str(i + 1)
            new_item.name = data['name']
            new_item.specification = data.get('specification', '')
            new_item.description = data.get('description', '')
            new_item.unit = data.get('unit', '')
            new_item.quantity = data.get('quantity', 0)
            new_item.unit_price = data.get('unit_price', 0)
            new_item.total_price = data.get('total_price', 0)
            new_item.remark = data.get('remark', '')
            new_item.level = 1

            tree_item = QTreeWidgetItem(self.detail_tree)
            tree_item.setText(0, "")  # 缩进列（空）
            tree_item.setText(1, new_item.sequence)
            tree_item.setText(2, new_item.name)
            tree_item.setText(3, new_item.specification)
            tree_item.setText(4, new_item.description)  # 项目特征描述
            tree_item.setText(5, new_item.unit)
            tree_item.setText(6, f"{new_item.quantity:,.2f}")
            tree_item.setText(7, f"{new_item.unit_price:,.2f}")
            tree_item.setText(8, f"{new_item.labor_unit_price:,.2f}")
            tree_item.setText(9, f"{new_item.material_unit_price:,.2f}")
            tree_item.setText(10, f"{new_item.material_loss_rate:,.2f}")
            tree_item.setText(11, f"{new_item.auxiliary_unit_price:,.2f}")
            tree_item.setText(12, f"{new_item.machine_unit_price:,.2f}")
            tree_item.setText(13, f"{new_item.other_unit_price:,.2f}")
            tree_item.setText(14, f"{new_item.total_price:,.2f}")
            tree_item.setText(15, f"{new_item.labor_total:,.2f}")
            tree_item.setText(16, f"{new_item.material_total:,.2f}")
            tree_item.setText(17, f"{new_item.auxiliary_total:,.2f}")
            tree_item.setText(18, f"{new_item.machine_total:,.2f}")
            tree_item.setText(19, f"{new_item.other_total:,.2f}")
            tree_item.setText(20, f"{new_item.management_total:,.2f}")
            tree_item.setText(21, f"{new_item.tax_total:,.2f}")
            tree_item.setText(22, f"{new_item.comprehensive_total:,.2f}")
            tree_item.setText(23, new_item.remark)
            tree_item.setData(0, Qt.UserRole, new_item)

            tree_item.setTextAlignment(1, Qt.AlignCenter)  # 序号列居中
            for col in range(6, 23):  # 数值列右对齐
                tree_item.setTextAlignment(col, Qt.AlignRight | Qt.AlignVCenter)

        self.calculate_total()
        MessageDialog.information(self, "成功", f"已重置 {len(all_items_data)} 个节点为级别1")

    def show_tip(self):
        """显示提示信息"""
        self.tip_label.show()
        self.content_widget.hide()

    def show_content(self):
        """显示内容区域"""
        self.tip_label.hide()
        self.content_widget.show()

    def on_column_settings(self):
        """打开列显示设置对话框"""
        dialog = ColumnSettingsDialog(self.visible_columns, self)
        if dialog.exec() == QDialog.Accepted:
            self.visible_columns = dialog.get_visible_columns()
            self.apply_column_visibility()

    def apply_column_visibility(self):
        """应用列可见性设置"""
        if not hasattr(self, 'detail_tree'):
            return
            
        header = self.detail_tree.header()
        
        # 缩进列（第0列）始终显示，专门用于层级缩进
        # 宽度设置为120px，可以容纳约5-6级缩进（每级20px + 图标宽度）
        self.detail_tree.setColumnHidden(0, False)
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        self.detail_tree.setColumnWidth(0, 120)  # 足够的宽度显示多级缩进
        
        # 序号列（第1列）始终显示
        self.detail_tree.setColumnHidden(1, False)
        header.setSectionResizeMode(1, QHeaderView.Fixed)
        self.detail_tree.setColumnWidth(1, 60)
        
        # 隐藏其他未选中的列（从第2列开始）
        for col_idx in range(2, 24):
            if col_idx in self.visible_columns:
                self.detail_tree.setColumnHidden(col_idx, False)
            else:
                self.detail_tree.setColumnHidden(col_idx, True)
        
        # 设置列宽模式
        # 分部分项工程名称(2)、规格型号(3)、项目特征描述(4)固定宽度180
        fixed_width_columns = [2, 3, 4]
        for col_idx in range(24):
            if col_idx in self.visible_columns or col_idx in [0, 1]:  # 包括缩进列和序号列
                if col_idx in fixed_width_columns:  # 固定宽度列
                    header.setSectionResizeMode(col_idx, QHeaderView.Fixed)
                    self.detail_tree.setColumnWidth(col_idx, 240)
                else:
                    header.setSectionResizeMode(col_idx, QHeaderView.ResizeToContents)

        # 设置所有表头文字居中
        for col_idx in range(24):
            header_item = self.detail_tree.headerItem()
            if header_item:
                header_item.setTextAlignment(col_idx, Qt.AlignCenter)

    def on_add_detail(self):
        """添加明细"""
        if not self.current_bidding_id:
            MessageDialog.warning(self, "提示", "请先选择投标")
            return

        current = self.detail_tree.currentItem()
        
        tree_item = QTreeWidgetItem()
        detail_item = DetailItem()
        
        if current:
            # 作为子节点添加
            current.addChild(tree_item)
            parent_detail = current.data(0, Qt.UserRole)
            detail_item.level = parent_detail.level + 1 if parent_detail else 2
            detail_item.sequence = f"{parent_detail.sequence}.1" if parent_detail else "1"
            current.setExpanded(True)
        else:
            # 添加到根
            self.detail_tree.addTopLevelItem(tree_item)
            detail_item.level = 1
            detail_item.sequence = str(self.detail_tree.topLevelItemCount())
        
        detail_item.name = "新明细项"
        
        tree_item.setText(0, "")  # 缩进列（空）
        tree_item.setText(1, detail_item.sequence)
        tree_item.setText(2, detail_item.name)
        tree_item.setText(3, "")  # 规格型号
        tree_item.setText(4, "")  # 项目特征描述
        tree_item.setText(5, "m²")  # 单位
        tree_item.setText(6, "0.00")  # 工程量
        tree_item.setText(7, "0.00")  # 综合单价
        tree_item.setText(8, "0.00")  # 人工单价
        tree_item.setText(9, "0.00")  # 主材单价
        tree_item.setText(10, "0.00")  # 主材损耗%
        tree_item.setText(11, "0.00")  # 辅材单价
        tree_item.setText(12, "0.00")  # 机械单价
        tree_item.setText(13, "0.00")  # 其他单价
        tree_item.setText(14, "0.00")  # 合价(净)
        tree_item.setText(15, "0.00")  # 人工合价
        tree_item.setText(16, "0.00")  # 主材合价
        tree_item.setText(17, "0.00")  # 辅材合价
        tree_item.setText(18, "0.00")  # 机械合价
        tree_item.setText(19, "0.00")  # 其他合价
        tree_item.setText(20, "0.00")  # 管理费合计
        tree_item.setText(21, "0.00")  # 税金合价
        tree_item.setText(22, "0.00")  # 综合合价
        tree_item.setText(23, "")  # 备注
        tree_item.setData(0, Qt.UserRole, detail_item)

        tree_item.setTextAlignment(1, Qt.AlignCenter)  # 序号列居中
        for col in range(6, 23):  # 数值列右对齐
            tree_item.setTextAlignment(col, Qt.AlignRight | Qt.AlignVCenter)

        self.detail_tree.setCurrentItem(tree_item)
        self.calculate_total()

    def on_delete_detail(self):
        """删除明细"""
        selected_items = self.detail_tree.selectedItems()
        if not selected_items:
            MessageDialog.warning(self, "提示", "请先选择要删除的明细")
            return

        reply = MessageDialog.question(
            self, "确认删除",
            f"确定要删除选中的 {len(selected_items)} 个明细吗？",
            MessageDialog.Yes | MessageDialog.No,
            MessageDialog.No
        )

        if reply == MessageDialog.Yes:
            for item in selected_items:
                parent = item.parent()
                if parent:
                    parent.removeChild(item)
                else:
                    index = self.detail_tree.indexOfTopLevelItem(item)
                    self.detail_tree.takeTopLevelItem(index)
            
            self.calculate_total()

    def on_save_detail(self):
        """保存明细到数据库"""
        if not self.current_bidding_id:
            MessageDialog.warning(self, "提示", "请先选择投标")
            return

        if not self.current_summary_item:
            MessageDialog.warning(self, "提示", "请先选择汇总表项目")
            return

        # 从UI收集数据
        detail = self.collect_data_from_ui()
        if not detail:
            MessageDialog.warning(self, "提示", "没有数据需要保存")
            return

        # 调试信息
        summary_item_id = getattr(self.current_summary_item, 'id', 0)
        print(f"[保存明细] 投标ID: {detail.bidding_id}, 汇总项ID: {summary_item_id}")
        print(f"[保存明细] 明细项目数: {len(detail.items)}")
        if detail.items:
            print(f"[保存明细] 第一个项目: {detail.items[0].name}, 子节点数: {len(detail.items[0].children)}")

        # 调用Logic层保存
        try:
            success, result = self.detail_manager.save_detail(detail, update_summary=True)
            if success:
                self.current_detail_id = result
                print(f"[保存明细] 保存成功，明细ID: {result}")
                
                # 获取汇总表行信息和版本信息
                summary_item_name = getattr(self.current_summary_item, 'name', '未知')
                summary_item_code = getattr(self.current_summary_item, 'sequence', '')
                version = detail.version or "V1.0"
                
                # 构建提示信息
                item_info = f"{summary_item_code} {summary_item_name}" if summary_item_code else summary_item_name
                MessageDialog.information(self, "保存成功", 
                    f"明细数据保存成功\n"
                    f"汇总项: {item_info}\n"
                    f"版本: {version}\n"
                    f"共 {len(detail.items)} 条明细记录")
                
                # 通知父页面刷新汇总表（带红色字体标记）
                print(f"[保存明细] 准备刷新汇总表，parent_page={self.parent_page}")
                if self.parent_page and hasattr(self.parent_page, 'refresh_summary_with_detail'):
                    print(f"[保存明细] 调用 refresh_summary_with_detail")
                    self.parent_page.refresh_summary_with_detail(self.current_summary_item)
                elif self.parent_page and hasattr(self.parent_page, 'refresh_summary'):
                    print(f"[保存明细] 调用 refresh_summary")
                    self.parent_page.refresh_summary()
            else:
                print(f"[保存明细] 保存失败: {result}")
                MessageDialog.warning(self, "保存失败", str(result))
        except Exception as e:
            print(f"[保存明细] 异常: {e}")
            import traceback
            traceback.print_exc()
            MessageDialog.critical(self, "错误", f"保存失败：{str(e)}")

    def collect_data_from_ui(self) -> BiddingDetail:
        """从UI收集数据构建BiddingDetail对象"""
        detail = BiddingDetail()
        detail.id = self.current_detail_id
        detail.bidding_id = self.current_bidding_id
        detail.summary_item_id = getattr(self.current_summary_item, 'id', 0)
        detail.version = "V1.0"  # 默认版本，后续可支持版本选择
        detail.created_by = ""  # 可从登录用户获取
        detail.remark = ""

        # 递归收集树形数据
        detail.items = self._collect_tree_items()

        return detail

    def _collect_tree_items(self) -> list:
        """递归收集树形控件中的数据"""
        items = []

        def collect_recursive(tree_item: QTreeWidgetItem, parent_item: DetailItemModel = None) -> DetailItemModel:
            """递归收集单个节点及其子节点"""
            # 获取节点数据
            detail_item_ui = tree_item.data(0, Qt.UserRole)
            if not detail_item_ui:
                return None

            # 创建数据模型对象
            detail_item = DetailItemModel()
            detail_item.sequence = detail_item_ui.sequence
            detail_item.name = detail_item_ui.name
            detail_item.specification = detail_item_ui.specification
            detail_item.description = detail_item_ui.description
            detail_item.unit = detail_item_ui.unit
            detail_item.quantity = detail_item_ui.quantity
            detail_item.unit_price = detail_item_ui.unit_price
            detail_item.labor_price = detail_item_ui.labor_unit_price
            detail_item.main_material_price = detail_item_ui.material_unit_price
            detail_item.main_material_loss_rate = detail_item_ui.material_loss_rate
            detail_item.aux_material_price = detail_item_ui.auxiliary_unit_price
            detail_item.machinery_price = detail_item_ui.machine_unit_price
            detail_item.other_price = detail_item_ui.other_unit_price
            detail_item.total_price = detail_item_ui.total_price
            detail_item.labor_total = detail_item_ui.labor_total
            detail_item.material_total = detail_item_ui.material_total
            detail_item.auxiliary_total = detail_item_ui.auxiliary_total
            detail_item.machine_total = detail_item_ui.machine_total
            detail_item.other_total = detail_item_ui.other_total
            detail_item.management_total = detail_item_ui.management_total
            detail_item.tax_total = detail_item_ui.tax_total
            detail_item.comprehensive_total = detail_item_ui.comprehensive_total
            detail_item.remark = detail_item_ui.remark
            detail_item.level = detail_item_ui.level
            detail_item.parent_id = parent_item.id if parent_item else None

            # 递归收集子节点
            for i in range(tree_item.childCount()):
                child_tree_item = tree_item.child(i)
                child_detail_item = collect_recursive(child_tree_item, detail_item)
                if child_detail_item:
                    detail_item.children.append(child_detail_item)

            return detail_item

        # 遍历顶层节点
        for i in range(self.detail_tree.topLevelItemCount()):
            tree_item = self.detail_tree.topLevelItem(i)
            detail_item = collect_recursive(tree_item, None)
            if detail_item:
                items.append(detail_item)

        return items

    def load_bidding_data(self, bidding_id: int):
        """加载投标明细数据 - 从数据库加载或显示提示"""
        self.current_bidding_id = bidding_id
        
        # 获取投标信息
        bidding = self.bidding_manager.get_bidding(bidding_id)
        if bidding:
            self.current_bidding_code = bidding.bidding_code
            self.current_bidding_name = bidding.bidding_name
        
        # 更新Excel文件选择器目录
        excel_dir = self.get_excel_directory()
        self.excel_selector.set_directory(excel_dir)
        
        # 显示提示信息，因为明细需要从汇总表双击行来加载具体数据
        self.show_tip()

    def on_item_changed(self, item: QTreeWidgetItem, column: int):
        """节点编辑事件"""
        detail_item = item.data(0, Qt.UserRole)
        if not detail_item:
            return

        # 文本列
        if column == 1:
            detail_item.sequence = item.text(1)
        elif column == 2:
            detail_item.name = item.text(2)
        elif column == 3:
            detail_item.specification = item.text(3)
        elif column == 4:
            detail_item.description = item.text(4)
        elif column == 5:
            detail_item.unit = item.text(5)
        # 数值列
        elif column == 6:
            try:
                detail_item.quantity = float(item.text(6).replace(',', '') or 0)
            except:
                pass
        elif column == 7:
            try:
                detail_item.unit_price = float(item.text(7).replace(',', '') or 0)
            except:
                pass
        elif column == 8:
            try:
                detail_item.labor_unit_price = float(item.text(8).replace(',', '') or 0)
            except:
                pass
        elif column == 9:
            try:
                detail_item.material_unit_price = float(item.text(9).replace(',', '') or 0)
            except:
                pass
        elif column == 10:
            try:
                detail_item.material_loss_rate = float(item.text(10).replace(',', '') or 0)
            except:
                pass
        elif column == 11:
            try:
                detail_item.auxiliary_unit_price = float(item.text(11).replace(',', '') or 0)
            except:
                pass
        elif column == 12:
            try:
                detail_item.machine_unit_price = float(item.text(12).replace(',', '') or 0)
            except:
                pass
        elif column == 13:
            try:
                detail_item.other_unit_price = float(item.text(13).replace(',', '') or 0)
            except:
                pass
        elif column == 15:
            try:
                detail_item.labor_total = float(item.text(15).replace(',', '') or 0)
            except:
                pass
        elif column == 16:
            try:
                detail_item.material_total = float(item.text(16).replace(',', '') or 0)
            except:
                pass
        elif column == 17:
            try:
                detail_item.auxiliary_total = float(item.text(17).replace(',', '') or 0)
            except:
                pass
        elif column == 18:
            try:
                detail_item.machine_total = float(item.text(18).replace(',', '') or 0)
            except:
                pass
        elif column == 19:
            try:
                detail_item.other_total = float(item.text(19).replace(',', '') or 0)
            except:
                pass
        elif column == 20:
            try:
                detail_item.management_total = float(item.text(20).replace(',', '') or 0)
            except:
                pass
        elif column == 21:
            try:
                detail_item.tax_total = float(item.text(21).replace(',', '') or 0)
            except:
                pass
        elif column == 22:
            try:
                detail_item.comprehensive_total = float(item.text(22).replace(',', '') or 0)
            except:
                pass
        elif column == 23:
            detail_item.remark = item.text(23)

        # 重新计算合价（当工程量或单价发生变化时）
        if column in [6, 7, 8, 9, 10, 11, 12, 13]:
            detail_item.calculate_total()
            # 更新所有合价列的显示
            item.setText(14, self._format_number(detail_item.total_price))
            item.setText(15, self._format_number(detail_item.labor_total))
            item.setText(16, self._format_number(detail_item.material_total))
            item.setText(17, self._format_number(detail_item.auxiliary_total))
            item.setText(18, self._format_number(detail_item.machine_total))
            item.setText(19, self._format_number(detail_item.other_total))
            item.setText(22, self._format_number(detail_item.comprehensive_total))

            # 更新父级节点的汇总值
            self._update_parent_totals(item)

        self.calculate_total()

        # 触发自动保存（防抖，延迟500ms）
        self._trigger_auto_save()

    def _update_parent_totals(self, item: QTreeWidgetItem):
        """更新父级节点的汇总值（向上递归）"""
        parent_item = item.parent()
        if not parent_item:
            return
        
        # 获取父节点的数据
        parent_data = parent_item.data(0, Qt.UserRole)
        if not parent_data:
            return
        
        # 重新计算父节点的汇总值
        total_price = 0.0
        labor_total = 0.0
        material_total = 0.0
        auxiliary_total = 0.0
        machine_total = 0.0
        other_total = 0.0
        management_total = 0.0
        tax_total = 0.0
        comprehensive_total = 0.0
        
        # 累加所有子节点的值
        for i in range(parent_item.childCount()):
            child_item = parent_item.child(i)
            child_data = child_item.data(0, Qt.UserRole)
            if child_data:
                total_price += child_data.total_price
                labor_total += child_data.labor_total
                material_total += child_data.material_total
                auxiliary_total += child_data.auxiliary_total
                machine_total += child_data.machine_total
                other_total += child_data.other_total
                management_total += child_data.management_total
                tax_total += child_data.tax_total
                comprehensive_total += child_data.comprehensive_total
        
        # 更新父节点的值
        parent_data.total_price = total_price
        parent_data.labor_total = labor_total
        parent_data.material_total = material_total
        parent_data.auxiliary_total = auxiliary_total
        parent_data.machine_total = machine_total
        parent_data.other_total = other_total
        parent_data.management_total = management_total
        parent_data.tax_total = tax_total
        parent_data.comprehensive_total = comprehensive_total
        
        # 更新父节点的显示
        parent_item.setText(14, self._format_number(total_price))
        parent_item.setText(15, self._format_number(labor_total))
        parent_item.setText(16, self._format_number(material_total))
        parent_item.setText(17, self._format_number(auxiliary_total))
        parent_item.setText(18, self._format_number(machine_total))
        parent_item.setText(19, self._format_number(other_total))
        parent_item.setText(20, self._format_number(management_total))
        parent_item.setText(21, self._format_number(tax_total))
        parent_item.setText(22, self._format_number(comprehensive_total))
        
        # 递归更新上级父节点
        self._update_parent_totals(parent_item)

    def _trigger_auto_save(self):
        """触发自动保存（防抖）"""
        if not self.current_bidding_id or not self.current_summary_item:
            return

        self._pending_save = True
        # 重新启动定时器（500ms后保存）
        self.auto_save_timer.stop()
        self.auto_save_timer.start(500)

    def _auto_save(self):
        """执行自动保存"""
        if not self._pending_save:
            return

        try:
            # 收集数据
            detail = self.collect_data_from_ui()
            if not detail.items:
                return

            # 调试：打印第一个顶层节点的数据
            if detail.items:
                first_item = detail.items[0]
                print(f"[自动保存调试] 第一个节点: {first_item.name}")
                print(f"[自动保存调试]   total_price={first_item.total_price}, labor_total={first_item.labor_total}")
                print(f"[自动保存调试]   子节点数: {len(first_item.children)}")
                if first_item.children:
                    for child in first_item.children:
                        print(f"[自动保存调试]     子节点: {child.name}, total_price={child.total_price}")

            # 保存到数据库（不显示成功提示，避免打扰用户）
            success, result = self.detail_manager.save_detail(detail, update_summary=True)
            if success:
                self.current_detail_id = result
                self._pending_save = False
                print(f"[自动保存] 成功，明细ID: {result}")

                # 通知父页面刷新汇总表
                if self.parent_page and hasattr(self.parent_page, 'refresh_summary_with_detail'):
                    self.parent_page.refresh_summary_with_detail(self.current_summary_item)
            else:
                print(f"[自动保存] 失败: {result}")
        except Exception as e:
            print(f"[自动保存] 异常: {e}")

    def calculate_total(self):
        """计算总价 - 只累加顶层节点的合价（顶层节点已包含所有子节点的合价）"""
        total = 0.0

        # 只累加顶层节点的 total_price
        # 因为父节点的 total_price 已经通过 _calculate_parent_totals 计算为所有子节点的总和
        for i in range(self.detail_tree.topLevelItemCount()):
            item = self.detail_tree.topLevelItem(i)
            detail_item = item.data(0, Qt.UserRole)
            if detail_item:
                total += detail_item.total_price

        self.total_label.setText(f"¥ {total:,.2f}")
        return total

    def _is_dark_theme(self) -> bool:
        """检测当前是否为深色主题"""
        # 获取应用程序的样式表
        app = QApplication.instance()
        if app:
            # 检查主窗口的背景色
            from PySide6.QtWidgets import QMainWindow
            main_window = None
            for widget in app.topLevelWidgets():
                if isinstance(widget, QMainWindow):
                    main_window = widget
                    break
            
            if main_window:
                # 获取背景色
                palette = main_window.palette()
                bg_color = palette.color(main_window.backgroundRole())
                # 如果背景色较暗（RGB平均值 < 128），认为是深色主题
                brightness = (bg_color.red() + bg_color.green() + bg_color.blue()) / 3
                return brightness < 128
        
        # 默认返回False（浅色主题）
        return False

    def expand_all(self):
        """展开全部"""
        self.detail_tree.expandAll()

    def collapse_all(self):
        """折叠全部"""
        self.detail_tree.collapseAll()

    def on_item_double_clicked(self, item: QTreeWidgetItem, column: int):
        """双击单元格打开行编辑对话框"""
        # 获取当前行的数据
        detail_item = item.data(0, Qt.UserRole)
        if not detail_item:
            return

        # 打开行编辑对话框
        dialog = RowEditDialog(detail_item, self)
        if dialog.exec() == QDialog.Accepted:
            # 获取编辑后的数据
            updated_data = dialog.get_updated_data()

            # 更新数据模型
            detail_item.sequence = updated_data['sequence']
            detail_item.name = updated_data['name']
            detail_item.specification = updated_data['specification']
            detail_item.description = updated_data['description']
            detail_item.unit = updated_data['unit']
            detail_item.quantity = updated_data['quantity']
            detail_item.unit_price = updated_data['unit_price']
            detail_item.labor_unit_price = updated_data['labor_unit_price']
            detail_item.material_unit_price = updated_data['material_unit_price']
            detail_item.material_loss_rate = updated_data['material_loss_rate']
            detail_item.auxiliary_unit_price = updated_data['auxiliary_unit_price']
            detail_item.machine_unit_price = updated_data['machine_unit_price']
            detail_item.other_unit_price = updated_data['other_unit_price']

            # 重新计算合价
            detail_item.calculate_total()

            # 更新树形控件显示
            self._update_tree_item_display(item, detail_item)

            # 更新父级节点的汇总值
            self._update_parent_totals(item)

            # 更新总价
            self.calculate_total()

            # 触发自动保存
            self._trigger_auto_save()

    def _update_tree_item_display(self, tree_item: QTreeWidgetItem, detail_item: DetailItem):
        """更新树节点的显示"""
        tree_item.setText(1, detail_item.sequence)
        tree_item.setText(2, detail_item.name)
        tree_item.setText(3, detail_item.specification)
        tree_item.setText(4, detail_item.description)
        tree_item.setText(5, detail_item.unit)
        tree_item.setText(6, self._format_number(detail_item.quantity))
        tree_item.setText(7, self._format_number(detail_item.unit_price))
        tree_item.setText(8, self._format_number(detail_item.labor_unit_price))
        tree_item.setText(9, self._format_number(detail_item.material_unit_price))
        tree_item.setText(10, self._format_number(detail_item.material_loss_rate))
        tree_item.setText(11, self._format_number(detail_item.auxiliary_unit_price))
        tree_item.setText(12, self._format_number(detail_item.machine_unit_price))
        tree_item.setText(13, self._format_number(detail_item.other_unit_price))
        tree_item.setText(14, self._format_number(detail_item.total_price))
        tree_item.setText(15, self._format_number(detail_item.labor_total))
        tree_item.setText(16, self._format_number(detail_item.material_total))
        tree_item.setText(17, self._format_number(detail_item.auxiliary_total))
        tree_item.setText(18, self._format_number(detail_item.machine_total))
        tree_item.setText(19, self._format_number(detail_item.other_total))
        tree_item.setText(20, self._format_number(detail_item.management_total))
        tree_item.setText(21, self._format_number(detail_item.tax_total))
        tree_item.setText(22, self._format_number(detail_item.comprehensive_total))

    def show_context_menu(self, position):
        """显示右键菜单"""
        menu = QMenu(self)
        menu.setObjectName("contextMenu")

        # 层级操作
        level_up_action = QAction("⬆️ 提升层级", self)
        level_up_action.triggered.connect(self.move_item_level_up)
        menu.addAction(level_up_action)

        level_down_action = QAction("⬇️ 降低层级", self)
        level_down_action.triggered.connect(self.move_item_level_down)
        menu.addAction(level_down_action)

        menu.addSeparator()

        # 添加操作
        add_child_action = QAction("➕ 添加子级", self)
        add_child_action.triggered.connect(self.on_add_detail)
        menu.addAction(add_child_action)

        menu.addSeparator()

        # 删除操作
        delete_action = QAction("🗑️ 删除", self)
        delete_action.triggered.connect(self.on_delete_detail)
        menu.addAction(delete_action)

        menu.exec(self.detail_tree.viewport().mapToGlobal(position))

    def move_item_level_up(self):
        """提升节点层级"""
        selected_items = self.detail_tree.selectedItems()
        if not selected_items:
            MessageDialog.warning(self, "提示", "请先选择要提升层级的节点")
            return

        for item in selected_items:
            parent = item.parent()
            if not parent:
                continue

            grandparent = parent.parent()
            new_item = QTreeWidgetItem()
            
            # 复制数据
            detail_item = item.data(0, Qt.UserRole)
            new_detail = DetailItem()
            new_detail.sequence = detail_item.sequence
            new_detail.name = detail_item.name
            new_detail.specification = detail_item.specification
            new_detail.description = detail_item.description
            new_detail.unit = detail_item.unit
            new_detail.quantity = detail_item.quantity
            new_detail.unit_price = detail_item.unit_price
            new_detail.labor_unit_price = detail_item.labor_unit_price
            new_detail.material_unit_price = detail_item.material_unit_price
            new_detail.auxiliary_unit_price = detail_item.auxiliary_unit_price
            new_detail.machine_unit_price = detail_item.machine_unit_price
            new_detail.other_unit_price = detail_item.other_unit_price
            new_detail.total_price = detail_item.total_price
            new_detail.labor_total = detail_item.labor_total
            new_detail.material_total = detail_item.material_total
            new_detail.auxiliary_total = detail_item.auxiliary_total
            new_detail.machine_total = detail_item.machine_total
            new_detail.other_total = detail_item.other_total
            new_detail.management_total = detail_item.management_total
            new_detail.tax_total = detail_item.tax_total
            new_detail.comprehensive_total = detail_item.comprehensive_total
            new_detail.remark = detail_item.remark
            new_detail.level = max(1, detail_item.level - 1)
            
            # 复制显示
            for col in range(24):
                new_item.setText(col, item.text(col))
                new_item.setFont(col, item.font(col))
            new_item.setData(0, Qt.UserRole, new_detail)

            if not grandparent:
                index = self.detail_tree.indexOfTopLevelItem(parent)
                self.detail_tree.insertTopLevelItem(index + 1, new_item)
            else:
                parent_index = grandparent.indexOfChild(parent)
                grandparent.insertChild(parent_index + 1, new_item)
            
            parent.removeChild(item)

        self.calculate_total()

    def move_item_level_down(self):
        """降低节点层级"""
        selected_items = self.detail_tree.selectedItems()
        if not selected_items:
            MessageDialog.warning(self, "提示", "请先选择要降低层级的节点")
            return

        for item in selected_items:
            parent = item.parent()
            
            if not parent:
                # 根节点，找前一个根节点
                index = self.detail_tree.indexOfTopLevelItem(item)
                if index > 0:
                    prev_sibling = self.detail_tree.topLevelItem(index - 1)
                    
                    new_item = QTreeWidgetItem()
                    detail_item = item.data(0, Qt.UserRole)
                    new_detail = DetailItem()
                    new_detail.sequence = detail_item.sequence
                    new_detail.name = detail_item.name
                    new_detail.specification = detail_item.specification
                    new_detail.description = detail_item.description
                    new_detail.unit = detail_item.unit
                    new_detail.quantity = detail_item.quantity
                    new_detail.unit_price = detail_item.unit_price
                    new_detail.labor_unit_price = detail_item.labor_unit_price
                    new_detail.material_unit_price = detail_item.material_unit_price
                    new_detail.auxiliary_unit_price = detail_item.auxiliary_unit_price
                    new_detail.machine_unit_price = detail_item.machine_unit_price
                    new_detail.other_unit_price = detail_item.other_unit_price
                    new_detail.total_price = detail_item.total_price
                    new_detail.labor_total = detail_item.labor_total
                    new_detail.material_total = detail_item.material_total
                    new_detail.auxiliary_total = detail_item.auxiliary_total
                    new_detail.machine_total = detail_item.machine_total
                    new_detail.other_total = detail_item.other_total
                    new_detail.management_total = detail_item.management_total
                    new_detail.tax_total = detail_item.tax_total
                    new_detail.comprehensive_total = detail_item.comprehensive_total
                    new_detail.remark = detail_item.remark
                    new_detail.level = detail_item.level + 1
                    
                    for col in range(24):
                        new_item.setText(col, item.text(col))
                    new_item.setData(0, Qt.UserRole, new_detail)

                    prev_sibling.addChild(new_item)
                    prev_sibling.setExpanded(True)
                    self.detail_tree.takeTopLevelItem(index)
            else:
                # 找前一个兄弟
                index = parent.indexOfChild(item)
                if index > 0:
                    prev_sibling = parent.child(index - 1)

                    new_item = QTreeWidgetItem()
                    detail_item = item.data(0, Qt.UserRole)
                    new_detail = DetailItem()
                    new_detail.sequence = detail_item.sequence
                    new_detail.name = detail_item.name
                    new_detail.specification = detail_item.specification
                    new_detail.description = detail_item.description
                    new_detail.unit = detail_item.unit
                    new_detail.quantity = detail_item.quantity
                    new_detail.unit_price = detail_item.unit_price
                    new_detail.labor_unit_price = detail_item.labor_unit_price
                    new_detail.material_unit_price = detail_item.material_unit_price
                    new_detail.material_loss_rate = detail_item.material_loss_rate
                    new_detail.auxiliary_unit_price = detail_item.auxiliary_unit_price
                    new_detail.machine_unit_price = detail_item.machine_unit_price
                    new_detail.other_unit_price = detail_item.other_unit_price
                    new_detail.total_price = detail_item.total_price
                    new_detail.labor_total = detail_item.labor_total
                    new_detail.material_total = detail_item.material_total
                    new_detail.auxiliary_total = detail_item.auxiliary_total
                    new_detail.machine_total = detail_item.machine_total
                    new_detail.other_total = detail_item.other_total
                    new_detail.management_total = detail_item.management_total
                    new_detail.tax_total = detail_item.tax_total
                    new_detail.comprehensive_total = detail_item.comprehensive_total
                    new_detail.remark = detail_item.remark
                    new_detail.level = detail_item.level + 1

                    for col in range(24):
                        new_item.setText(col, item.text(col))
                    new_item.setData(0, Qt.UserRole, new_detail)
                    
                    prev_sibling.addChild(new_item)
                    prev_sibling.setExpanded(True)
                    parent.removeChild(item)

        self.calculate_total()


class RowEditDialog(QDialog):
    """行编辑对话框 - 编辑整行数据"""

    def __init__(self, detail_item: DetailItem, parent=None):
        super().__init__(parent)
        self.detail_item = detail_item
        self.setWindowTitle(f"编辑明细 - {detail_item.name}")
        self.setMinimumSize(600, 500)
        self.setup_ui()

    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # 基本信息组
        basic_group = QFrame()
        basic_layout = QGridLayout(basic_group)
        basic_layout.setSpacing(10)

        # 序号
        basic_layout.addWidget(QLabel("序号:"), 0, 0)
        self.sequence_edit = QLineEdit(self.detail_item.sequence)
        basic_layout.addWidget(self.sequence_edit, 0, 1)

        # 名称
        basic_layout.addWidget(QLabel("名称:"), 1, 0)
        self.name_edit = QLineEdit(self.detail_item.name)
        basic_layout.addWidget(self.name_edit, 1, 1)

        # 规格型号
        basic_layout.addWidget(QLabel("规格型号:"), 2, 0)
        self.specification_edit = QLineEdit(self.detail_item.specification)
        basic_layout.addWidget(self.specification_edit, 2, 1)

        # 项目特征描述
        basic_layout.addWidget(QLabel("项目特征描述:"), 3, 0)
        self.description_edit = QTextEdit(self.detail_item.description)
        self.description_edit.setMaximumHeight(80)
        basic_layout.addWidget(self.description_edit, 3, 1)

        # 单位
        basic_layout.addWidget(QLabel("单位:"), 4, 0)
        self.unit_edit = QLineEdit(self.detail_item.unit)
        basic_layout.addWidget(self.unit_edit, 4, 1)

        layout.addWidget(basic_group)

        # 工程量和单价组
        price_group = QFrame()
        price_layout = QGridLayout(price_group)
        price_layout.setSpacing(10)

        # 工程量
        price_layout.addWidget(QLabel("工程量:"), 0, 0)
        self.quantity_edit = QLineEdit(str(self.detail_item.quantity))
        price_layout.addWidget(self.quantity_edit, 0, 1)

        # 综合单价
        price_layout.addWidget(QLabel("综合单价:"), 1, 0)
        self.unit_price_edit = QLineEdit(str(self.detail_item.unit_price))
        price_layout.addWidget(self.unit_price_edit, 1, 1)

        # 人工单价
        price_layout.addWidget(QLabel("人工单价:"), 2, 0)
        self.labor_unit_price_edit = QLineEdit(str(self.detail_item.labor_unit_price))
        price_layout.addWidget(self.labor_unit_price_edit, 2, 1)

        # 主材单价
        price_layout.addWidget(QLabel("主材单价:"), 3, 0)
        self.material_unit_price_edit = QLineEdit(str(self.detail_item.material_unit_price))
        price_layout.addWidget(self.material_unit_price_edit, 3, 1)

        # 主材损耗率
        price_layout.addWidget(QLabel("主材损耗率(%)"), 4, 0)
        self.material_loss_rate_edit = QLineEdit(str(self.detail_item.material_loss_rate))
        price_layout.addWidget(self.material_loss_rate_edit, 4, 1)

        # 辅材单价
        price_layout.addWidget(QLabel("辅材单价:"), 5, 0)
        self.auxiliary_unit_price_edit = QLineEdit(str(self.detail_item.auxiliary_unit_price))
        price_layout.addWidget(self.auxiliary_unit_price_edit, 5, 1)

        # 机械单价
        price_layout.addWidget(QLabel("机械单价:"), 6, 0)
        self.machine_unit_price_edit = QLineEdit(str(self.detail_item.machine_unit_price))
        price_layout.addWidget(self.machine_unit_price_edit, 6, 1)

        # 其他单价
        price_layout.addWidget(QLabel("其他单价:"), 7, 0)
        self.other_unit_price_edit = QLineEdit(str(self.detail_item.other_unit_price))
        price_layout.addWidget(self.other_unit_price_edit, 7, 1)

        layout.addWidget(price_group)

        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        ok_btn = PrimaryPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        button_layout.addWidget(ok_btn)

        cancel_btn = PushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)

    def get_updated_data(self) -> dict:
        """获取更新后的数据"""
        return {
            'sequence': self.sequence_edit.text(),
            'name': self.name_edit.text(),
            'specification': self.specification_edit.text(),
            'description': self.description_edit.toPlainText(),
            'unit': self.unit_edit.text(),
            'quantity': float(self.quantity_edit.text() or 0),
            'unit_price': float(self.unit_price_edit.text() or 0),
            'labor_unit_price': float(self.labor_unit_price_edit.text() or 0),
            'material_unit_price': float(self.material_unit_price_edit.text() or 0),
            'material_loss_rate': float(self.material_loss_rate_edit.text() or 0),
            'auxiliary_unit_price': float(self.auxiliary_unit_price_edit.text() or 0),
            'machine_unit_price': float(self.machine_unit_price_edit.text() or 0),
            'other_unit_price': float(self.other_unit_price_edit.text() or 0),
        }
