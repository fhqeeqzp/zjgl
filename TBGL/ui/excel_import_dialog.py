"""
Excel导入对话框
用于从Excel文件导入汇总表数据，支持工作簿选择、列映射、数据预览
"""
import os
import re
from pathlib import Path
from typing import List, Dict, Optional, Tuple

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QComboBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QGroupBox, QFormLayout,
    QMessageBox, QFileDialog, QSplitter, QWidget,
    QCheckBox, QScrollArea, QGridLayout, QFrame,
    QProgressBar, QTextEdit
)
from PySide6.QtCore import Qt, Signal, QThread, QTimer
from PySide6.QtGui import QFont, QColor
from PySide6.QtWidgets import QApplication

from ui.fluent_widgets import PushButton, PrimaryPushButton


class ExcelLoadWorker(QThread):
    """Excel加载工作线程 - 避免阻塞UI"""

    loaded = Signal(list)  # 发送工作表名称列表
    sheet_data_loaded = Signal(list, list, int)  # 发送表头选项、样本数据、总行数
    error = Signal(str)  # 发送错误信息

    def __init__(self, excel_path: str):
        super().__init__()
        self.excel_path = excel_path
        self.sheet_name = None
        self._is_running = True

    def set_sheet(self, sheet_name: str):
        """设置要加载的工作表"""
        self.sheet_name = sheet_name

    def run(self):
        """在线程中执行加载"""
        try:
            # 加载工作簿获取工作表列表
            workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()

            if self._is_running:
                self.loaded.emit(sheet_names)

            # 如果指定了工作表，加载其数据
            if self.sheet_name and self._is_running:
                self.load_sheet_data(self.sheet_name)

        except Exception as e:
            if self._is_running:
                self.error.emit(str(e))

    def load_sheet_data(self, sheet_name: str):
        """加载指定工作表的数据 - 读取前10行供用户选择表头"""
        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook[sheet_name]

            # 读取前10行数据，让用户选择哪一行作为表头
            preview_rows = []
            for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                preview_rows.append([str(cell).strip() if cell is not None else "" for cell in row])

            workbook.close()

            if self._is_running:
                self.sheet_data_loaded.emit(preview_rows, [], len(preview_rows))

        except Exception as e:
            if self._is_running:
                self.error.emit(str(e))

    def stop(self):
        """停止线程"""
        self._is_running = False
        self.wait(1000)


class ImportProgressDialog(QDialog):
    """导入进度对话框 - 显示导入进度和明细"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("正在导入数据...")
        self.setMinimumSize(600, 400)
        self.setMaximumSize(800, 600)
        self.setModal(True)
        
        self.setup_ui()
        
        # 设置样式
        self.setStyleSheet("""
            QDialog { background-color: #2b2b2b; color: #ffffff; }
            QLabel { color: #ffffff; background-color: transparent; }
            QProgressBar { 
                border: 1px solid #555555; 
                border-radius: 4px; 
                background-color: #3c3c3c; 
                height: 20px;
                text-align: center;
            }
            QProgressBar::chunk { 
                background-color: #4EC9FF; 
                border-radius: 3px;
            }
            QTextEdit { 
                background-color: #1e1e1e; 
                color: #d4d4d4; 
                border: 1px solid #444444;
                border-radius: 4px;
                font-family: 'Consolas', 'Microsoft YaHei', monospace;
                font-size: 12px;
                padding: 8px;
            }
        """)
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 状态标签
        self.status_label = QLabel("正在准备导入...")
        self.status_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #4EC9FF;")
        layout.addWidget(self.status_label)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("%p% (%v/%m)")
        layout.addWidget(self.progress_bar)
        
        # 当前行信息
        self.current_row_label = QLabel("")
        self.current_row_label.setStyleSheet("color: #aaaaaa; font-size: 12px;")
        layout.addWidget(self.current_row_label)
        
        # 明细显示区域
        detail_label = QLabel("📋 导入明细（最近20行）：")
        detail_label.setStyleSheet("color: #ffffff; font-size: 12px;")
        layout.addWidget(detail_label)
        
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        layout.addWidget(self.detail_text)
        
        # 统计信息
        self.stats_label = QLabel("已导入: 0 行")
        self.stats_label.setStyleSheet("color: #aaaaaa; font-size: 12px;")
        layout.addWidget(self.stats_label)
        
        # 取消按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.cancel_btn = PushButton("取消导入")
        self.cancel_btn.setFixedSize(100, 35)
        self.cancel_btn.clicked.connect(self.on_cancel)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
        
        self.is_cancelled = False
        self.imported_count = 0
        self.error_count = 0
    
    def set_total_rows(self, total: int):
        """设置总行数"""
        self.progress_bar.setMaximum(total)
        self.total_rows = total
    
    def update_progress(self, current: int, row_data: dict = None):
        """更新进度"""
        self.progress_bar.setValue(current)
        
        # 更新状态标签
        percentage = (current / self.total_rows * 100) if self.total_rows > 0 else 0
        self.status_label.setText(f"正在导入... {percentage:.1f}%")
        
        # 更新当前行信息
        if row_data:
            sequence = row_data.get('sequence', '')
            name = row_data.get('name', '')
            self.current_row_label.setText(f"当前行: {current}/{self.total_rows} - {sequence} {name[:30]}")
            
            # 添加到明细显示
            self.add_detail_line(current, row_data)
        
        # 更新统计
        self.stats_label.setText(f"已导入: {self.imported_count} 行 | 错误: {self.error_count} 行")
        
        # 处理UI事件，保持界面响应
        QApplication.processEvents()
    
    def add_detail_line(self, row_num: int, row_data: dict):
        """添加明细行"""
        sequence = row_data.get('sequence', '')
        name = row_data.get('name', '')
        level = row_data.get('level', 1)
        
        # 根据层级添加缩进
        indent = "  " * (level - 1)
        line = f"[{row_num:4d}] {indent}{sequence:8s} {name[:40]}"
        
        self.detail_text.append(line)
        
        # 滚动到底部
        scrollbar = self.detail_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def on_cancel(self):
        """取消导入"""
        self.is_cancelled = True
        self.status_label.setText("正在取消...")
        self.status_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #ff6b6b;")
    
    def set_completed(self, success_count: int, error_count: int):
        """设置完成状态"""
        self.progress_bar.setValue(self.progress_bar.maximum())
        self.status_label.setText("✓ 导入完成")
        self.status_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #4ade80;")
        self.stats_label.setText(f"成功: {success_count} 行 | 错误: {error_count} 行")
        self.cancel_btn.setText("关闭")
        self.cancel_btn.setEnabled(True)
    
    def set_error(self, error_msg: str):
        """设置错误状态"""
        self.status_label.setText(f"✗ 导入失败")
        self.status_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #ff6b6b;")
        self.detail_text.append(f"\n错误: {error_msg}")
        self.cancel_btn.setText("关闭")


class ColumnMappingWidget(QWidget):
    """列映射选择组件 - 带预览的直观选择器"""
    
    mapping_changed = Signal()
    
    def __init__(self, field_key: str, field_name: str, parent=None):
        super().__init__(parent)
        self.field_key = field_key
        self.field_name = field_name
        self.headers = []
        self.sample_data = []
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI - 适配两列布局"""
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 3, 5, 3)
        layout.setSpacing(8)
        
        # 字段标签
        label = QLabel(f"{self.field_name}:")
        label.setMinimumWidth(120)
        label.setMaximumWidth(140)
        label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        label.setStyleSheet("color: #ffffff; background-color: transparent; font-size: 12px;")
        layout.addWidget(label)
        
        # 下拉选择框
        self.combo = QComboBox()
        self.combo.setMinimumWidth(200)
        self.combo.setMaximumWidth(280)
        self.combo.currentIndexChanged.connect(self.on_selection_changed)
        layout.addWidget(self.combo)
        
        # 预览标签（简化显示）
        self.preview_label = QLabel("")
        self.preview_label.setStyleSheet("color: #888888; font-size: 10px; background-color: transparent;")
        self.preview_label.setMinimumWidth(80)
        self.preview_label.setMaximumWidth(100)
        layout.addWidget(self.preview_label)
        
        layout.addStretch()
    
    def set_headers(self, headers: List[str], sample_data: List[List]):
        """设置表头和样本数据"""
        self.headers = headers
        self.sample_data = sample_data
        
        self.combo.clear()
        self.combo.addItem("-- 请选择 --", None)
        
        for idx, header in enumerate(headers):
            # 处理空表头
            header_display = header if header else f"(空列{idx+1})"
            
            # 获取样本值
            sample = ""
            if sample_data and idx < len(sample_data[0]):
                sample = str(sample_data[0][idx]) if sample_data[0][idx] is not None else ""
                if len(sample) > 15:
                    sample = sample[:15] + "..."
            
            # 显示格式: 列A: 表头名 (样本值)
            col_letter = self._get_column_letter(idx)
            display_text = f"{col_letter}: {header_display}"
            if sample:
                display_text += f"  [例: {sample}]"
            
            self.combo.addItem(display_text, idx)
    
    def _get_column_letter(self, n: int) -> str:
        """将列索引转换为Excel列字母 (0->A, 25->Z, 26->AA, ...)"""
        result = ""
        while n >= 0:
            result = chr(n % 26 + ord('A')) + result
            n = n // 26 - 1
        return result if result else "A"
    
    def on_selection_changed(self):
        """选择改变时更新预览"""
        col_idx = self.combo.currentData()
        if col_idx is not None and self.sample_data:
            # 显示前3个样本值
            samples = []
            for row in self.sample_data[:3]:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        samples.append(str(val))
            
            if samples:
                preview = " | ".join(samples)
                if len(preview) > 50:
                    preview = preview[:50] + "..."
                self.preview_label.setText(f"数据: {preview}")
            else:
                self.preview_label.setText("")
        else:
            self.preview_label.setText("")
        
        self.mapping_changed.emit()
    
    def get_selected_column(self) -> Optional[int]:
        """获取选中的列索引"""
        return self.combo.currentData()
    
    def set_selected_column(self, col_idx: int):
        """设置选中的列"""
        for i in range(self.combo.count()):
            if self.combo.itemData(i) == col_idx:
                self.combo.setCurrentIndex(i)
                break


class ExcelImportDialog(QDialog):
    """Excel导入对话框 - 优化版"""
    
    # 字段定义: (字段标识, 字段显示名称, 是否必填, 匹配关键词列表)
    FIELD_DEFINITIONS = [
        ('name', '工程项目及费用名称', True, [
            '工程项目及费用名称', '项目名称', '工程名称', '费用名称', 
            '名称', '项目', '工程', '费用', 'item', 'name', 'project'
        ]),
        ('sequence', '序号', False, [
            '序号', '编号', 'no', 'no.', 'number', 'id', '序号'
        ]),
    ]
    
    def __init__(self, excel_path: str, bidding_name: str = "", parent=None):
        super().__init__(parent)
        self.excel_path = excel_path
        self.bidding_name = bidding_name
        self.sheet_names = []
        self.headers = []
        self.sample_data = []
        self.preview_rows = []  # 存储前10行预览数据
        self.header_row_index = 0  # 用户选择的表头行索引
        self.load_worker = None
        self.current_sheet_name = None

        self.setWindowTitle("Excel导入 - 列映射配置")
        self.setMinimumSize(1000, 800)

        self.setup_ui()
        self.load_excel_async()
    
    def setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 设置对话框样式
        self.setStyleSheet("""
            QDialog { background-color: #2b2b2b; color: #ffffff; }
            QLabel { color: #ffffff; background-color: transparent; }
            QComboBox { background-color: #3c3c3c; color: #ffffff; border: 1px solid #555555; border-radius: 4px; padding: 5px; min-height: 25px; }
            QComboBox:hover { border-color: #0078d4; }
            QComboBox::drop-down { border: none; width: 25px; }
            QComboBox QAbstractItemView { background-color: #3c3c3c; color: #ffffff; border: 1px solid #555555; selection-background-color: #0078d4; }
            QGroupBox { color: #ffffff; border: 1px solid #555555; border-radius: 5px; margin-top: 10px; font-weight: bold; background-color: transparent; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; background-color: #2b2b2b; }
            QTableWidget { background-color: #3c3c3c; color: #ffffff; border: 1px solid #555555; gridline-color: #555555; }
            QTableWidget::item { background-color: #3c3c3c; color: #ffffff; padding: 5px; }
            QTableWidget::item:selected { background-color: #0078d4; color: #ffffff; }
            QHeaderView::section { background-color: #404040; color: #ffffff; padding: 5px; border: 1px solid #555555; font-weight: bold; }
            QScrollArea { border: none; background-color: transparent; }
            QFrame { background-color: #3c3c3c; border-radius: 5px; }
            QProgressBar { border: 1px solid #555555; border-radius: 3px; background-color: #3c3c3c; height: 8px; }
            QProgressBar::chunk { background-color: #0078d4; border-radius: 2px; }
        """)
        
        # 文件信息
        info_frame = QFrame()
        info_frame.setStyleSheet("background-color: #3c3c3c; border-radius: 5px; padding: 5px;")
        info_layout = QHBoxLayout(info_frame)
        info_label = QLabel(f"📄 文件: {os.path.basename(self.excel_path)}")
        info_label.setStyleSheet("color: #ffffff; background-color: transparent;")
        info_layout.addWidget(info_label)
        info_layout.addStretch()
        layout.addWidget(info_frame)
        
        # 工作簿选择
        sheet_layout = QHBoxLayout()
        sheet_label = QLabel("📑 工作簿:")
        sheet_label.setStyleSheet("color: #ffffff; background-color: transparent;")
        sheet_layout.addWidget(sheet_label)
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(200)
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)
        sheet_layout.addWidget(self.sheet_combo)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumWidth(150)
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.hide()
        sheet_layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #aaaaaa; font-size: 12px; background-color: transparent;")
        sheet_layout.addWidget(self.status_label)
        
        sheet_layout.addStretch()
        layout.addLayout(sheet_layout)

        # 表头行选择
        header_row_layout = QHBoxLayout()
        header_row_label = QLabel("📋 表头行:")
        header_row_label.setStyleSheet("color: #ffffff; background-color: transparent;")
        header_row_layout.addWidget(header_row_label)
        self.header_row_combo = QComboBox()
        self.header_row_combo.setMinimumWidth(200)
        self.header_row_combo.setToolTip("选择哪一行作为表头")
        self.header_row_combo.currentIndexChanged.connect(self.on_header_row_changed)
        header_row_layout.addWidget(self.header_row_combo)

        header_row_hint = QLabel("💡 提示: 请选择包含列标题的行（如'序号'、'项目名称'等）")
        header_row_hint.setStyleSheet("color: #aaaaaa; font-size: 12px; background-color: transparent;")
        header_row_layout.addWidget(header_row_hint)

        header_row_layout.addStretch()
        layout.addLayout(header_row_layout)

        # 分割器
        splitter = QSplitter(Qt.Vertical)

        # 上半部分：列映射
        mapping_widget = QWidget()
        mapping_layout = QVBoxLayout(mapping_widget)
        mapping_layout.setContentsMargins(0, 0, 0, 0)

        mapping_frame = QFrame()
        mapping_frame.setStyleSheet("background-color: #3c3c3c; border-radius: 5px;")
        mapping_inner_layout = QVBoxLayout(mapping_frame)
        mapping_inner_layout.setContentsMargins(15, 15, 15, 15)

        # 创建列映射组件
        self.mapping_widgets = {}
        for field_key, field_name, is_required, _ in self.FIELD_DEFINITIONS:
            display_name = f"* {field_name}" if is_required else field_name
            widget = ColumnMappingWidget(field_key, display_name)
            widget.mapping_changed.connect(self.on_mapping_changed)
            self.mapping_widgets[field_key] = widget
            mapping_inner_layout.addWidget(widget)

        # 自动识别按钮
        btn_layout = QHBoxLayout()
        auto_btn = PushButton("🔍 智能识别列")
        auto_btn.setToolTip("根据表头名称自动匹配列")
        auto_btn.clicked.connect(self.auto_detect_columns)
        btn_layout.addWidget(auto_btn)

        clear_btn = PushButton("🗑️ 清除选择")
        clear_btn.clicked.connect(self.clear_mappings)
        btn_layout.addWidget(clear_btn)

        btn_layout.addStretch()
        mapping_inner_layout.addLayout(btn_layout)

        mapping_layout.addWidget(mapping_frame)
        splitter.addWidget(mapping_widget)
        
        # 下半部分：原始数据预览
        preview_widget = QWidget()
        preview_layout = QVBoxLayout(preview_widget)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        
        preview_group = QGroupBox("Excel原始数据预览（前5行）")
        preview_group_layout = QVBoxLayout(preview_group)
        
        self.preview_table = QTableWidget()
        self.preview_table.setMaximumHeight(200)
        
        preview_group_layout.addWidget(self.preview_table)
        preview_layout.addWidget(preview_group)
        splitter.addWidget(preview_widget)
        
        splitter.setSizes([400, 300])
        layout.addWidget(splitter)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = PushButton("取消")
        cancel_btn.setFixedSize(100, 35)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        btn_layout.addSpacing(15)
        
        self.import_btn = PrimaryPushButton("确认导入")
        self.import_btn.setFixedSize(100, 35)
        self.import_btn.clicked.connect(self.on_import)
        btn_layout.addWidget(self.import_btn)
        
        layout.addLayout(btn_layout)
    
    def load_excel_async(self):
        """异步加载Excel文件"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "提示", "请先安装openpyxl: pip install openpyxl")
            self.reject()
            return
        
        self.progress_bar.show()
        self.status_label.setText("正在加载Excel文件...")
        self.import_btn.setEnabled(False)
        
        self.load_worker = ExcelLoadWorker(self.excel_path)
        self.load_worker.loaded.connect(self.on_excel_loaded)
        self.load_worker.sheet_data_loaded.connect(self.on_sheet_data_loaded)
        self.load_worker.error.connect(self.on_load_error)
        self.load_worker.start()
    
    def on_excel_loaded(self, sheet_names: list):
        """Excel文件加载完成回调"""
        self.sheet_names = sheet_names
        
        self.sheet_combo.clear()
        for name in sheet_names:
            self.sheet_combo.addItem(name)
        
        if sheet_names:
            self.current_sheet_name = sheet_names[0]
            self.load_worker.set_sheet(sheet_names[0])
    
    def on_sheet_data_loaded(self, preview_rows: list, _, row_count: int):
        """工作表数据加载完成回调 - 显示前10行供用户选择表头"""
        # print(f"[Excel导入] 加载完成，共 {row_count} 行预览数据")

        self.preview_rows = preview_rows

        # 填充表头行选择下拉框
        self.header_row_combo.clear()
        for i, row in enumerate(preview_rows):
            # 显示行号和前3列内容作为提示
            preview_text = " | ".join(str(cell)[:10] for cell in row[:3] if cell)
            if len(preview_text) > 30:
                preview_text = preview_text[:30] + "..."
            self.header_row_combo.addItem(f"第 {i+1} 行: {preview_text}", i)

        # 自动识别表头行
        detected_row = self._detect_header_row(preview_rows)
        self.header_row_index = detected_row

        # 设置下拉框为自动识别的行
        self.header_row_combo.setCurrentIndex(detected_row)

        self.progress_bar.hide()
        self.status_label.setText(f"✓ 已加载 {row_count} 行预览数据，自动识别表头为第 {detected_row + 1} 行")
        self.import_btn.setEnabled(True)

        # 根据选择的行更新表头和数据
        self._update_header_and_data()

        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None

    def on_header_row_changed(self, index):
        """表头行选择改变"""
        if index >= 0 and self.preview_rows:
            self.header_row_index = index
            self._update_header_and_data()

    def _update_header_and_data(self):
        """根据选择的表头行更新数据"""
        if not self.preview_rows or self.header_row_index >= len(self.preview_rows):
            return

        # 选择的行作为表头
        self.headers = self.preview_rows[self.header_row_index]

        # 后面的行作为样本数据
        self.sample_data = self.preview_rows[self.header_row_index + 1:]

        # print(f"[Excel导入] 表头行: {self.header_row_index + 1}, 表头: {self.headers}")
        # print(f"[Excel导入] 样本数据行数: {len(self.sample_data)}")

        # 更新UI
        self.update_mapping_widgets()
        self.update_raw_preview()
        self.auto_detect_columns()
    
    def on_load_error(self, error_msg: str):
        """加载错误回调"""
        self.progress_bar.hide()
        self.status_label.setText("加载失败")
        self.import_btn.setEnabled(True)
        QMessageBox.critical(self, "错误", f"无法加载Excel文件:\n{error_msg}")
        
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
    
    def load_sheet_async(self, sheet_name: str):
        """异步加载指定工作表"""
        if not sheet_name:
            return
        
        self.current_sheet_name = sheet_name
        
        self.progress_bar.show()
        self.status_label.setText(f"正在加载工作表: {sheet_name}...")
        self.import_btn.setEnabled(False)
        
        self.load_worker = ExcelLoadWorker(self.excel_path)
        self.load_worker.sheet_data_loaded.connect(self.on_sheet_data_loaded)
        self.load_worker.error.connect(self.on_load_error)
        self.load_worker.set_sheet(sheet_name)
        self.load_worker.start()
    
    def update_mapping_widgets(self):
        """更新列映射组件"""
        for widget in self.mapping_widgets.values():
            widget.set_headers(self.headers, self.sample_data)
    
    def update_raw_preview(self):
        """更新原始数据预览"""
        if not self.headers:
            return
        
        self.preview_table.clear()
        self.preview_table.setColumnCount(len(self.headers))
        self.preview_table.setHorizontalHeaderLabels(self.headers)
        
        row_count = min(5, len(self.sample_data))
        self.preview_table.setRowCount(row_count)
        
        for row_idx, row_data in enumerate(self.sample_data[:5]):
            for col_idx, value in enumerate(row_data):
                if value is None:
                    value = ""
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.preview_table.setItem(row_idx, col_idx, item)
        
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
    
    def _normalize_text(self, text: str) -> str:
        """标准化文本：移除干扰字符，统一大小写"""
        if not text:
            return ""
        # 转换为字符串并清理
        text = str(text).strip()
        # 移除常见干扰字符
        text = re.sub(r'[：:\(\)（）\[\]【】\s\.\-_]', '', text)
        # 转为小写
        return text.lower()

    def _detect_header_row(self, preview_rows: List[List[str]]) -> int:
        """自动识别表头行 - 返回最可能是表头的行索引"""
        if not preview_rows:
            return 0

        # 收集所有字段的匹配关键词
        all_keywords = []
        for _, _, _, keywords in self.FIELD_DEFINITIONS:
            all_keywords.extend(keywords)

        best_row_idx = 0
        best_score = 0.0

        # print(f"[自动识别表头] 开始识别，共 {len(preview_rows)} 行")

        for row_idx, row in enumerate(preview_rows):
            score = 0.0
            matched_keywords = []

            for cell in row:
                cell_norm = self._normalize_text(cell)
                if not cell_norm:
                    continue

                # 检查是否匹配任何关键词
                for keyword in all_keywords:
                    keyword_norm = self._normalize_text(keyword)
                    if not keyword_norm:
                        continue

                    # 完全匹配
                    if cell_norm == keyword_norm:
                        score += 1.0
                        matched_keywords.append(f"{cell}={keyword}")
                    # 包含匹配
                    elif keyword_norm in cell_norm:
                        score += 0.8
                        matched_keywords.append(f"{cell}~{keyword}")
                    # 部分匹配（至少2个字符）
                    elif len(keyword_norm) >= 2:
                        # 检查是否有共同子串
                        for i in range(len(keyword_norm)):
                            for j in range(i + 2, min(i + 5, len(keyword_norm) + 1)):
                                substr = keyword_norm[i:j]
                                if substr in cell_norm:
                                    score += 0.3 * (len(substr) / len(keyword_norm))

            # 归一化分数（除以列数，避免列多的行占优势）
            normalized_score = score / max(len(row), 1)

            # print(f"[自动识别表头] 第 {row_idx + 1} 行: 得分={normalized_score:.2f}, 匹配={matched_keywords}")

            if normalized_score > best_score:
                best_score = normalized_score
                best_row_idx = row_idx

        # print(f"[自动识别表头] 结果: 第 {best_row_idx + 1} 行是表头 (得分: {best_score:.2f})")
        return best_row_idx

    def _calculate_similarity(self, header: str, keyword: str) -> float:
        """计算表头与关键词的相似度"""
        header_norm = self._normalize_text(header)
        keyword_norm = self._normalize_text(keyword)
        
        if not header_norm or not keyword_norm:
            return 0.0
        
        # 完全匹配
        if header_norm == keyword_norm:
            return 1.0
        
        # 包含匹配
        if keyword_norm in header_norm:
            # 根据匹配位置给分
            if header_norm.startswith(keyword_norm):
                return 0.9
            elif header_norm.endswith(keyword_norm):
                return 0.8
            else:
                return 0.7
        
        # 部分匹配（编辑距离）
        # 简单实现：计算共同子串长度
        max_common = 0
        for i in range(len(keyword_norm)):
            for j in range(i + 1, len(keyword_norm) + 1):
                substr = keyword_norm[i:j]
                if substr in header_norm and len(substr) > max_common:
                    max_common = len(substr)
        
        if max_common > 0:
            return 0.5 * (max_common / max(len(header_norm), len(keyword_norm)))
        
        return 0.0
    
    def auto_detect_columns(self):
        """智能识别列 - 全新设计"""
        if not self.headers:
            # print("[智能识别] 错误：没有表头数据")
            return
        
        # print(f"\n[智能识别] ====== 开始识别 ======")
        # print(f"[智能识别] 表头列表: {self.headers}")
        
        assigned_columns = set()
        
        for field_key, field_name, is_required, keywords in self.FIELD_DEFINITIONS:
            # print(f"\n[智能识别] 识别字段: {field_name} (必填: {is_required})")
            
            best_match_idx = None
            best_score = 0.0
            best_header = ""
            best_keyword = ""
            
            for idx, header in enumerate(self.headers):
                if idx in assigned_columns:
                    continue
                
                # 计算该列与所有关键词的最高匹配分数
                for keyword in keywords:
                    score = self._calculate_similarity(header, keyword)
                    
                    if score > 0:
                        pass  # print(f"[智能识别]   列{idx} '{header}' vs '{keyword}' = {score:.2f}")

                    if score > best_score:
                        best_score = score
                        best_match_idx = idx
                        best_header = header
                        best_keyword = keyword
            
            # 设置匹配阈值
            threshold = 0.6 if is_required else 0.5
            
            if best_match_idx is not None and best_score >= threshold:
                pass  # print(f"[智能识别] ✓ 成功匹配: {field_name} -> 列{best_match_idx} '{best_header}' (得分: {best_score:.2f}, 关键词: {best_keyword})")
                self.mapping_widgets[field_key].set_selected_column(best_match_idx)
                assigned_columns.add(best_match_idx)
            else:
                pass  # print(f"[智能识别] ✗ 未匹配: {field_name} (最佳得分: {best_score:.2f}, 阈值: {threshold})")

        pass  # print(f"[智能识别] ====== 识别完成 ======\n")
        self.on_mapping_changed()
    
    def clear_mappings(self):
        """清除所有列映射"""
        for widget in self.mapping_widgets.values():
            widget.combo.setCurrentIndex(0)
    
    def on_sheet_changed(self, index):
        """工作簿改变"""
        if index >= 0 and self.sheet_names:
            sheet_name = self.sheet_names[index]
            if sheet_name != self.current_sheet_name:
                self.load_sheet_async(sheet_name)
    
    def on_mapping_changed(self):
        """列映射改变"""
        pass
    
    def get_column_mapping(self) -> Dict[str, int]:
        """获取列映射配置"""
        mapping = {}
        for field_key, widget in self.mapping_widgets.items():
            col_idx = widget.get_selected_column()
            if col_idx is not None:
                mapping[field_key] = col_idx
        return mapping
    
    def parse_sequence_level(self, sequence: str) -> int:
        """解析序号层级"""
        if not sequence:
            return 1
        
        sequence = str(sequence).strip()
        
        # 中文数字 -> 层级1
        chinese_nums = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
        if any(c in sequence for c in chinese_nums):
            return 1
        
        # 计算小数点数量
        dots = sequence.count('.')
        return dots + 2
    
    def import_data_with_progress(self, progress_dialog: ImportProgressDialog) -> List[Dict]:
        """导入数据（带进度显示）"""
        if not self.current_sheet_name or not self.excel_path:
            return []

        mapping = self.get_column_mapping()

        # 检查必填字段
        for field_key, field_name, is_required, _ in self.FIELD_DEFINITIONS:
            if is_required and field_key not in mapping:
                QMessageBox.warning(self, "提示", f"必须配置 {field_name} 列")
                return []

        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook[self.current_sheet_name]

            # 数据从表头行的下一行开始
            start_row = self.header_row_index + 2  # +2 因为行号从1开始，还要跳过表头行

            # 计算总行数
            total_rows = sheet.max_row - start_row + 1
            # print(f"[import_data] 从第 {start_row} 行开始读取数据，共 {total_rows} 行")
            
            # 设置进度对话框的总行数
            progress_dialog.set_total_rows(total_rows)

            imported_items = []
            error_count = 0
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=1):
                # 检查是否取消
                if progress_dialog.is_cancelled:
                    # print("[import_data] 用户取消导入")
                    break
                
                item_data = {}
                
                try:
                    for field_key, col_idx in mapping.items():
                        if col_idx < len(row):
                            value = row[col_idx]

                            if field_key != 'sequence' and field_key != 'name':
                                try:
                                    value = float(value) if value is not None else 0.0
                                except (ValueError, TypeError):
                                    value = 0.0
                            else:
                                value = str(value).strip() if value is not None else ""

                            item_data[field_key] = value
                        else:
                            item_data[field_key] = "" if field_key in ['sequence', 'name'] else 0.0

                    # 计算层级
                    sequence = str(item_data.get('sequence', ''))
                    if sequence and sequence.strip():
                        item_data['level'] = self.parse_sequence_level(sequence)
                    else:
                        item_data['level'] = 1

                    imported_items.append(item_data)
                    progress_dialog.imported_count += 1
                    
                except Exception as e:
                    error_count += 1
                    progress_dialog.error_count += 1
                    # print(f"[import_data] 第 {row_idx} 行导入错误: {e}")
                
                # 更新进度（每5行更新一次，避免UI卡顿）
                if row_idx % 5 == 0 or row_idx == total_rows:
                    progress_dialog.update_progress(row_idx, item_data if item_data else None)

            workbook.close()
            
            # 设置完成状态
            progress_dialog.set_completed(len(imported_items), error_count)
            
            return imported_items

        except Exception as e:
            progress_dialog.set_error(str(e))
            QMessageBox.critical(self, "错误", f"导入数据失败:\n{e}")
            return []
    
    def on_import(self):
        """确认导入"""
        # 创建并显示进度对话框
        progress_dialog = ImportProgressDialog(self)
        progress_dialog.show()
        
        # 处理UI事件，确保对话框显示
        QApplication.processEvents()
        
        # 执行导入
        data = self.import_data_with_progress(progress_dialog)
        
        # 如果用户没有取消且导入成功，等待用户关闭进度对话框
        if not progress_dialog.is_cancelled and data:
            # 存储导入的数据供后续使用
            self._imported_data = data
            # 等待用户点击关闭按钮
            progress_dialog.exec()
            # 用户关闭后，接受对话框
            self.accept()
        elif progress_dialog.is_cancelled:
            # 用户取消，不关闭主对话框
            QMessageBox.information(self, "提示", f"已取消导入，成功导入 {len(data)} 行数据")
        
    def get_imported_data(self) -> List[Dict]:
        """获取导入的数据"""
        return getattr(self, '_imported_data', [])
    
    def closeEvent(self, event):
        """关闭事件"""
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
        event.accept()


class ExcelFileSelector(QWidget):
    """Excel文件选择器组件"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.excel_files = []
        self.current_path = ""
        self.setup_ui()
    
    def setup_ui(self):
        """设置UI"""
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)
        
        layout.addWidget(QLabel("Excel文件:"))
        
        self.file_combo = QComboBox()
        self.file_combo.setMinimumWidth(200)
        layout.addWidget(self.file_combo)
        
        self.refresh_btn = PushButton("🔄 刷新")
        self.refresh_btn.clicked.connect(self.refresh_files)
        layout.addWidget(self.refresh_btn)
        
        self.upload_btn = PushButton("📤 上传")
        self.upload_btn.clicked.connect(self.upload_file)
        layout.addWidget(self.upload_btn)
        
        layout.addStretch()
    
    def set_directory(self, directory: str):
        """设置目录并加载Excel文件"""
        self.current_path = directory
        self.refresh_files()
    
    def refresh_files(self):
        """刷新文件列表"""
        self.file_combo.clear()
        self.file_combo.addItem("-- 请选择Excel文件 --", None)
        self.excel_files = []
        
        if not self.current_path or not os.path.exists(self.current_path):
            return
        
        try:
            path = Path(self.current_path)
            for ext in ['*.xlsx', '*.xls']:
                for file_path in sorted(path.glob(ext)):
                    self.excel_files.append(str(file_path))
                    self.file_combo.addItem(file_path.name, str(file_path))
        except Exception as e:
            pass  # print(f"刷新文件列表失败: {e}")

    def upload_file(self):
        """上传文件"""
        if not self.current_path:
            QMessageBox.warning(self, "提示", "请先设置目录")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "",
            "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        
        if file_path:
            try:
                import shutil
                target_path = os.path.join(self.current_path, os.path.basename(file_path))
                
                counter = 1
                original_name, ext = os.path.splitext(target_path)
                while os.path.exists(target_path):
                    target_path = f"{original_name}_{counter}{ext}"
                    counter += 1
                
                shutil.copy2(file_path, target_path)
                self.refresh_files()
                
                for i in range(self.file_combo.count()):
                    if self.file_combo.itemData(i) == target_path:
                        self.file_combo.setCurrentIndex(i)
                        break
                
                QMessageBox.information(self, "成功", "文件上传成功")
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"上传失败:\n{e}")
    
    def get_selected_file(self) -> Optional[str]:
        """获取选中的文件路径"""
        return self.file_combo.currentData()
