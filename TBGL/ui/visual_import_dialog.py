"""
可视化导入对话框
用于从Excel文件导入明细数据，支持可视化识别、列映射、行类型标记
"""
import os
import re
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Set
from enum import Enum, auto

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QPushButton, QFrame, QWidget,
    QProgressBar, QTextEdit, QApplication, QCheckBox, QLineEdit,
    QSplitter, QScrollArea, QGridLayout, QFormLayout, QMenu,
    QAbstractItemView, QStyledItemDelegate, QStyleOptionViewItem,
    QStyle
)
from PySide6.QtCore import Qt, Signal, QThread, QSize
from PySide6.QtGui import QFont, QColor, QBrush, QIcon, QPixmap, QPainter

from ui.message_dialog import MessageDialog
from ui.fluent_widgets import PushButton, PrimaryPushButton
from .excel_import_dialog import ExcelLoadWorker


class FullDataLoadWorker(QThread):
    """完整数据加载工作线程 - 加载所有行数据"""
    
    data_loaded = Signal(list, int)  # 发送所有行数据、总行数
    error = Signal(str)  # 发送错误信息
    
    def __init__(self, excel_path: str, sheet_name: str):
        super().__init__()
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self._is_running = True
    
    def set_sheet(self, sheet_name: str):
        """设置要加载的工作表名称"""
        self.sheet_name = sheet_name
    
    def run(self):
        """在线程中执行加载"""
        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            sheet = workbook[self.sheet_name]
            
            # 读取所有行数据
            all_rows = []
            row_count = 0
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not self._is_running:
                    break
                all_rows.append([str(cell).strip() if cell is not None else "" for cell in row])
                row_count += 1
            
            workbook.close()
            
            if self._is_running:
                self.data_loaded.emit(all_rows, row_count)
                
        except Exception as e:
            if self._is_running:
                self.error.emit(str(e))
    
    def stop(self):
        """停止线程"""
        self._is_running = False
        self.wait(1000)


class RowType(Enum):
    """行类型枚举"""
    UNKNOWN = "未知"
    HEADER = "表头"      # 表头行 - 无效
    FOOTER = "表尾"      # 表尾行 - 无效
    DIVISION_1 = "分部一级"   # 父级
    DIVISION_2 = "分部二级"
    DIVISION_3 = "分部三级"
    DIVISION_4 = "分部四级"
    DIVISION_5 = "分部五级"
    ITEM = "清单行"      # 有项目名称、单位、工程量的行
    MERGE = "合并行"     # 需要合并到清单行的内容
    INVALID = "无效行"   # 其他无效行


class VisualImportDialog(QDialog):
    """可视化导入对话框"""
    
    # 字段定义: (字段标识, 字段显示名称, 是否必填, 匹配关键词列表)
    FIELD_DEFINITIONS = [
        ('sequence', '序号', False, ['序号', '编号', 'no', 'no.', 'number', 'id', '流水号']),
        ('name', '分部分项工程名称', True, ['分部分项工程名称', '工程名称', '项目名称', '名称', '项目', '工程', 'name', 'item', '工程或费用名称', '费用名称']),
        ('specification', '规格型号', False, ['规格型号', '规格', '型号', 'spec', 'specification', 'type', 'model']),
        ('description', '项目特征描述', False, ['项目特征描述', '特征描述', '项目特征', '描述', '特征', 'description', '工作内容', '内容', '工作']),
        ('unit', '单位', False, ['单位', '计量单位', 'unit', '计量']),
        ('quantity', '工程量', False, ['工程量', '数量', 'quantity', 'amount', '工程数量', '数 量']),
        ('unit_price', '综合单价', False, ['综合单价', '单价', 'unit price', 'price', '单价(元)']),
        ('labor_unit_price', '人工单价', False, ['人工单价', '人工费单价', '人工', 'labor price']),
        ('material_unit_price', '主材单价', False, ['主材单价', '材料单价', '材料费单价']),
        ('material_loss_rate', '主材损耗%', False, ['主材损耗', '材料损耗', '损耗率', '损耗%', '损耗', '主材损耗率']),
        ('auxiliary_unit_price', '辅材单价', False, ['辅材单价', '辅助材料单价', '辅材', 'auxiliary price']),
        ('machine_unit_price', '机械单价', False, ['机械单价', '机械费单价', '机械台班单价', '机械', 'machine price']),
        ('other_unit_price', '其他单价', False, ['其他单价', '其他费单价', '其他', 'other price']),
        ('remark', '备注', False, ['备注', '说明', 'note', 'remark', '注释']),
    ]
    
    # 行类型样式配置 - 浅色主题
    ROW_TYPE_STYLES_LIGHT = {
        RowType.HEADER: {'bg': '#FFE4E1', 'fg': '#8B0000', 'icon': '⬆'},
        RowType.FOOTER: {'bg': '#FFE4E1', 'fg': '#8B0000', 'icon': '⬇'},
        RowType.INVALID: {'bg': '#F5F5F5', 'fg': '#808080', 'icon': '✗'},
        RowType.DIVISION_1: {'bg': '#E6F3FF', 'fg': '#0066CC', 'icon': '①'},
        RowType.DIVISION_2: {'bg': '#E6F3FF', 'fg': '#0066CC', 'icon': '②'},
        RowType.DIVISION_3: {'bg': '#E6F3FF', 'fg': '#0066CC', 'icon': '③'},
        RowType.DIVISION_4: {'bg': '#E6F3FF', 'fg': '#0066CC', 'icon': '④'},
        RowType.DIVISION_5: {'bg': '#E6F3FF', 'fg': '#0066CC', 'icon': '⑤'},
        RowType.ITEM: {'bg': '#F0FFF0', 'fg': '#006400', 'icon': '✓'},
        RowType.MERGE: {'bg': '#FFF8DC', 'fg': '#B8860B', 'icon': '➕'},
        RowType.UNKNOWN: {'bg': '#FFFFFF', 'fg': '#000000', 'icon': '?'},
    }
    
    # 行类型样式配置 - 深色主题
    ROW_TYPE_STYLES_DARK = {
        RowType.HEADER: {'bg': '#3D2020', 'fg': '#FF9999', 'icon': '⬆'},
        RowType.FOOTER: {'bg': '#3D2020', 'fg': '#FF9999', 'icon': '⬇'},
        RowType.INVALID: {'bg': '#2A2A2A', 'fg': '#666666', 'icon': '✗'},
        RowType.DIVISION_1: {'bg': '#1A2A3D', 'fg': '#66B2FF', 'icon': '①'},
        RowType.DIVISION_2: {'bg': '#1A2A3D', 'fg': '#66B2FF', 'icon': '②'},
        RowType.DIVISION_3: {'bg': '#1A2A3D', 'fg': '#66B2FF', 'icon': '③'},
        RowType.DIVISION_4: {'bg': '#1A2A3D', 'fg': '#66B2FF', 'icon': '④'},
        RowType.DIVISION_5: {'bg': '#1A2A3D', 'fg': '#66B2FF', 'icon': '⑤'},
        RowType.ITEM: {'bg': '#1A3D1A', 'fg': '#66FF66', 'icon': '✓'},
        RowType.MERGE: {'bg': '#3D3D1A', 'fg': '#FFFF66', 'icon': '➕'},
        RowType.UNKNOWN: {'bg': '#1E1E1E', 'fg': '#E6E6E6', 'icon': '?'},
    }
    
    def __init__(self, excel_path: str, bidding_name: str = "", summary_name: str = "", parent=None, is_dark_theme: bool = False):
        super().__init__(parent)
        self.excel_path = excel_path
        self.bidding_name = bidding_name
        self.summary_name = summary_name
        self.sheet_names = []
        self.headers = []
        self.preview_rows = []
        self.header_row_index = 0
        self.header_row_count = 1  # 表头行数（支持多行表头）
        self.load_worker = None
        self.current_sheet_name = None
        self.drag_pos = None
        self.is_dark_theme = is_dark_theme
        
        # 根据主题选择样式
        self.ROW_TYPE_STYLES = self.ROW_TYPE_STYLES_DARK if is_dark_theme else self.ROW_TYPE_STYLES_LIGHT
        
        # 行类型识别结果
        self.row_types = {}  # row_index -> RowType
        self.checked_rows = set()  # 被选中的行
        self.column_mappings = {}  # field_key -> col_index
        
        # 列识别完成标志
        self.column_detection_done = False
        
        # 设置无边框窗口
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        
        # 设置对象名以便QSS样式应用
        self.setObjectName("visualImportDialog")
        
        self.setMinimumSize(1200, 800)
        
        self.setup_ui()
        self.load_excel_async()
    
    def setup_ui(self):
        """设置UI - 自定义标题栏"""
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # 创建内容容器
        self.content_frame = QFrame()
        self.content_frame.setObjectName("contentFrame")
        layout = QVBoxLayout(self.content_frame)
        layout.setSpacing(10)
        layout.setContentsMargins(0, 0, 0, 15)
        
        # ========== 自定义标题栏 ==========
        title_bar = QFrame()
        title_bar.setObjectName("titleBarFrame")
        title_bar.setFixedHeight(50)
        title_bar_layout = QHBoxLayout(title_bar)
        title_bar_layout.setContentsMargins(20, 0, 20, 0)
        title_bar_layout.setSpacing(10)
        
        # 标题
        title = QLabel("📊 可视化导入 - 投标明细")
        title.setFont(QFont("Microsoft YaHei", 14, QFont.Bold))
        title.setObjectName("titleLabel")
        title_bar_layout.addWidget(title)
        
        title_bar_layout.addStretch()
        
        # 关闭按钮
        close_btn = QPushButton("×")
        close_btn.setObjectName("closeButton")
        close_btn.setProperty("class", "titlebar-button")
        close_btn.setFixedSize(46, 32)
        close_btn.setFont(QFont("Microsoft YaHei", 14))
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.clicked.connect(self.reject)
        title_bar_layout.addWidget(close_btn)
        
        layout.addWidget(title_bar)
        
        # ========== 顶部信息区域 ==========
        info_widget = QWidget()
        info_widget.setObjectName("infoWidget")
        info_layout = QHBoxLayout(info_widget)
        info_layout.setContentsMargins(20, 5, 20, 5)
        info_layout.setSpacing(15)
        
        # 文件信息
        file_label = QLabel(f"📄 {os.path.basename(self.excel_path)}")
        file_label.setObjectName("fileLabel")
        info_layout.addWidget(file_label)
        
        # 工作簿选择
        info_layout.addWidget(QLabel("工作簿:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setObjectName("sheetCombo")
        self.sheet_combo.setMinimumWidth(150)
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)
        info_layout.addWidget(self.sheet_combo)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progressBar")
        self.progress_bar.setMaximumWidth(100)
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.hide()
        info_layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("")
        self.status_label.setObjectName("statusLabel")
        info_layout.addWidget(self.status_label)
        
        info_layout.addStretch()
        
        # 自动识别列按钮
        self.auto_detect_col_btn = PushButton("🔍 自动识别列")
        self.auto_detect_col_btn.setObjectName("autoDetectColButton")
        self.auto_detect_col_btn.setToolTip("自动识别表头列映射（支持多行表头）")
        self.auto_detect_col_btn.setFixedSize(120, 32)
        self.auto_detect_col_btn.clicked.connect(self.auto_detect_columns)
        info_layout.addWidget(self.auto_detect_col_btn)
        
        # 自动识别行按钮（初始禁用）
        self.auto_detect_row_btn = PushButton("🔍 自动识别行")
        self.auto_detect_row_btn.setObjectName("autoDetectRowButton")
        self.auto_detect_row_btn.setToolTip("先识别列后才能识别行")
        self.auto_detect_row_btn.setFixedSize(120, 32)
        self.auto_detect_row_btn.setEnabled(False)
        self.auto_detect_row_btn.clicked.connect(self.auto_detect_rows)
        info_layout.addWidget(self.auto_detect_row_btn)
        
        # 清除按钮
        clear_btn = PushButton("🗑️ 清除")
        clear_btn.setObjectName("clearButton")
        clear_btn.setFixedSize(80, 32)
        clear_btn.clicked.connect(self.clear_all)
        info_layout.addWidget(clear_btn)
        
        layout.addWidget(info_widget)
        
        # ========== 列识别区域 ==========
        self.column_mapping_widget = QWidget()
        self.column_mapping_widget.setObjectName("columnMappingWidget")
        column_mapping_layout = QHBoxLayout(self.column_mapping_widget)
        column_mapping_layout.setContentsMargins(20, 5, 20, 5)
        column_mapping_layout.setSpacing(5)
        
        column_mapping_layout.addWidget(QLabel("列识别:"))
        self.column_mapping_layout = column_mapping_layout  # 保存引用以便动态添加标签
        
        # ========== 数据预览表格 ==========
        self.preview_table = QTableWidget()
        self.preview_table.setObjectName("previewTable")
        self.preview_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.preview_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.verticalHeader().setVisible(True)
        self.preview_table.verticalHeader().setDefaultSectionSize(35)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.preview_table.customContextMenuRequested.connect(self.show_context_menu)
        
        layout.addWidget(self.preview_table)
        
        # ========== 底部按钮区域 ==========
        btn_layout = QHBoxLayout()
        btn_layout.setContentsMargins(20, 10, 20, 0)
        btn_layout.addStretch()
        
        cancel_btn = PushButton("取消")
        cancel_btn.setObjectName("secondaryButton")
        cancel_btn.setFixedSize(100, 35)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        btn_layout.addSpacing(20)
        
        self.import_btn = PrimaryPushButton("确认导入")
        self.import_btn.setObjectName("primaryButton")
        self.import_btn.setFixedSize(100, 35)
        self.import_btn.clicked.connect(self.on_import)
        btn_layout.addWidget(self.import_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 添加内容框架到主布局
        main_layout.addWidget(self.content_frame)
    
    def load_excel_async(self):
        """异步加载Excel文件"""
        if not OPENPYXL_AVAILABLE:
            MessageDialog.warning(self, "提示", "请先安装openpyxl: pip install openpyxl")
            self.reject()
            return
        
        self.progress_bar.show()
        self.status_label.setText("正在加载Excel文件...")
        self.import_btn.setEnabled(False)
        
        self.load_worker = ExcelLoadWorker(self.excel_path)
        self.load_worker.loaded.connect(self.on_excel_loaded)
        self.load_worker.sheet_data_loaded.connect(self.on_sheet_data_loaded)
        self.load_worker.error.connect(self.on_load_error)
        self.load_worker.start()
    
    def on_excel_loaded(self, sheet_names: list):
        """Excel文件加载完成回调"""
        self.sheet_names = sheet_names
        
        self.sheet_combo.clear()
        for name in sheet_names:
            self.sheet_combo.addItem(name)
        
        if sheet_names:
            # 智能选择工作簿
            preferred_keywords = ['清单', '明细', '分部分项', '工程', '项目']
            selected_idx = 0
            best_score = 0
            
            for idx, name in enumerate(sheet_names):
                name_lower = name.lower()
                score = sum(1 for kw in preferred_keywords if kw in name_lower)
                if score > best_score:
                    best_score = score
                    selected_idx = idx
            
            self.current_sheet_name = sheet_names[selected_idx]
            self.sheet_combo.setCurrentIndex(selected_idx)
            self.load_worker.set_sheet(self.current_sheet_name)
    
    def on_sheet_data_loaded(self, preview_rows: list, _, row_count: int):
        """工作表数据加载完成回调"""
        self.preview_rows = preview_rows
        
        # 初始化行类型
        self.row_types = {i: RowType.UNKNOWN for i in range(len(preview_rows))}
        self.checked_rows = set()
        self.column_detection_done = False
        self.auto_detect_row_btn.setEnabled(False)
        
        # 更新预览表格
        self.update_preview_table()
        
        self.progress_bar.hide()
        self.status_label.setText(f"✓ 已加载 {len(preview_rows)} 行数据，请先点击'自动识别列'")
        self.import_btn.setEnabled(True)
        
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
    
    def update_preview_table(self):
        """更新预览表格 - 添加选择列和标识列"""
        if not self.preview_rows:
            return
        
        # 获取最大列数
        max_cols = max(len(row) for row in self.preview_rows) if self.preview_rows else 0
        
        # 设置列数：选择列 + 标识列 + 数据列
        self.preview_table.setColumnCount(2 + max_cols)  # 选择列 + 标识列 + 数据列
        
        # 设置行数：列识别行 + 数据行
        self.preview_table.setRowCount(1 + len(self.preview_rows))
        
        # 设置表头
        headers = ["选择", "标识"] + [f"列{i+1}" for i in range(max_cols)]
        self.preview_table.setHorizontalHeaderLabels(headers)
        
        # 设置列宽
        self.preview_table.setColumnWidth(0, 50)   # 选择列
        self.preview_table.setColumnWidth(1, 80)   # 标识列
        
        # 设置第0行为列识别行（使用不同背景色）
        self._setup_column_mapping_row(max_cols)
        
        # 填充数据行（从第1行开始）
        for row_idx, row_data in enumerate(self.preview_rows):
            table_row = row_idx + 1  # 第0行是列识别行
            
            # 选择列 - 使用复选框
            checkbox_item = QTableWidgetItem()
            checkbox_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            if row_idx in self.checked_rows:
                checkbox_item.setCheckState(Qt.Checked)
            else:
                checkbox_item.setCheckState(Qt.Unchecked)
            self.preview_table.setItem(table_row, 0, checkbox_item)
            
            # 标识列 - 显示行类型
            row_type = self.row_types.get(row_idx, RowType.UNKNOWN)
            type_item = QTableWidgetItem(f"{self.ROW_TYPE_STYLES[row_type]['icon']} {row_type.value}")
            type_item.setData(Qt.UserRole, row_type)
            type_item.setFlags(type_item.flags() & ~Qt.ItemIsEditable)
            self.preview_table.setItem(table_row, 1, type_item)
            
            # 数据列
            for col_idx, value in enumerate(row_data):
                if value is None:
                    value = ""
                item = QTableWidgetItem(str(value))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.preview_table.setItem(table_row, 2 + col_idx, item)
            
            # 应用行样式
            self.apply_row_style(table_row, row_type)
    
    def _setup_column_mapping_row(self, max_cols: int):
        """设置第0行为列识别行 - 显示每列的识别结果并可点击修改"""
        # 字段名到显示名称的映射
        field_display_names = {
            'sequence': '序号',
            'name': '项目名称',
            'specification': '规格型号',
            'description': '项目特征',
            'unit': '单位',
            'quantity': '工程量',
            'unit_price': '综合单价',
            'labor_unit_price': '人工单价',
            'material_unit_price': '主材单价',
            'material_loss_rate': '损耗率',
            'auxiliary_unit_price': '辅材单价',
            'machine_unit_price': '机械单价',
            'other_unit_price': '其他单价',
            'remark': '备注',
        }
        
        # 第0列（选择列）- 显示"列识别"标签
        header_item = QTableWidgetItem("列识别")
        header_item.setFlags(header_item.flags() & ~Qt.ItemIsEditable)
        header_item.setBackground(QBrush(QColor(self.ROW_TYPE_STYLES[RowType.HEADER]['bg'])))
        header_item.setForeground(QBrush(QColor(self.ROW_TYPE_STYLES[RowType.HEADER]['fg'])))
        header_item.setTextAlignment(Qt.AlignCenter)
        font = QFont()
        font.setBold(True)
        header_item.setFont(font)
        self.preview_table.setItem(0, 0, header_item)
        
        # 第1列（标识列）- 留空
        empty_item = QTableWidgetItem("")
        empty_item.setFlags(empty_item.flags() & ~Qt.ItemIsEditable)
        empty_item.setBackground(QBrush(QColor(self.ROW_TYPE_STYLES[RowType.HEADER]['bg'])))
        self.preview_table.setItem(0, 1, empty_item)
        
        # 数据列 - 显示识别结果
        for col_idx in range(max_cols):
            # 查找该列对应的字段
            field_name = None
            for field, col in self.column_mappings.items():
                if col == col_idx:
                    field_name = field
                    break
            
            if field_name:
                display_name = field_display_names.get(field_name, field_name)
                cell_text = display_name
                tooltip = f"列 {col_idx+1} 已识别为: {display_name}\n点击修改"
                bg_color = QColor("#4CAF50")  # 绿色表示已识别
                fg_color = QColor("white")
            else:
                cell_text = f"列{col_idx+1}"
                tooltip = f"列 {col_idx+1} 未识别\n点击设置"
                bg_color = QColor("#9E9E9E")  # 灰色表示未识别
                fg_color = QColor("white")
            
            item = QTableWidgetItem(cell_text)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setBackground(QBrush(bg_color))
            item.setForeground(QBrush(fg_color))
            item.setTextAlignment(Qt.AlignCenter)
            item.setToolTip(tooltip)
            item.setData(Qt.UserRole, col_idx)  # 存储列索引
            font = QFont()
            font.setBold(True)
            item.setFont(font)
            self.preview_table.setItem(0, 2 + col_idx, item)
        
        # 设置第0行的行高
        self.preview_table.setRowHeight(0, 35)
        
        # 连接单元格点击事件到列识别修改
        self.preview_table.cellClicked.connect(self.on_column_mapping_cell_clicked)
    
    def on_column_mapping_cell_clicked(self, row: int, column: int):
        """处理列识别行的点击事件"""
        if row != 0:  # 只处理第0行
            return
        
        if column < 2:  # 不处理选择列和标识列
            return
        
        col_idx = column - 2  # 转换为数据列索引
        self.show_column_mapping_menu_for_col(col_idx)
    
    def show_column_mapping_menu_for_col(self, col_idx: int):
        """显示指定列的映射选择菜单"""
        menu = QMenu(self)
        menu.setObjectName("columnMappingMenu")
        
        # 字段选项
        field_options = [
            ("--未识别--", None),
            ("序号", "sequence"),
            ("项目名称", "name"),
            ("规格型号", "specification"),
            ("项目特征", "description"),
            ("单位", "unit"),
            ("工程量", "quantity"),
            ("综合单价", "unit_price"),
            ("人工单价", "labor_unit_price"),
            ("主材单价", "material_unit_price"),
            ("损耗率", "material_loss_rate"),
            ("辅材单价", "auxiliary_unit_price"),
            ("机械单价", "machine_unit_price"),
            ("其他单价", "other_unit_price"),
            ("备注", "remark"),
        ]
        
        # 查找当前列已映射的字段
        current_field = None
        for field, col in self.column_mappings.items():
            if col == col_idx:
                current_field = field
                break
        
        for label, field_key in field_options:
            action = menu.addAction(label)
            action.setCheckable(True)
            action.setChecked(field_key == current_field)
            action.triggered.connect(lambda checked, c=col_idx, f=field_key: self.set_column_mapping(c, f))
        
        # 在单元格位置显示菜单
        header_item = self.preview_table.item(0, 2 + col_idx)
        if header_item:
            rect = self.preview_table.visualItemRect(header_item)
            global_pos = self.preview_table.viewport().mapToGlobal(rect.bottomLeft())
            menu.exec_(global_pos)
    
    def update_column_mapping_display(self, max_cols: int):
        """更新列识别行显示 - 在列映射改变后调用"""
        # 重新设置第0行
        self._setup_column_mapping_row(max_cols)
    
    def apply_row_style(self, table_row: int, row_type: RowType):
        """应用行样式"""
        style = self.ROW_TYPE_STYLES.get(row_type, self.ROW_TYPE_STYLES[RowType.UNKNOWN])
        
        for col_idx in range(self.preview_table.columnCount()):
            item = self.preview_table.item(table_row, col_idx)
            if item:
                item.setBackground(QBrush(QColor(style['bg'])))
                item.setForeground(QBrush(QColor(style['fg'])))
    
    def set_column_mapping(self, col_idx: int, field_key: Optional[str]):
        """设置列映射"""
        # 清除该列之前的映射
        fields_to_remove = [f for f, c in self.column_mappings.items() if c == col_idx]
        for f in fields_to_remove:
            del self.column_mappings[f]
        
        # 如果选择了字段，设置新映射
        if field_key:
            # 如果该字段已映射到其他列，先清除
            if field_key in self.column_mappings:
                del self.column_mappings[field_key]
            self.column_mappings[field_key] = col_idx
        
        # 更新显示
        max_cols = max(len(row) for row in self.preview_rows) if self.preview_rows else 0
        self.update_column_mapping_display(max_cols)
        
        # 检查是否已识别必要列
        if 'name' in self.column_mappings:
            self.column_detection_done = True
            self.auto_detect_row_btn.setEnabled(True)
            self.auto_detect_row_btn.setToolTip("点击自动识别行类型")
            self.status_label.setText(f"✓ 列映射已更新，已识别 {len(self.column_mappings)} 列，可点击'自动识别行'")
        else:
            self.status_label.setText(f"✓ 列映射已更新，已识别 {len(self.column_mappings)} 列，请继续设置项目名称列")
    
    def auto_detect_columns(self):
        """自动识别列映射 - 支持多行表头"""
        if not self.preview_rows:
            return
        
        # 1. 检测表头行
        header_row_idx = self._detect_header_row()
        if header_row_idx is None:
            self.status_label.setText("✗ 未能识别表头行，请手动设置")
            return
        
        # 2. 检测多行表头
        self._detect_multi_row_headers(header_row_idx)
        
        # 3. 更新列映射
        self._update_column_mappings_from_header()
        
        # 4. 标记表头行为无效行
        for i in range(header_row_idx, header_row_idx + self.header_row_count):
            if i < len(self.preview_rows):
                self.row_types[i] = RowType.INVALID
        
        # 5. 更新表格显示
        self.update_preview_table()
        
        # 6. 启用行识别按钮
        self.column_detection_done = True
        self.auto_detect_row_btn.setEnabled(True)
        self.auto_detect_row_btn.setToolTip("自动识别行类型（分部、清单、合并行）")
        
        self.status_label.setText(f"✓ 列识别完成，识别出{self.header_row_count}行表头，现在可以识别行")
    
    def _detect_header_row(self) -> Optional[int]:
        """自动识别表头行"""
        if not self.preview_rows:
            return None
        
        all_keywords = []
        for _, _, _, keywords in self.FIELD_DEFINITIONS:
            all_keywords.extend(keywords)
        
        best_row_idx = None
        best_score = 0.0
        
        for row_idx, row in enumerate(self.preview_rows):
            score = 0.0
            for cell in row:
                cell_text = str(cell).lower().strip()
                for keyword in all_keywords:
                    if keyword.lower() in cell_text:
                        score += 1.0
            
            normalized_score = score / max(len(row), 1)
            if normalized_score > best_score and normalized_score > 0.3:
                best_score = normalized_score
                best_row_idx = row_idx
        
        return best_row_idx
    
    def _detect_multi_row_headers(self, header_row_idx: int):
        """检测多行表头"""
        if header_row_idx >= len(self.preview_rows):
            self.header_row_count = 1
            return
        
        # 尝试合并后续2-3行作为完整表头
        rows_to_merge = [self.preview_rows[header_row_idx]]
        for i in range(1, 3):  # 最多合并后续2行
            if header_row_idx + i < len(self.preview_rows):
                next_row = self.preview_rows[header_row_idx + i]
                # 如果下一行有非空单元格，可能是多行表头的一部分
                if any(cell and str(cell).strip() for cell in next_row):
                    rows_to_merge.append(next_row)
                else:
                    break
        
        # 合并多行表头
        max_cols = max(len(row) for row in rows_to_merge)
        merged_headers = []
        for col_idx in range(max_cols):
            header_parts = []
            for row in rows_to_merge:
                if col_idx < len(row) and row[col_idx]:
                    cell_val = str(row[col_idx]).strip()
                    if cell_val and cell_val not in header_parts:
                        header_parts.append(cell_val)
            
            if header_parts:
                merged_headers.append(''.join(header_parts))
            else:
                merged_headers.append('')
        
        self.headers = merged_headers
        self.header_row_index = header_row_idx
        self.header_row_count = len(rows_to_merge)
    
    def _update_column_mappings_from_header(self):
        """根据识别的表头更新列映射"""
        for field_key, field_name, is_required, keywords in self.FIELD_DEFINITIONS:
            best_col = None
            best_score = 0.0
            
            for col_idx, header in enumerate(self.headers):
                header_text = str(header).lower()
                for keyword in keywords:
                    keyword_lower = keyword.lower()
                    if keyword_lower in header_text:
                        score = len(keyword_lower) / len(header_text) if header_text else 0
                        if score > best_score:
                            best_score = score
                            best_col = col_idx
            
            if best_col is not None and best_score > 0.3:
                self.column_mappings[field_key] = best_col
        
        # 更新列映射显示
        max_cols = max(len(row) for row in self.preview_rows) if self.preview_rows else 0
        self.update_column_mapping_display(max_cols)
    
    def auto_detect_rows(self):
        """自动识别行类型"""
        if not self.column_detection_done:
            MessageDialog.warning(self, "提示", "请先点击'自动识别列'按钮")
            return
        
        if not self.preview_rows:
            return
        
        name_col = self.column_mappings.get('name')
        unit_col = self.column_mappings.get('unit')
        quantity_col = self.column_mappings.get('quantity')
        
        # 首先标记所有表头行为无效行
        for row_idx in range(0, self.header_row_index + self.header_row_count):
            self.row_types[row_idx] = RowType.HEADER
        
        # 从表头之后开始识别数据行
        start_row = self.header_row_index + self.header_row_count
        
        # 跟踪当前分部层级深度
        division_stack = []
        
        for row_idx in range(start_row, len(self.preview_rows)):
            row_data = self.preview_rows[row_idx]
            
            # 获取关键列的值
            name_val = str(row_data[name_col]).strip() if name_col and name_col < len(row_data) else ""
            unit_val = str(row_data[unit_col]).strip() if unit_col and unit_col < len(row_data) else ""
            quantity_val = str(row_data[quantity_col]).strip() if quantity_col and quantity_col < len(row_data) else ""
            
            # 首先检查是否是表尾关键词
            row_text = ' '.join(str(cell).lower() for cell in row_data if cell)
            footer_keywords = ['合计', '总计', '汇总', '小计', '共', '元']
            is_footer = any(keyword in row_text for keyword in footer_keywords)
            
            if is_footer:
                self.row_types[row_idx] = RowType.FOOTER
                continue
            
            # 获取项目特征列的值（用于区分分部行和合并行）
            desc_col = self.column_mappings.get('description')
            desc_val = str(row_data[desc_col]).strip() if desc_col and desc_col < len(row_data) else ""
            has_description = bool(desc_val and len(desc_val) > 0)
            
            # 识别规则
            if not name_val:
                # 空行 - 先标记为无效，后续合并行检测可能会重新标记
                self.row_types[row_idx] = RowType.INVALID
            elif unit_val and quantity_val:
                # 有单位、有工程量 - 清单行，需要勾选
                self.row_types[row_idx] = RowType.ITEM
                self.checked_rows.add(row_idx)
                # 清空分部栈，因为清单行是分部的叶子节点
                division_stack = []
            elif name_val and quantity_val and not unit_val:
                # 有项目名称、有工程量但无单位 - 也识别为清单行（处理特殊情况）
                self.row_types[row_idx] = RowType.ITEM
                self.checked_rows.add(row_idx)
                # 清空分部栈，因为清单行是分部的叶子节点
                division_stack = []
            elif name_val and not unit_val and not quantity_val:
                # 有项目名称、无单位、无工程量 - 可能是分部行或合并行
                # 关键区分条件：是否有项目特征
                # - 有项目特征 = 合并行（是清单行描述的延续）
                # - 无项目特征 = 分部行（是分类标题）
                
                if has_description:
                    # 有项目特征，说明是合并行（清单行内容的延续）
                    # 先标记为无效，后续合并行检测会统一处理
                    self.row_types[row_idx] = RowType.INVALID
                else:
                    # 无项目特征，说明是分部行
                    # 根据项目名称判断分部级别
                    division_level = self._detect_division_level(name_val, division_stack)
                    
                    if division_level == 1:
                        self.row_types[row_idx] = RowType.DIVISION_1
                        division_stack = [row_idx]
                    elif division_level == 2:
                        self.row_types[row_idx] = RowType.DIVISION_2
                        division_stack = division_stack[:1] + [row_idx]
                    elif division_level == 3:
                        self.row_types[row_idx] = RowType.DIVISION_3
                        division_stack = division_stack[:2] + [row_idx]
                    elif division_level == 4:
                        self.row_types[row_idx] = RowType.DIVISION_4
                        division_stack = division_stack[:3] + [row_idx]
                    else:
                        self.row_types[row_idx] = RowType.DIVISION_5
                        division_stack = division_stack[:4] + [row_idx]
                    
                    self.checked_rows.add(row_idx)
            else:
                # 其他情况 - 标记为无效，后续合并行检测会处理
                self.row_types[row_idx] = RowType.INVALID
        
        # 检测合并行（在清单行之后，有特征描述内容）
        self._detect_merge_rows(start_row)
        
        # 检测表尾（从后往前找，通常是汇总信息）
        self._detect_footer_rows()
        
        # 更新表格显示
        self.update_preview_table()
        
        # 勾选需要导入的行
        self._update_checked_rows()
        
        valid_count = len([t for t in self.row_types.values() if t in [RowType.ITEM, RowType.DIVISION_1, RowType.DIVISION_2, RowType.DIVISION_3, RowType.DIVISION_4, RowType.DIVISION_5, RowType.MERGE]])
        total_rows = len(self.preview_rows)
        self.status_label.setText(f"✓ 行识别完成，共 {total_rows} 行，识别出 {valid_count} 行有效数据（已自动勾选）")
    
    def _is_division_name_format(self, name_val: str) -> bool:
        """判断项目名称是否符合分部命名格式
        
        分部行的命名通常有以下特征：
        1. 以中文数字开头（一、二、三...）
        2. 以阿拉伯数字+点开头（1. 2. 1.1.）
        3. 以括号数字开头（(1) (2) 或 1) 2)）
        4. 包含"分部"、"章"、"节"等关键词
        5. 名称较短（通常不超过20字）
        """
        import re
        
        # 如果名称太长，不太可能是分部行
        if len(name_val) > 30:
            return False
        
        # 中文数字开头
        chinese_nums = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
        for num in chinese_nums:
            if name_val.startswith(num):
                return True
        
        # 阿拉伯数字+点开头
        if re.match(r'^\d+\.', name_val):
            return True
        
        # 括号数字开头
        if re.match(r'^[\(（]\d+[\)）]', name_val):
            return True
        
        # 数字+右括号
        if re.match(r'^\d+\)', name_val):
            return True
        
        # 包含分部关键词
        division_keywords = ['分部', '章', '节', '项', '目']
        for keyword in division_keywords:
            if keyword in name_val:
                return True
        
        return False
    
    def _detect_division_level(self, name_val: str, division_stack: list) -> int:
        """检测分部级别
        
        根据名称特征判断分部层级：
        - 一级分部：以中文数字开头（一、二、三...）
        - 二级分部：以阿拉伯数字+点开头（1. 2. 3.）
        - 三级分部：以括号数字开头（(1) (2) 或 1) 2)）
        - 四级分部：以双括号或特殊标记
        - 五级分部：其他格式
        """
        import re
        
        # 中文数字开头 - 一级分部
        chinese_nums = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
        
        # 检查是否是中文数字开头（可能后跟、或.）
        for num in chinese_nums:
            if name_val.startswith(num):
                # 检查是否是独立的章节号
                if len(name_val) > 1 and name_val[1] in ['、', '.', ' ']:
                    return 1
        
        # 阿拉伯数字+点开头 - 二级分部 (1. 2. 1.1. 等)
        if re.match(r'^\d+\.', name_val):
            # 检查小数点数量判断级别
            dots = name_val.count('.')
            if dots == 1:
                # 检查数字位数
                num_part = name_val.split('.')[0]
                if num_part.isdigit() and 1 <= int(num_part) <= 99:
                    return 2
            elif dots == 2:
                return 3
            elif dots >= 3:
                return 4
        
        # 括号数字开头 - 三级或四级分部
        if re.match(r'^[\(（]\d+[\)）]', name_val):
            return 3
        
        # 数字+右括号 - 也是三级分部
        if re.match(r'^\d+\)', name_val):
            return 3
        
        # 如果栈为空，默认为一级
        if not division_stack:
            return 1
        
        # 根据当前栈深度判断
        current_depth = len(division_stack)
        if current_depth < 5:
            return current_depth + 1
        else:
            return 5
    
    def _detect_merge_rows(self, start_row: int):
        """检测合并行
        
        合并行的特征：
        1. 在清单行之后
        2. 有内容（描述、规格等）但没有单位/工程量
        3. 可能有多行连续的情况
        """
        name_col = self.column_mappings.get('name')
        desc_col = self.column_mappings.get('description')
        spec_col = self.column_mappings.get('specification')
        unit_col = self.column_mappings.get('unit')
        quantity_col = self.column_mappings.get('quantity')
        
        if name_col is None:
            return
        
        prev_item_row = None
        merge_sequence = []  # 记录连续的合并行
        
        for row_idx in range(start_row, len(self.preview_rows)):
            current_type = self.row_types.get(row_idx)
            
            if current_type == RowType.ITEM:
                # 遇到新的清单行，先处理之前的合并行序列
                if prev_item_row is not None and merge_sequence:
                    for merge_row in merge_sequence:
                        if self.row_types.get(merge_row) == RowType.INVALID:
                            self.row_types[merge_row] = RowType.MERGE
                            self.checked_rows.add(merge_row)
                    merge_sequence = []
                
                prev_item_row = row_idx
                merge_sequence = []
                
            elif prev_item_row is not None and current_type == RowType.INVALID:
                row_data = self.preview_rows[row_idx]
                name_val = str(row_data[name_col]).strip() if name_col < len(row_data) else ""
                desc_val = str(row_data[desc_col]).strip() if desc_col and desc_col < len(row_data) else ""
                spec_val = str(row_data[spec_col]).strip() if spec_col and spec_col < len(row_data) else ""
                unit_val = str(row_data[unit_col]).strip() if unit_col and unit_col < len(row_data) else ""
                quantity_val = str(row_data[quantity_col]).strip() if quantity_col and quantity_col < len(row_data) else ""
                
                # 合并行特征：
                # 1. 有项目特征内容（这是主要特征）
                # 2. 没有单位和工程量
                # 3. 可能有项目名称（作为描述的延续）
                # 简化判断：只要有项目特征内容且无单位工程量，就是合并行
                has_desc_content = bool(desc_val and len(desc_val) > 0)
                no_unit_qty = not unit_val and not quantity_val
                
                is_merge_row = has_desc_content and no_unit_qty
                
                if is_merge_row:
                    merge_sequence.append(row_idx)
                else:
                    # 不满足合并行条件，处理之前的序列
                    if merge_sequence:
                        for merge_row in merge_sequence:
                            if self.row_types.get(merge_row) == RowType.INVALID:
                                self.row_types[merge_row] = RowType.MERGE
                                self.checked_rows.add(merge_row)
                        merge_sequence = []
        
        # 处理最后一批合并行
        if prev_item_row is not None and merge_sequence:
            for merge_row in merge_sequence:
                if self.row_types.get(merge_row) == RowType.INVALID:
                    self.row_types[merge_row] = RowType.MERGE
                    self.checked_rows.add(merge_row)
    
    def _detect_footer_rows(self):
        """检测表尾行"""
        # 从后往前找，通常是包含"合计"、"总计"等关键词的行
        footer_keywords = ['合计', '总计', '汇总', '小计', '共']
        
        for row_idx in range(len(self.preview_rows) - 1, -1, -1):
            if self.row_types.get(row_idx) != RowType.UNKNOWN:
                continue
            
            row_data = self.preview_rows[row_idx]
            row_text = ' '.join(str(cell).lower() for cell in row_data if cell)
            
            for keyword in footer_keywords:
                if keyword in row_text:
                    self.row_types[row_idx] = RowType.INVALID
                    break
    
    def _update_checked_rows(self):
        """更新勾选状态到表格"""
        for table_row in range(self.preview_table.rowCount()):
            if table_row == 0:  # 跳过列识别行
                continue
            data_row_idx = table_row - 1  # 转换为数据行索引
            checkbox_item = self.preview_table.item(table_row, 0)
            if checkbox_item:
                if data_row_idx in self.checked_rows:
                    checkbox_item.setCheckState(Qt.Checked)
                else:
                    checkbox_item.setCheckState(Qt.Unchecked)
    
    def mark_selected_rows(self, row_type: RowType):
        """标记选中的行为指定类型"""
        # 获取选中的行（表格行索引）
        selected_table_rows = set()
        for item in self.preview_table.selectedItems():
            table_row = item.row()
            selected_table_rows.add(table_row)
        
        # 如果没有选中行，使用勾选行
        if not selected_table_rows:
            for table_row in range(self.preview_table.rowCount()):
                checkbox_item = self.preview_table.item(table_row, 0)
                if checkbox_item and checkbox_item.checkState() == Qt.Checked:
                    selected_table_rows.add(table_row)
        
        # 标记行类型
        # 注意：表格第0行是列识别行，第1行开始是数据行（对应preview_rows[0]）
        for table_row in selected_table_rows:
            if table_row == 0:  # 跳过列识别行
                continue
            
            data_row_idx = table_row - 1  # 转换为数据行索引
            self.row_types[data_row_idx] = row_type
            
            # 更新标识列
            type_item = QTableWidgetItem(f"{self.ROW_TYPE_STYLES[row_type]['icon']} {row_type.value}")
            type_item.setData(Qt.UserRole, row_type)
            type_item.setFlags(type_item.flags() & ~Qt.ItemIsEditable)
            self.preview_table.setItem(table_row, 1, type_item)
            
            # 应用样式
            self.apply_row_style(table_row, row_type)
        
        self.status_label.setText(f"✓ 已标记 {len(selected_table_rows)} 行")
    
    def select_all_rows(self):
        """全选所有行"""
        for table_row in range(self.preview_table.rowCount()):
            if table_row == 0:  # 跳过列识别行
                continue
            checkbox_item = self.preview_table.item(table_row, 0)
            if checkbox_item:
                checkbox_item.setCheckState(Qt.Checked)
                data_row_idx = table_row - 1
                self.checked_rows.add(data_row_idx)
    
    def invert_selection(self):
        """反选所有行"""
        for table_row in range(self.preview_table.rowCount()):
            if table_row == 0:  # 跳过列识别行
                continue
            checkbox_item = self.preview_table.item(table_row, 0)
            if checkbox_item:
                current = checkbox_item.checkState()
                new_state = Qt.Unchecked if current == Qt.Checked else Qt.Checked
                checkbox_item.setCheckState(new_state)
                data_row_idx = table_row - 1
                if new_state == Qt.Checked:
                    self.checked_rows.add(data_row_idx)
                else:
                    self.checked_rows.discard(data_row_idx)
    
    def clear_all(self):
        """清除所有标记"""
        self.row_types = {i: RowType.UNKNOWN for i in range(len(self.preview_rows))}
        self.checked_rows = set()
        self.column_mappings = {}
        self.column_detection_done = False
        self.auto_detect_row_btn.setEnabled(False)
        self.auto_detect_row_btn.setToolTip("先识别列后才能识别行")
        self.update_preview_table()
        self.status_label.setText("✓ 已清除所有标记，请重新识别")
    
    def show_context_menu(self, position):
        """显示右键菜单"""
        menu = QMenu(self)
        menu.setObjectName("contextMenu")
        
        # 行类型子菜单
        row_type_menu = menu.addMenu("标记为")
        
        row_types = [
            ("表头", RowType.HEADER),
            ("表尾", RowType.FOOTER),
            ("无效行", RowType.INVALID),
            ("分部一级", RowType.DIVISION_1),
            ("分部二级", RowType.DIVISION_2),
            ("分部三级", RowType.DIVISION_3),
            ("分部四级", RowType.DIVISION_4),
            ("分部五级", RowType.DIVISION_5),
            ("清单行", RowType.ITEM),
            ("合并行", RowType.MERGE),
        ]
        
        for label, row_type in row_types:
            action = row_type_menu.addAction(label)
            action.triggered.connect(lambda checked, rt=row_type: self.mark_selected_rows(rt))
        
        menu.addSeparator()
        
        # 全选/反选
        select_all_action = menu.addAction("☑️ 全选")
        select_all_action.triggered.connect(self.select_all_rows)
        
        invert_action = menu.addAction("🔃 反选")
        invert_action.triggered.connect(self.invert_selection)
        
        menu.exec_(self.preview_table.viewport().mapToGlobal(position))
    
    def on_sheet_changed(self, index):
        """工作簿改变"""
        if index >= 0 and self.sheet_names:
            sheet_name = self.sheet_names[index]
            if sheet_name != self.current_sheet_name:
                self.load_sheet_async(sheet_name)
    
    def load_sheet_async(self, sheet_name: str):
        """异步加载指定工作表 - 加载全部数据"""
        if not sheet_name:
            return
        
        self.current_sheet_name = sheet_name
        
        self.progress_bar.show()
        self.status_label.setText(f"正在加载工作表: {sheet_name}...")
        self.import_btn.setEnabled(False)
        
        # 使用新的完整数据加载器
        self.load_worker = FullDataLoadWorker(self.excel_path, sheet_name)
        self.load_worker.data_loaded.connect(self.on_full_data_loaded)
        self.load_worker.error.connect(self.on_load_error)
        self.load_worker.start()
    
    def on_full_data_loaded(self, all_rows: list, row_count: int):
        """完整数据加载完成回调"""
        self.preview_rows = all_rows
        
        # 初始化行类型
        self.row_types = {i: RowType.UNKNOWN for i in range(len(all_rows))}
        self.checked_rows = set()
        self.column_detection_done = False
        self.auto_detect_row_btn.setEnabled(False)
        
        # 更新预览表格
        self.update_preview_table()
        
        self.progress_bar.hide()
        self.status_label.setText(f"✓ 已加载 {row_count} 行数据，请先点击'自动识别列'")
        self.import_btn.setEnabled(True)
        
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
    
    def on_load_error(self, error_msg: str):
        """加载错误回调"""
        self.progress_bar.hide()
        self.status_label.setText("加载失败")
        self.import_btn.setEnabled(True)
        MessageDialog.critical(self, "错误", f"无法加载Excel文件:\n{error_msg}")
        
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
    
    def on_import(self):
        """确认导入"""
        # 收集要导入的数据
        imported_items = []
        
        # 获取列映射
        name_col = self.column_mappings.get('name')
        if name_col is None:
            MessageDialog.warning(self, "提示", "请先设置项目名称列映射")
            return
        
        # 遍历所有行，导入清单行和分部行
        current_division = None
        current_item = None
        merge_buffer = []
        
        for row_idx, row_data in enumerate(self.preview_rows):
            row_type = self.row_types.get(row_idx, RowType.UNKNOWN)
            
            # 跳过无效行
            if row_type in [RowType.HEADER, RowType.FOOTER, RowType.INVALID]:
                continue
            
            # 获取行数据
            item_data = self._extract_row_data(row_data)
            
            if row_type in [RowType.DIVISION_1, RowType.DIVISION_2, RowType.DIVISION_3,
                           RowType.DIVISION_4, RowType.DIVISION_5]:
                # 分部行
                item_data['level'] = self._get_division_level(row_type)
                item_data['is_division'] = True
                imported_items.append(item_data)
                current_division = item_data
                current_item = None
                merge_buffer = []
                
            elif row_type == RowType.ITEM:
                # 清单行 - 先处理之前的合并行
                if merge_buffer and current_item:
                    # 将合并行内容附加到上一个清单行
                    self._apply_merge_rows(current_item, merge_buffer)
                
                item_data['level'] = 2 if current_division else 1
                item_data['is_division'] = False
                imported_items.append(item_data)
                current_item = item_data
                merge_buffer = []
                
            elif row_type == RowType.MERGE:
                # 合并行 - 缓存起来
                merge_buffer.append(item_data)
        
        # 处理最后的合并行
        if merge_buffer and current_item:
            self._apply_merge_rows(current_item, merge_buffer)
        
        # 存储导入的数据
        self._imported_data = imported_items
        
        MessageDialog.information(self, "导入成功", f"成功导入 {len(imported_items)} 条数据")
        self.accept()
    
    def _extract_row_data(self, row_data: List) -> Dict:
        """从行数据中提取字段值"""
        item_data = {}
        
        for field_key, col_idx in self.column_mappings.items():
            if col_idx < len(row_data):
                value = row_data[col_idx]
                
                # 数值字段处理
                numeric_fields = ['quantity', 'unit_price', 'labor_unit_price',
                                'material_unit_price', 'material_loss_rate',
                                'auxiliary_unit_price', 'machine_unit_price', 'other_unit_price']
                if field_key in numeric_fields:
                    try:
                        value = float(value) if value is not None else 0.0
                    except (ValueError, TypeError):
                        value = 0.0
                else:
                    value = str(value).strip() if value is not None else ""
                
                item_data[field_key] = value
            else:
                # 默认值
                if field_key in ['quantity', 'unit_price']:
                    item_data[field_key] = 0.0
                else:
                    item_data[field_key] = ""
        
        return item_data
    
    def _get_division_level(self, row_type: RowType) -> int:
        """获取分部层级"""
        level_map = {
            RowType.DIVISION_1: 1,
            RowType.DIVISION_2: 2,
            RowType.DIVISION_3: 3,
            RowType.DIVISION_4: 4,
            RowType.DIVISION_5: 5,
        }
        return level_map.get(row_type, 1)
    
    def _apply_merge_rows(self, item_data: Dict, merge_rows: List[Dict]):
        """将合并行内容应用到清单行"""
        for merge_row in merge_rows:
            # 合并描述字段
            merge_desc = merge_row.get('description', '')
            if merge_desc:
                current_desc = item_data.get('description', '')
                if current_desc:
                    item_data['description'] = current_desc + '\n' + merge_desc
                else:
                    item_data['description'] = merge_desc
            
            # 合并其他字段（如果有值）
            for field_key in ['specification', 'remark']:
                merge_val = merge_row.get(field_key, '')
                if merge_val and not item_data.get(field_key):
                    item_data[field_key] = merge_val
    
    def get_imported_data(self) -> List[Dict]:
        """获取导入的数据"""
        return getattr(self, '_imported_data', [])
    
    def closeEvent(self, event):
        """关闭事件"""
        if self.load_worker:
            self.load_worker.stop()
            self.load_worker = None
        event.accept()
    
    # ========== 鼠标拖动支持 ==========
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drag_pos = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()
    
    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and self.drag_pos is not None:
            self.move(event.globalPos() - self.drag_pos)
            event.accept()
