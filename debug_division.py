#!/usr/bin/env python3
"""调试分部行被错误合并的问题"""
from openpyxl import load_workbook
import re
from typing import Dict

def _should_merge_with_previous(current_row: tuple, previous_item: Dict, mapping: Dict[str, int]) -> bool:
    """当前的合并判断逻辑"""
    if not previous_item:
        return False
    
    name_col = mapping.get('name')
    desc_col = mapping.get('description')
    sequence_col = mapping.get('sequence')
    code_col = mapping.get('code')
    
    current_name = str(current_row[name_col]).strip() if name_col is not None and name_col < len(current_row) and current_row[name_col] else ""
    current_desc = str(current_row[desc_col]).strip() if desc_col is not None and desc_col < len(current_row) and current_row[desc_col] else ""
    current_seq = str(current_row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(current_row) and current_row[sequence_col] else ""
    current_code = str(current_row[code_col]).strip() if code_col is not None and code_col < len(current_row) and current_row[code_col] else ""
    
    prev_name = str(previous_item.get('name', '')).strip()
    prev_desc = str(previous_item.get('description', '')).strip()
    
    print(f"  当前行: seq='{current_seq}', code='{current_code}', name='{current_name[:30]}...'")
    print(f"  前一行: name='{prev_name[:30]}...', desc='{prev_desc[:30]}...'")
    
    # 如果当前行有项目编码，说明是新记录，不合并
    if current_code and re.match(r'^\d{9,}$', current_code.replace('-', '')):
        print(f"  → 有项目编码，不合并")
        return False
    
    # 如果当前行有序号且是数字格式，说明是新记录，不合并
    if current_seq and re.match(r'^\d+$', current_seq):
        print(f"  → 有序号，不合并")
        return False
    
    # 如果当前行没有序号，但有项目名称，可能是前一行的延续
    if not current_seq and current_name and not current_code:
        print(f"  → 无序号有名称，合并")
        return True
    
    # 如果当前行项目名称非常短（少于3个字符），可能是被截断的
    if len(current_name) < 3 and len(prev_name) > 0:
        print(f"  → 名称太短，合并")
        return True
    
    # 如果前一行项目名称以不完整的方式结束
    incomplete_endings_name = ['混凝', '预拌', '泵送', '强度', '等级', '基础', '独立', '类型', '种类', '形式']
    for ending in incomplete_endings_name:
        if prev_name.endswith(ending):
            print(f"  → 前一行名称以'{ending}'结尾，合并")
            return True
    
    # 如果前一行项目特征描述以不完整的方式结束
    incomplete_endings_desc = ['混凝', '预拌', '泵送', '强度等级:', '基础类型:', '混凝土种类:', '种类:', '等级:', '类型:']
    for ending in incomplete_endings_desc:
        if prev_desc.endswith(ending):
            print(f"  → 前一行描述以'{ending}'结尾，合并")
            return True
    
    print(f"  → 不合并")
    return False

# 加载Excel文件
wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

mapping = {
    'sequence': 0,
    'code': 1,
    'name': 3,
    'description': 5,
}

# 查找钢结构、装饰装修、门窗等行
print("=== 查找可能被错误合并的分部行 ===\n")

for i in range(40, 80):  # 查看40-80行
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    name = str(row[3]).strip() if row[3] else ''
    
    if any(keyword in name for keyword in ['钢结构', '装饰装修', '门窗', '屋面', '防水', '保温']):
        print(f"\n行{i}: {name}")
        print(f"  完整行数据: seq={row[0]}, code={row[1]}, name={name}, desc={row[5]}")
        
        # 模拟前一行
        prev_row = list(sheet.iter_rows(min_row=i-1, max_row=i-1, values_only=True))[0]
        previous_item = {
            'sequence': str(prev_row[0]) if prev_row[0] else '',
            'code': str(prev_row[1]) if prev_row[1] else '',
            'name': str(prev_row[3]) if prev_row[3] else '',
            'description': str(prev_row[5]) if prev_row[5] else '',
        }
        
        print(f"\n  检查是否应该合并:")
        should_merge = _should_merge_with_previous(row, previous_item, mapping)
        print(f"  结果: {'合并' if should_merge else '不合并'}")

wb.close()
