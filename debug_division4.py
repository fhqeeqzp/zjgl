#!/usr/bin/env python3
"""调试为什么钢结构等没有被识别为分部"""
from openpyxl import load_workbook

wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

print("=== 检查钢结构、装饰装修、门窗等行 ===\n")

division_keywords = ['工程', '分部', '分项', '措施项目']

for i in range(95, 160):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    seq = str(row[0]).strip() if row[0] else ''
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[3]).strip() if row[3] else ''
    
    if name:
        # 检查是否包含分部关键词
        has_keyword = any(keyword in name for keyword in division_keywords)
        # 检查是否没有序号和编码
        no_seq_code = not seq and not code
        
        if has_keyword:
            print(f"行{i:3d}: seq='{seq}', code='{code}', name='{name}'")
            print(f"       包含关键词: {has_keyword}, 无序号编码: {no_seq_code}")
            if has_keyword and no_seq_code:
                print(f"       → 应该识别为分部")
            else:
                print(f"       → 未识别为分部的原因: {'有序号或编码' if not no_seq_code else ''}")

wb.close()
