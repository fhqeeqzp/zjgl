#!/usr/bin/env python3
"""调试合并逻辑 - 查找正确的续行"""
from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

# 读取第15-30行
print("=== 查看第15-30行 ===")
for i in range(15, 31):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    seq = str(row[0]) if row[0] else ''
    code = str(row[1]) if row[1] else ''
    name = str(row[3])[:30] if row[3] else ''
    desc = str(row[5])[:40] if row[5] else ''
    print(f"行{i:2d}: seq={seq:4s} code={code:15s} name={name:30s} desc={desc:40s}")

wb.close()
