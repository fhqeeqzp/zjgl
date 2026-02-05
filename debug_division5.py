#!/usr/bin/env python3
"""查找所有包含关键词的行"""
from openpyxl import load_workbook

wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

print("=== 查找钢结构、装饰装修、门窗 ===\n")

keywords = ['钢结构', '装饰装修', '门窗']

for i in range(1, 200):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    name = str(row[3]).strip() if row[3] else ''
    
    for keyword in keywords:
        if keyword in name:
            seq = str(row[0]).strip() if row[0] else ''
            code = str(row[1]).strip() if row[1] else ''
            print(f"行{i:3d}: seq='{seq}', code='{code}', name='{name}'")

wb.close()
