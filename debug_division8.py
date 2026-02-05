#!/usr/bin/env python3
"""查看'超过1m 钢支撑'和'径(mm) 20'的前一行"""
from openpyxl import load_workbook

wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

print("=== 查看'超过1m 钢支撑'相关行 ===\n")
for i in range(25, 32):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    seq = str(row[0]).strip() if row[0] else ''
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[3]).strip() if row[3] else ''
    desc = str(row[5]).strip() if row[5] else ''
    marker = " <-- " if i in [26, 28] else ""
    print(f"行{i:3d}: seq='{seq}', code='{code}', name='{name}'{marker}")

print("\n=== 查看'径(mm) 20'相关行 ===\n")
for i in range(90, 97):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    seq = str(row[0]).strip() if row[0] else ''
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[3]).strip() if row[3] else ''
    desc = str(row[5]).strip() if row[5] else ''
    marker = " <-- " if i in [85, 94] else ""
    print(f"行{i:3d}: seq='{seq}', code='{code}', name='{name}'{marker}")

wb.close()
