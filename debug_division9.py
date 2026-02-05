#!/usr/bin/env python3
"""查找'径(mm) 20'应该合并的前一行"""
from openpyxl import load_workbook

wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

print("=== 查找'径(mm) 20'相关行 ===\n")

# 查看80-95行
for i in range(80, 96):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    seq = str(row[0]).strip() if row[0] else ''
    code = str(row[1]).strip() if row[1] else ''
    name = str(row[3]).strip() if row[3] else ''
    desc = str(row[5]).strip() if row[5] else ''
    
    # 只显示有内容的行
    if name or seq or code:
        print(f"行{i:3d}: seq='{seq}', code='{code}', name='{name}'")
        if desc:
            print(f"       desc='{desc[:50]}...'")

wb.close()
