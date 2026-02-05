#!/usr/bin/env python3
"""调试分部识别逻辑"""
from openpyxl import load_workbook

wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

mapping = {
    'sequence': 0,
    'code': 1,
    'name': 3,
    'description': 5,
}

def _is_division_row(row: tuple, mapping: dict) -> bool:
    name_col = mapping.get('name')
    sequence_col = mapping.get('sequence')
    code_col = mapping.get('code')
    
    if name_col is None or name_col >= len(row):
        return False
    
    name_val = str(row[name_col]).strip() if row[name_col] else ""
    if not name_val:
        return False
    
    division_keywords = ['工程', '分部', '分项', '措施项目']
    is_division = any(keyword in name_val for keyword in division_keywords)
    
    seq_val = str(row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(row) and row[sequence_col] else ""
    code_val = str(row[code_col]).strip() if code_col is not None and code_col < len(row) and row[code_col] else ""
    
    print(f"  名称: '{name_val}', 包含关键词: {is_division}")
    print(f"  序号: '{seq_val}', 编码: '{code_val}'")
    print(f"  无序号: {not seq_val}, 无编码: {not code_val}")
    
    if is_division and not seq_val and not code_val:
        print(f"  → 是分部")
        return True
    
    print(f"  → 不是分部")
    return False

# 测试行100
print("=== 测试行100: 钢结构 ===")
row100 = list(sheet.iter_rows(min_row=100, max_row=100, values_only=True))[0]
_is_division_row(row100, mapping)

print("\n=== 测试行111: 装饰装修 ===")
row111 = list(sheet.iter_rows(min_row=111, max_row=111, values_only=True))[0]
_is_division_row(row111, mapping)

print("\n=== 测试行154: 门窗 ===")
row154 = list(sheet.iter_rows(min_row=154, max_row=154, values_only=True))[0]
_is_division_row(row154, mapping)

wb.close()
