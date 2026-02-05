#!/usr/bin/env python3
"""查看100-150行"""
from openpyxl import load_workbook

wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
sheet = wb['1.辅助斜坡道空气预热间土建']

print("=== 查看所有行（100-150行）===\n")

for i in range(100, 151):
    row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    seq = str(row[0])[:5] if row[0] else ''
    code = str(row[1])[:12] if row[1] else ''
    name = str(row[3])[:40] if row[3] else ''
    desc = str(row[5])[:30] if row[5] else ''
    
    # 只显示有内容的行
    if name:
        print(f"行{i:3d}: seq={seq:5s} code={code:12s} name={name:40s}")

wb.close()
