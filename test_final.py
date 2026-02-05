#!/usr/bin/env python3
"""最终测试 - 验证修复结果"""
import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook
import re
from typing import Dict, List

def _is_header_row(row: tuple, mapping: Dict[str, int]) -> bool:
    header_title_patterns = [
        '分部分项工程和单价措施项目清单与计价表',
        '分部分项工程',
        '工程名称：',
        '标段：',
    ]
    header_column_keywords = ['序号', '项目编码', '项目名称', '项目特征', '计量单位', '工程量', '综合单价', '合价']
    summary_keywords = ['本页小计', '合计', '总计', '小计', '共', '页', '表—']
    
    name_col = mapping.get('name')
    desc_col = mapping.get('description')
    sequence_col = mapping.get('sequence')
    code_col = mapping.get('code')
    
    if name_col is not None and name_col < len(row):
        name_val = str(row[name_col]).strip() if row[name_col] else ""
        for keyword in summary_keywords:
            if keyword in name_val:
                return True
        for pattern in header_title_patterns:
            if pattern in name_val:
                return True
        if '项目名称' in name_val and len(name_val) < 15:
            return True
    
    if desc_col is not None and desc_col < len(row):
        desc_val = str(row[desc_col]).strip() if row[desc_col] else ""
        for keyword in summary_keywords:
            if keyword in desc_val:
                return True
        if '项目特征描述' in desc_val and len(desc_val) < 15:
            return True
    
    header_col_count = 0
    for col_idx in [sequence_col, code_col, name_col, desc_col]:
        if col_idx is not None and col_idx < len(row):
            cell_val = str(row[col_idx]).strip() if row[col_idx] else ""
            for keyword in header_column_keywords:
                if keyword in cell_val:
                    header_col_count += 1
                    break
    
    if header_col_count >= 2:
        return True
    
    if sequence_col is not None and sequence_col < len(row):
        seq_val = str(row[sequence_col]).strip() if row[sequence_col] else ""
        if seq_val == '序号' or '序号' in seq_val:
            return True
    
    return False

def _is_division_row(row: tuple, mapping: Dict[str, int]) -> bool:
    name_col = mapping.get('name')
    sequence_col = mapping.get('sequence')
    code_col = mapping.get('code')
    
    if name_col is None or name_col >= len(row):
        return False
    
    name_val = str(row[name_col]).strip() if row[name_col] else ""
    if not name_val:
        return False
    
    division_keywords = ['工程', '分部', '分项', '措施项目']
    is_division = any(keyword in name_val for keyword in division_keywords)
    
    seq_val = str(row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(row) and row[sequence_col] else ""
    code_val = str(row[code_col]).strip() if code_col is not None and code_col < len(row) and row[code_col] else ""
    
    if is_division and not seq_val and not code_val:
        return True
    
    return False

def _should_merge_with_previous(current_row: tuple, previous_item: Dict, mapping: Dict[str, int]) -> bool:
    """修复后的合并判断逻辑"""
    if not previous_item:
        return False
    
    name_col = mapping.get('name')
    desc_col = mapping.get('description')
    sequence_col = mapping.get('sequence')
    code_col = mapping.get('code')
    
    current_name = str(current_row[name_col]).strip() if name_col is not None and name_col < len(current_row) and current_row[name_col] else ""
    current_desc = str(current_row[desc_col]).strip() if desc_col is not None and desc_col < len(current_row) and current_row[desc_col] else ""
    current_seq = str(current_row[sequence_col]).strip() if sequence_col is not None and sequence_col < len(current_row) and current_row[sequence_col] else ""
    current_code = str(current_row[code_col]).strip() if code_col is not None and code_col < len(current_row) and current_row[code_col] else ""
    
    prev_name = str(previous_item.get('name', '')).strip()
    prev_desc = str(previous_item.get('description', '')).strip()
    
    # 如果当前行有项目编码，说明是新记录，不合并
    if current_code and re.match(r'^\d{9,}$', current_code.replace('-', '')):
        return False
    
    # 如果当前行有序号且是数字格式，说明是新记录，不合并
    if current_seq and re.match(r'^\d+$', current_seq):
        return False
    
    # 检查前一行是否以不完整的方式结束（这是合并的前提条件）
    prev_is_incomplete = False
    
    incomplete_endings_name = ['混凝', '预拌', '泵送', '强度', '等级', '基础', '独立', '类型', '种类', '形式']
    for ending in incomplete_endings_name:
        if prev_name.endswith(ending):
            prev_is_incomplete = True
            break
    
    if not prev_is_incomplete:
        incomplete_endings_desc = ['混凝', '预拌', '泵送', '强度等级:', '基础类型:', '混凝土种类:', '种类:', '等级:', '类型:']
        for ending in incomplete_endings_desc:
            if prev_desc.endswith(ending):
                prev_is_incomplete = True
                break
    
    # 如果前一行是完整的（如"钢结构"、"装饰装修"等分部行），则不合并
    if not prev_is_incomplete:
        return False
    
    # 如果当前行没有序号，但有项目名称，且前一行不完整，可能是前一行的延续
    if not current_seq and current_name and not current_code:
        return True
    
    # 如果当前行项目名称非常短（少于3个字符），且前一行不完整
    if len(current_name) < 3 and len(prev_name) > 0:
        return True
    
    continuation_starts = ['土', '送', '级', '凝', '础', '类', '型']
    if current_name and current_name[0] in continuation_starts:
        for ending in ['混凝', '预拌', '种类:', '类型:', '等级:']:
            if prev_name.endswith(ending) or prev_desc.endswith(ending):
                return True
    
    if current_desc and prev_desc:
        desc_continuations = [
            ('混凝', '土'),
            ('预拌', '混凝'),
            ('种类:', '预拌'),
            ('等级:', 'C'),
        ]
        for prev_ending, curr_start in desc_continuations:
            if prev_desc.endswith(prev_ending) and current_desc.startswith(curr_start):
                return True
    
    return False

def _merge_rows(previous_item: Dict, current_row: tuple, mapping: Dict[str, int]) -> Dict:
    merged = previous_item.copy()
    
    name_col = mapping.get('name')
    desc_col = mapping.get('description')
    
    if name_col is not None and name_col < len(current_row):
        current_name = str(current_row[name_col]).strip() if current_row[name_col] else ""
        if current_name:
            prev_name = str(previous_item.get('name', '')).strip()
            if prev_name and current_name:
                overlap = 0
                for i in range(1, min(len(prev_name), len(current_name)) + 1):
                    if prev_name[-i:] == current_name[:i]:
                        overlap = i
                if overlap > 0:
                    merged['name'] = prev_name + current_name[overlap:]
                else:
                    merged['name'] = prev_name + current_name
    
    if desc_col is not None and desc_col < len(current_row):
        current_desc = str(current_row[desc_col]).strip() if current_row[desc_col] else ""
        if current_desc:
            prev_desc = str(previous_item.get('description', '')).strip()
            if prev_desc and current_desc:
                overlap = 0
                for i in range(1, min(len(prev_desc), len(current_desc)) + 1):
                    if prev_desc[-i:] == current_desc[:i]:
                        overlap = i
                if overlap > 0:
                    merged['description'] = prev_desc + current_desc[overlap:]
                else:
                    merged['description'] = prev_desc + current_desc
    
    return merged

def parse_sequence_level(sequence: str) -> int:
    if not sequence:
        return 1
    sequence = str(sequence).strip()
    dots = sequence.count('.')
    return dots + 1

def test_import():
    print("=== 最终导入测试（修复版）===\n")
    
    wb = load_workbook('WD/陈台沟填充站项目/投标附件/土建工程工程量清单.xlsx', data_only=True)
    sheet = wb['1.辅助斜坡道空气预热间土建']
    
    mapping = {
        'sequence': 0,
        'code': 1,
        'name': 3,
        'description': 5,
    }
    
    imported_items = []
    current_division = None
    pending_item = None
    
    start_row = 5
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if _is_header_row(row, mapping):
            continue
        
        if _is_division_row(row, mapping):
            name_col = mapping.get('name')
            if name_col is not None and name_col < len(row) and row[name_col]:
                division_name = str(row[name_col]).strip()
                current_division = division_name
                division_item = {
                    'sequence': '',
                    'name': division_name,
                    'specification': '',
                    'description': '',
                    'unit': '',
                    'quantity': 0.0,
                    'unit_price': 0.0,
                    'level': 2,
                    'remark': '',
                    'is_division': True
                }
                if pending_item:
                    imported_items.append(pending_item)
                    pending_item = None
                imported_items.append(division_item)
            continue
        
        item_data = {}
        for field_key, col_idx in mapping.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    item_data[field_key] = str(value).strip()
                else:
                    item_data[field_key] = ""
            else:
                item_data[field_key] = ""
        
        name = str(item_data.get('name', '')).strip()
        if not name:
            continue
        
        sequence = str(item_data.get('sequence', ''))
        if sequence and sequence.strip():
            base_level = parse_sequence_level(sequence)
            if current_division:
                item_data['level'] = base_level + 2
            else:
                item_data['level'] = base_level + 1
        else:
            code = str(item_data.get('code', '')).strip()
            if code and re.match(r'^\d{9,}$', code.replace('-', '')):
                item_data['level'] = 3 if current_division else 2
            else:
                item_data['level'] = 1
        
        if pending_item and _should_merge_with_previous(row, pending_item, mapping):
            pending_item = _merge_rows(pending_item, row, mapping)
        else:
            if pending_item:
                imported_items.append(pending_item)
            pending_item = item_data
    
    if pending_item:
        imported_items.append(pending_item)
    
    wb.close()
    
    print(f"共导入 {len(imported_items)} 条数据\n")
    
    print("前30条数据预览:")
    print("-" * 100)
    for i, item in enumerate(imported_items[:30], 1):
        level = item.get('level', 1)
        indent = "  " * (level - 1)
        name = item.get('name', '')[:50]
        is_division = item.get('is_division', False)
        marker = "【分部】" if is_division else ""
        print(f"{i:3d}. {indent}{marker}{name}")
    
    print("\n" + "=" * 100)
    print("关键数据验证:")
    
    # 检查"现浇混凝土基础 独立基础 混凝土"
    found = False
    for item in imported_items:
        if '现浇混凝土基础 独立基础 混凝土' in item.get('name', ''):
            print(f"\n✓ 找到合并后的项目: {item['name']}")
            print(f"  项目特征: {item['description'][:80]}...")
            found = True
            break
    
    if not found:
        print("\n✗ 未找到合并后的'现浇混凝土基础 独立基础 混凝土'")
    
    # 检查分部
    divisions = [item for item in imported_items if item.get('is_division')]
    print(f"\n✓ 找到 {len(divisions)} 个分部:")
    for div in divisions:
        print(f"  - {div['name']}")
    
    # 验证关键分部是否独立
    print("\n✓ 验证关键分部是否独立（不应被合并）:")
    key_divisions = ['钢结构', '装饰装修', '门窗']
    for key in key_divisions:
        found_items = [item for item in imported_items if key in item.get('name', '')]
        for item in found_items:
            if item.get('is_division'):
                print(f"  ✓ {item['name']} - 正确识别为独立分部")
            else:
                print(f"  ✗ {item['name']} - 被错误合并")

if __name__ == '__main__':
    test_import()
